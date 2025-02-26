import requests
from bs4 import BeautifulSoup
import json
import os
import sys
import datetime
import time
import math
import copy
from logging.handlers import TimedRotatingFileHandler
from QtStockPriceMainWindow import Ui_MainWindow  # 導入轉換後的 UI 類
import logging
from QtStockCapitalIncreaseEditDialog import Ui_Dialog as Ui_StockCapitalIncreaseDialog
from QtStockTradingEditDialog import Ui_Dialog as Ui_StockTradingDialog
from QtStockRegularTradingEditDialog import Ui_Dialog as Ui_StockRegularTradingDialog
from QtStockDividendEditDialog import Ui_Dialog as Ui_StockDividendDialog
from QtStockCapitalReductionEditDialog import Ui_Dialog as Ui_StockCapitalReductionDialog
from QtCashTransferEditDialog import Ui_Dialog as Ui_CashTransferDialog
from QtDuplicateOptionDialog import Ui_Dialog as Ui_DuplicateOptionDialog
from QtSaveCheckDialog import Ui_Dialog as Ui_SaveCheckDialog
from PySide6.QtWidgets import QApplication, QMainWindow, QDialog, QButtonGroup, QMessageBox, QStyledItemDelegate, QFileDialog, QHeaderView, QVBoxLayout, QHBoxLayout, \
                              QLabel, QLineEdit, QDialogButtonBox, QTabBar, QWidget, QTableView, QComboBox, QPushButton, QSizePolicy, QSpacerItem, QCheckBox, QDoubleSpinBox, \
                              QProgressBar, QTabWidget
from PySide6.QtGui import QStandardItemModel, QStandardItem, QIcon, QBrush
from PySide6.QtCore import Qt, QModelIndex, QRect, QSignalBlocker, QSize, QThread, QObject, Signal
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from enum import Enum, IntEnum
from decimal import Decimal
from scipy.optimize import newton

#有考慮當沖交易稅減半
#有考慮 ETF 交易稅減少
#有考慮法人交易不需扣補充保費

#打包指令
# cd D:\_2.code\PythonStockPrice   
# pyinstaller --hidden-import "babel.numbers" --add-data "resources;./resources" --onefile --noconsole StockPriceMainWindow.py
# pyinstaller --hidden-import "babel.numbers" --add-data "resources;./resources" --onefile --console StockPriceMainWindow.py
# pyinstaller --hidden-import "babel.numbers" --add-data "resources;./resources" --noconsole StockPriceMainWindow.py

# 要把.ui檔變成.py
# cd D:\_2.code\PythonStockPrice
# pyside6-uic QtStockPriceMainWindowTemplate.ui -o QtStockPriceMainWindowTemplate.py
# pyside6-uic QtStockPriceMainWindow.ui -o QtStockPriceMainWindow.py
# pyside6-uic QtStockTradingEditDialog.ui -o QtStockTradingEditDialog.py
# pyside6-uic QtStockRegularTradingEditDialog.ui -o QtStockRegularTradingEditDialog.py
# pyside6-uic QtStockDividendEditDialog.ui -o QtStockDividendEditDialog.py
# pyside6-uic QtStockCapitalReductionEditDialog.ui -o QtStockCapitalReductionEditDialog.py
# pyside6-uic QtStockCapitalIncreaseEditDialog.ui -o QtStockCapitalIncreaseEditDialog.py
# pyside6-uic QtDuplicateOptionDialog.ui -o QtDuplicateOptionDialog.py
# pyside6-uic QtCashTransferEditDialog.ui -o QtCashTransferEditDialog.py
# pyside6-uic QtSaveCheckDialog.ui -o QtSaveCheckDialog.py

# 以下兩個網站都可以下載"上市"ETF的股利
# https://www.twse.com.tw/zh/products/securities/etf/products/div.html
# https://www.twse.com.tw/zh/ETFortune/dividendList

# 以下這個網站可以下載"上櫃"ETF的股利
# https://www.tpex.org.tw/zh-tw/announce/market/ex/cal.html

# 靜態掃描
# pylint --disable=all --enable=E1120,E1121 StockPriceMainWindow.py 只顯示參數數量錯誤
# pylint -E StockPriceMainWindow.py 只顯示錯誤級別的資訊


# region 設定 錯誤資訊 logging
# 設定日誌
logger = logging.getLogger()
logger.setLevel(logging.ERROR)

current_date = datetime.datetime.today().strftime("%Y-%m-%d")
log_filename = f"log_{current_date}.txt"
# 日誌檔案處理器
file_handler = TimedRotatingFileHandler(
    filename = log_filename,  # 基本文件名
    when = "midnight",  # 依據每天午夜分割日誌
    interval = 1,  # 每 1 天生成一個新檔案
    backupCount = 7,  # 保留最近 7 天的日誌檔案
    encoding = "utf-8"  # 確保支援 UTF-8 編碼
)
file_handler.setLevel(logging.ERROR)
file_formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
file_handler.setFormatter(file_formatter)

# 終端處理器
console_handler = logging.StreamHandler()
console_handler.setLevel(logging.ERROR)
console_formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
console_handler.setFormatter(console_formatter)

# 添加處理器到 logger
logger.addHandler(file_handler)
logger.addHandler(console_handler)

# 自定義未處理例外的鉤子
def handle_exception(exc_type, exc_value, exc_traceback):
    if issubclass(exc_type, KeyboardInterrupt):
        # 如果是鍵盤中斷，保持預設行為
        sys.__excepthook__(exc_type, exc_value, exc_traceback)
        return
    # 記錄例外
    logging.error("Uncaught exception", exc_info=(exc_type, exc_value, exc_traceback))
# 將自定義的鉤子設定為全局的未處理例外處理器
sys.excepthook = handle_exception
# endregion

g_user_dir = os.path.expanduser("~")  #開發模式跟打包模式下都是C:\Users\foxki
g_exe_dir = os.path.dirname(__file__) #開發模式下是D:\_2.code\PythonStockPrice #打包模式後是C:\Users\foxki\AppData\Local\Temp\_MEI60962 最後那個資料夾是暫時性的隨機名稱
g_exe2_dir = os.path.dirname( sys.executable ) #開發模式下是C:\Users\foxki\AppData\Local\Programs\Python\Python312 #打包模式後是:D:\_2.code\PythonStockPrice\dist
g_abs_dir = os.path.dirname( os.path.abspath(__file__) ) #開發模式下是D:\_2.code\PythonStockPrice #打包模式後是C:\Users\foxki\AppData\Local\Temp\_MEI60962 最後那個資料夾是暫時性的隨機名稱
print( "g_user_dir :" + g_user_dir ) #開發模式下是C:\Users\foxki
print( "g_exe_dir :" + g_exe_dir ) #開發模式下是D:\_2.code\PythonStockPrice #打包模式後是C:\Users\foxki\AppData\Local\Temp\_MEI60962 最後那個資料夾是暫時性的隨機名稱
print( "g_exe2_dir :" + g_exe2_dir ) #開發模式下是C:\Users\foxki\AppData\Local\Programs\Python\Python312 #打包模式後是:D:\_2.code\PythonStockPrice\dist
print( "g_abs_dir :" + g_abs_dir ) #開發模式下是D:\_2.code\PythonStockPrice #打包模式後是C:\Users\foxki\AppData\Local\Temp\_MEI60962 最後那個資料夾是暫時性的隨機名稱


if getattr( sys, 'frozen', False ):
    # PyInstaller 打包後執行時
    g_exe_root_dir = os.path.dirname(__file__) #C:\Users\foxki\AppData\Local\Temp\_MEI60962
    g_data_dir = os.path.join( g_user_dir, "AppData", "Local", "FoxInfo" ) #C:\Users\foxki\AppData\Local\FoxInfo
else:
    # VSCode執行 Python 腳本時
    g_exe_root_dir = os.path.dirname( os.path.abspath(__file__) )
    g_data_dir = g_exe_root_dir

window_icon_file_path = os.path.join( g_exe_root_dir, 'resources\\FoxInfo.png' ) 
edit_icon_file_path = os.path.join( g_exe_root_dir, 'resources\\Edit.svg' ) 
edit_icon = QIcon( edit_icon_file_path ) 
delete_icon_file_path = os.path.join( g_exe_root_dir, 'resources\\Delete.svg' ) 
delete_icon = QIcon( delete_icon_file_path ) 
export_icon_file_path = os.path.join( g_exe_root_dir, 'resources\\Export.svg' ) 
export_icon = QIcon( export_icon_file_path )
check_icon_file_path = os.path.join( g_exe_root_dir, 'resources\\CheckOn.svg' ) 
check_icon = QIcon( check_icon_file_path )
uncheck_icon_file_path = os.path.join( g_exe_root_dir, 'resources\\CheckOff.svg' ) 
uncheck_icon = QIcon( uncheck_icon_file_path )
down_icon_file_path = os.path.join( g_exe_root_dir, 'resources\\MoveDown.svg' ) 
down_icon = QIcon( down_icon_file_path )
up_icon_file_path = os.path.join( g_exe_root_dir, 'resources\\MoveUp.svg' ) 
up_icon = QIcon( up_icon_file_path )
styles_css_path = os.path.join( g_exe_root_dir, 'resources\\styles.css' ) 


class CenterIconDelegate( QStyledItemDelegate ):
    def paint( self, painter, option, index ):
        # 获取单元格数据
        icon = index.data( Qt.DecorationRole )  # 获取图标
        
        # 如果有图标
        if icon:
            rect = option.rect  # 单元格的绘制区域
            size = icon.actualSize( rect.size() ) * 0.7  # 图标实际尺寸
            
            # 计算居中位置
            x = rect.x() + ( rect.width() - size.width() ) // 2
            y = rect.y() + ( rect.height() - size.height() ) // 2
            target_rect = QRect( x, y, size.width(), size.height() )
            
            # 绘制图标
            icon.paint( painter, target_rect, Qt.AlignCenter )
        else:
            # 如果没有图标，使用默认绘制方法
            super().paint( painter, option, index )

class TradingType( IntEnum ):
    TEMPLATE = 0
    SELL = 1
    BUY = 2
    REGULAR_BUY = 3
    CAPITAL_INCREASE = 4
    DIVIDEND = 5
    CAPITAL_REDUCTION = 6

class TradingFeeType( Enum ):
    VARIABLE = 0
    CONSTANT = 1

class TradingData( Enum ):
    TRADING_DATE = 0
    TRADING_TYPE = 1 # 0:賣出, 1:買進, 2:股利, 3:減資
    TRADING_PRICE = 2
    TRADING_COUNT = 3
    TRADING_FEE_TYPE = 4 # 0:固定, 1:變動
    TRADING_FEE_DISCOUNT = 5
    TRADING_FEE_MINIMUM = 6
    TRADING_FEE_CONSTANT = 7
    STOCK_DIVIDEND_PER_SHARE = 8
    CASH_DIVIDEND_PER_SHARE = 9
    CAPITAL_REDUCTION_PER_SHARE = 10
    USE_AUTO_DIVIDEND_DATA = 11
    DAYING_TRADING = 12
    SORTED_INDEX_NON_SAVE = 13 #不會記錄
    TRADING_VALUE_NON_SAVE = 14 #不會記錄
    TRADING_FEE_NON_SAVE = 15 #不會記錄
    TRADING_TAX_NON_SAVE = 16 #不會記錄
    TRADING_COST_NON_SAVE = 17 #不會記錄
    STOCK_DIVIDEND_GAIN_NON_SAVE = 18 #不會記錄
    CASH_DIVIDEND_GAIN_NON_SAVE = 19 #不會記錄
    EXTRA_INSURANCE_FEE_NON_SAVE = 20 #不會記錄
    ACCUMULATED_COST_NON_SAVE = 21 #不會記錄
    ACCUMULATED_INVENTORY_NON_SAVE = 22 #不會記錄
    AVERAGE_COST_NON_SAVE = 23 #不會記錄
    IS_AUTO_DIVIDEND_DATA_NON_SAVE = 24 #不會記錄
    ALL_STOCK_DIVIDEND_GAIN_NON_SAVE = 25 #不會記錄
    ALL_CASH_DIVIDEND_GAIN_NON_SAVE = 26 #不會記錄
    IS_REALLY_DAYING_TRADING_NON_SAVE = 27 #不會記錄

class TradingCost( Enum ):
    TRADING_VALUE = 0
    TRADING_FEE = 1
    TRADING_TAX = 2
    TRADING_TOTAL_COST = 3

class TransferType( Enum ):
    TRANSFER_IN = 0
    TRANSFER_OUT = 1

class TransferData( Enum ):
    TRANSFER_DATE = 0
    TRANSFER_TYPE = 1 # 0:入金, 1:出金
    TRANSFER_VALUE = 2
    SORTED_INDEX_NON_SAVE = 3 #不會記錄
    TOTAL_VALUE_NON_SAVE = 4 #不會記錄

class Utility():
    def compute_cost( e_trading_type, f_trading_price, n_trading_count, f_trading_fee_discount, b_etf, b_daying_trading, b_bond ):
        f_trading_price = Decimal( str( f_trading_price ) )#原本10.45 * 100000 = 1044999.999999999 然後取 int 就變成1044999，所以改用Decimal
        n_trading_count = Decimal( str( n_trading_count ) )
        f_trading_fee_discount = Decimal( str( f_trading_fee_discount ) )
        dict_result = {}
        if e_trading_type == TradingType.BUY or e_trading_type == TradingType.SELL:
            n_trading_value = int( f_trading_price * n_trading_count )
            n_trading_fee = int( n_trading_value * Decimal( '0.001425' ) * f_trading_fee_discount )
            
            if n_trading_value != 0:
                if n_trading_count % 1000 == 0:
                    n_trading_fee = max( 20, n_trading_fee )
                else:#零股交易
                    n_trading_fee = max( 1, n_trading_fee )
            
            if e_trading_type == TradingType.SELL:
                if b_etf:
                    if b_bond:
                         n_trading_tax = 0
                    else:
                        n_trading_tax = int( n_trading_value * Decimal( '0.001' ) )
                elif b_daying_trading:
                    n_trading_tax = int( n_trading_value * Decimal( '0.0015' ) )
                else:
                    n_trading_tax = int( n_trading_value * Decimal( '0.003' ) )
            else:
                n_trading_tax = 0

            dict_result[ TradingCost.TRADING_VALUE ] = n_trading_value
            dict_result[ TradingCost.TRADING_FEE ] = n_trading_fee
            dict_result[ TradingCost.TRADING_TAX ] = n_trading_tax
            if e_trading_type == TradingType.BUY:
                dict_result[ TradingCost.TRADING_TOTAL_COST ] = n_trading_value + n_trading_fee + n_trading_tax
            else:
                dict_result[ TradingCost.TRADING_TOTAL_COST ] = n_trading_value - n_trading_fee - n_trading_tax
        else:   
            dict_result[ TradingCost.TRADING_VALUE ] = 0
            dict_result[ TradingCost.TRADING_FEE ] = 0
            dict_result[ TradingCost.TRADING_TAX ] = 0
            dict_result[ TradingCost.TRADING_TOTAL_COST ] = 0
        return dict_result

    def generate_trading_data( str_trading_date,            #交易日期
                               e_trading_type,              #交易種類
                               f_trading_price,             #交易價格
                               n_trading_count,             #交易股數
                               e_trading_fee_type,          #手續費種類
                               f_trading_fee_discount,      #手續費折扣
                               n_trading_fee_minimum,       #手續費最低金額
                               n_trading_fee_constant,      #手續費固定金額
                               f_stock_dividend_per_share,  #每股股票股利
                               f_cash_dividend_per_share,   #每股現金股利
                               f_capital_reduction_per_share, #每股減資金額
                               b_daying_trading  ):         #是否為當沖交易
        dict_trading_data = {}
        dict_trading_data[ TradingData.TRADING_DATE ] = str_trading_date
        dict_trading_data[ TradingData.TRADING_TYPE ] = e_trading_type
        dict_trading_data[ TradingData.TRADING_PRICE ] = f_trading_price
        dict_trading_data[ TradingData.TRADING_COUNT ] = n_trading_count
        dict_trading_data[ TradingData.TRADING_FEE_TYPE ] = e_trading_fee_type
        dict_trading_data[ TradingData.TRADING_FEE_DISCOUNT ] = f_trading_fee_discount
        dict_trading_data[ TradingData.TRADING_FEE_MINIMUM ] = n_trading_fee_minimum
        dict_trading_data[ TradingData.TRADING_FEE_CONSTANT ] = n_trading_fee_constant
        dict_trading_data[ TradingData.STOCK_DIVIDEND_PER_SHARE ] = f_stock_dividend_per_share
        dict_trading_data[ TradingData.CASH_DIVIDEND_PER_SHARE ] = f_cash_dividend_per_share
        dict_trading_data[ TradingData.CAPITAL_REDUCTION_PER_SHARE ] = f_capital_reduction_per_share
        dict_trading_data[ TradingData.DAYING_TRADING ] = b_daying_trading
        return dict_trading_data

    def xirr(cash_flows, dates):
        """
        計算 XIRR (年化報酬率)
        :param cash_flows: list of cash flows (現金流金額，正數為收入，負數為支出)
        :param dates: list of datetime objects (對應的日期)
        :return: 年化報酬率
        """
        def xnpv(rate):
            """
            計算 XNPV
            """
            return sum(cf / ((1 + rate) ** ((d - dates[0]).days / 365.0)) for cf, d in zip(cash_flows, dates))

        # 初始猜測的 XIRR 值
        try:
            return newton(lambda r: xnpv(r), 0.1)
        except RuntimeError:
            raise ValueError("無法找到解，請檢查輸入的現金流和日期。")

class EditTabTitleDialog( QDialog ):
    """一個小對話框用於編輯 Tab 標題"""
    def __init__( self, current_title, parent = None ):
        super().__init__( parent )
        self.setWindowTitle( "修改名稱" )
        self.resize( 300, 100 )

        self.layout = QVBoxLayout( self )

        self.label = QLabel("輸入名稱:")
        self.layout.addWidget( self.label )

        self.line_edit = QLineEdit( self )
        self.line_edit.setText( current_title )
        self.line_edit.selectAll()
        self.layout.addWidget( self.line_edit )

        self.horizontalLayout_4 = QHBoxLayout()
        self.qtOkPushButton = QPushButton( self )
        self.qtOkPushButton.setText( "確認" )
        self.qtCancelPushButton = QPushButton( self )
        self.qtCancelPushButton.setText( "取消" )

        self.horizontalLayout_4.addWidget(self.qtOkPushButton)
        self.horizontalLayout_4.addWidget(self.qtCancelPushButton)

        self.qtOkPushButton.clicked.connect( self.accept )
        self.qtCancelPushButton.clicked.connect( self.reject )

        self.layout.addLayout( self.horizontalLayout_4 )

    def get_new_title(self):
        """返回用戶輸入的標題"""
        return self.line_edit.text()

class ImportDataDuplicateOptionDialog( QDialog ):
    def __init__( self, parent = None ):
        super().__init__( parent )

        self.ui = Ui_DuplicateOptionDialog()
        self.ui.setupUi( self )
        window_icon = QIcon( window_icon_file_path ) 
        self.setWindowIcon( window_icon )
        self.ui.qtOkPushButton.clicked.connect( self.accept_data )
        self.ui.qtCancelPushButton.clicked.connect( self.cancel )
        self.b_overwrite = False

    def accept_data( self ):
        self.b_overwrite = self.ui.qtOverWriteRadioButton.isChecked()
        self.accept()
    
    def cancel( self ):
        self.reject()

class StockCapitalReductionEditDialog( QDialog ):
    def __init__( self, str_stock_number, str_stock_name, parent = None ):
        super().__init__( parent )

        self.ui = Ui_StockCapitalReductionDialog()
        self.ui.setupUi( self )
        window_icon = QIcon( window_icon_file_path ) 
        self.setWindowIcon( window_icon )

        self.ui.qtStockNumberLabel.setText( str_stock_number )
        self.ui.qtStockNameLabel.setText( str_stock_name )
        obj_current_date = datetime.datetime.today()
        self.ui.qtDateEdit.setDate( obj_current_date.date() )
        self.ui.qtDateEdit.setCalendarPopup( True )
        self.ui.qtOkPushButton.clicked.connect( self.accept_data )
        self.ui.qtCancelPushButton.clicked.connect( self.cancel )
        self.dict_trading_data = {}

    def setup_trading_date( self, str_date ):
        self.ui.qtDateEdit.setDate( datetime.datetime.strptime( str_date, "%Y-%m-%d" ).date() )

    def setup_stock_capital_reduction( self, f_stock_capital_reduction_per_share ):
        self.ui.qtCapitalReductionDoubleSpinBox.setValue( f_stock_capital_reduction_per_share )

    def accept_data( self ):
        f_stock_capital_reduction_per_share = self.ui.qtCapitalReductionDoubleSpinBox.value()
        if f_stock_capital_reduction_per_share != 0:

            self.dict_trading_data = Utility.generate_trading_data( self.ui.qtDateEdit.date().toString( "yyyy-MM-dd" ), #交易日期
                                                                    TradingType.CAPITAL_REDUCTION,                      #交易種類
                                                                    0,                                                  #交易價格                         
                                                                    0,                                                  #交易股數
                                                                    TradingFeeType.VARIABLE,                            #手續費種類
                                                                    1,                                                  #手續費折扣
                                                                    0,                                                  #手續費最低金額
                                                                    0,                                                  #手續費固定金額
                                                                    0,                                                  #每股股票股利
                                                                    0,                                                  #每股現金股利
                                                                    f_stock_capital_reduction_per_share,                #每股減資金額
                                                                    False )                                             #是否為當沖交易
            self.accept()
        else:
            self.reject()
    
    def cancel( self ):
        self.reject()

class CashTransferEditDialog( QDialog ):
    def __init__( self, str_account_name, parent = None ):
        super().__init__( parent )

        self.ui = Ui_CashTransferDialog()
        self.ui.setupUi( self )
        window_icon = QIcon( window_icon_file_path ) 
        self.setWindowIcon( window_icon )

        self.ui.qtAccountNameLabel.setText( str_account_name )
        obj_current_date = datetime.datetime.today()
        self.ui.qtDateEdit.setDate( obj_current_date.date() )
        self.ui.qtDateEdit.setCalendarPopup( True )
        self.ui.qtOkPushButton.clicked.connect( self.accept_data )
        self.ui.qtCancelPushButton.clicked.connect( self.cancel )
        self.dict_cash_transfer_data = {}

    def setup_transfer_date( self, str_date ):
        self.ui.qtDateEdit.setDate( datetime.datetime.strptime( str_date, "%Y-%m-%d" ).date() )

    def setup_transfer_type( self, e_transfer_type ):
        if e_transfer_type == TransferType.TRANSFER_IN:
            self.ui.qtTransferInRadioButton.setChecked( True )
        else:
            self.ui.qtTransferOutRadioButton.setChecked( True )

    def setup_transfer_value( self, n_transfer_value ):
        self.ui.qtCashDividendSpinBox.setValue( n_transfer_value )

    def accept_data( self ):
        n_transfer_value = self.ui.qtCashDividendSpinBox.value()
        if n_transfer_value != 0:
            self.dict_cash_transfer_data[ TransferData.TRANSFER_DATE ] = self.ui.qtDateEdit.date().toString( "yyyy-MM-dd" )
            self.dict_cash_transfer_data[ TransferData.TRANSFER_TYPE ] = TransferType.TRANSFER_IN if self.ui.qtTransferInRadioButton.isChecked() else TransferType.TRANSFER_OUT
            self.dict_cash_transfer_data[ TransferData.TRANSFER_VALUE ] = n_transfer_value
    
            self.accept()
        else:
            self.reject()
    
    def cancel( self ):
        self.reject()

class StockDividendEditDialog( QDialog ):
    def __init__( self, str_stock_number, str_stock_name, parent = None ):
        super().__init__( parent )

        self.ui = Ui_StockDividendDialog()
        self.ui.setupUi( self )
        window_icon = QIcon( window_icon_file_path ) 
        self.setWindowIcon( window_icon )

        self.ui.qtStockNumberLabel.setText( str_stock_number )
        self.ui.qtStockNameLabel.setText( str_stock_name )
        obj_current_date = datetime.datetime.today()
        self.ui.qtDateEdit.setDate( obj_current_date.date() )
        self.ui.qtDateEdit.setCalendarPopup( True )
        self.ui.qtOkPushButton.clicked.connect( self.accept_data )
        self.ui.qtCancelPushButton.clicked.connect( self.cancel )
        self.dict_trading_data = {}

    def setup_trading_date( self, str_date ):
        self.ui.qtDateEdit.setDate( datetime.datetime.strptime( str_date, "%Y-%m-%d" ).date() )

    def setup_stock_dividend( self, f_stock_dividend_per_share ):
        self.ui.qtStockDividendDoubleSpinBox.setValue( f_stock_dividend_per_share )

    def setup_cash_dividend( self, f_cash_dividend_per_share ):
        self.ui.qtCashDividendDoubleSpinBox.setValue( f_cash_dividend_per_share )

    def accept_data( self ):
        f_stock_dividend_per_share = self.ui.qtStockDividendDoubleSpinBox.value()
        f_cash_dividend_per_share = self.ui.qtCashDividendDoubleSpinBox.value()
        if f_stock_dividend_per_share != 0 or f_cash_dividend_per_share != 0:

            self.dict_trading_data = Utility.generate_trading_data( self.ui.qtDateEdit.date().toString( "yyyy-MM-dd" ), #交易日期
                                                                    TradingType.DIVIDEND,                               #交易種類
                                                                    0,                                                  #交易價格                         
                                                                    0,                                                  #交易股數
                                                                    TradingFeeType.VARIABLE,                            #手續費種類
                                                                    1,                                                  #手續費折扣
                                                                    0,                                                  #手續費最低金額
                                                                    0,                                                  #手續費固定金額
                                                                    f_stock_dividend_per_share,                         #每股股票股利
                                                                    f_cash_dividend_per_share,                          #每股現金股利
                                                                    0,                                                  #每股減資金額
                                                                    False )                                             #是否為當沖交易
            self.accept()
        else:
            self.reject()
    
    def cancel( self ):
        self.reject()

class StockTradingEditDialog( QDialog ):
    def __init__( self, str_stock_number, str_stock_name, b_etf, b_discount, f_discount_value, parent = None ):
        super().__init__( parent )

        self.ui = Ui_StockTradingDialog()
        self.ui.setupUi( self )
        
        window_icon = QIcon( window_icon_file_path ) 
        self.setWindowIcon( window_icon )

        self.ui.qtStockNumberLabel.setText( str_stock_number )
        self.ui.qtStockNameLabel.setText( str_stock_name )
        obj_current_date = datetime.datetime.today()
        self.ui.qtDateEdit.setDate( obj_current_date.date() )
        self.ui.qtDateEdit.setCalendarPopup( True )
        self.ui.qtDiscountCheckBox.setChecked( b_discount )
        self.ui.qtDiscountRateDoubleSpinBox.setValue( f_discount_value * 10 )
        self.ui.qtDiscountRateDoubleSpinBox.setEnabled( b_discount )

        self.ui.qtDiscountCheckBox.stateChanged.connect( self.on_discount_check_box_state_changed )
        self.ui.qtDiscountRateDoubleSpinBox.valueChanged.connect( self.compute_cost )
        self.ui.qtBuyRadioButton.toggled.connect( self.compute_cost )
        self.ui.qtSellRadioButton.toggled.connect( self.compute_cost )
        self.ui.qtCommonTradeRadioButton.toggled.connect( self.on_trading_type_changed )
        self.ui.qtOddTradeRadioButton.toggled.connect( self.on_trading_type_changed )
        self.ui.qtPriceDoubleSpinBox.valueChanged.connect( self.compute_cost )
        self.ui.qtCommonTradeCountSpinBox.valueChanged.connect( self.compute_cost )
        self.ui.qtOddTradeCountSpinBox.valueChanged.connect( self.compute_cost )
        self.ui.qtOkPushButton.clicked.connect( self.accept_data )
        self.ui.qtCancelPushButton.clicked.connect( self.cancel )
        self.b_etf = b_etf
        self.str_stock_name = str_stock_name
        # self.load_stylesheet("style.css")
        self.dict_trading_data = {}

    def load_stylesheet( self, file_path ):
        try:
            with open(file_path, "r", encoding="utf-8") as file:  # 指定 UTF-8 編碼
                stylesheet = file.read()
                self.setStyleSheet(stylesheet)
        except FileNotFoundError:
            print(f"CSS 檔案 {file_path} 找不到")
        except Exception as e:
            print(f"讀取 CSS 檔案時發生錯誤: {e}")

    def on_discount_check_box_state_changed( self, state ):
        if state == 2:
            self.ui.qtDiscountRateDoubleSpinBox.setEnabled( True )
        else:
            self.ui.qtDiscountRateDoubleSpinBox.setEnabled( False )

        self.compute_cost()

    def setup_trading_date( self, str_date ):
        self.ui.qtDateEdit.setDate( datetime.datetime.strptime( str_date, "%Y-%m-%d" ).date() )

    def setup_trading_type( self, e_trading_type ):
        if e_trading_type == TradingType.BUY:
            self.ui.qtBuyRadioButton.setChecked( True )
        else:
            self.ui.qtSellRadioButton.setChecked( True )

    def setup_trading_discount( self, f_discount_value ):
        if f_discount_value != 1:
            self.ui.qtDiscountCheckBox.setChecked( True )
            self.ui.qtDiscountRateDoubleSpinBox.setValue( f_discount_value * 10 )
            self.ui.qtDiscountRateDoubleSpinBox.setEnabled( True )
        else:
            self.ui.qtDiscountCheckBox.setChecked( False )
            self.ui.qtDiscountRateDoubleSpinBox.setValue( 6 )
            self.ui.qtDiscountRateDoubleSpinBox.setEnabled( False )

    def setup_trading_price( self, f_price ):
        self.ui.qtPriceDoubleSpinBox.setValue( f_price )

    def setup_trading_count( self, f_count ):
        if f_count % 1000 == 0:
            self.ui.qtCommonTradeRadioButton.setChecked( True )
            self.ui.qtCommonTradeCountSpinBox.setValue( f_count / 1000 )
            self.ui.qtOddTradeCountSpinBox.setValue( 0 )
        else:
            self.ui.qtOddTradeRadioButton.setChecked( True )
            self.ui.qtCommonTradeCountSpinBox.setValue( 0 )
            self.ui.qtOddTradeCountSpinBox.setValue( f_count )

    def setup_daying_trading( self, b_daying_trading ):
        if b_daying_trading:
            self.ui.qtDayingTradingCheckBox.setChecked( True )
        else:
            self.ui.qtDayingTradingCheckBox.setChecked( False )

    def accept_data( self ):

        if float( self.ui.qtTotalCostLineEdit.text().replace( ',', '' ) ) != 0:
            
            self.dict_trading_data = Utility.generate_trading_data( self.ui.qtDateEdit.date().toString( "yyyy-MM-dd" ), #交易日期
                                                                    self.get_trading_type(),                            #交易種類
                                                                    self.ui.qtPriceDoubleSpinBox.value(),               #交易價格
                                                                    self.get_trading_count(),                           #交易股數
                                                                    TradingFeeType.VARIABLE,                            #手續費種類
                                                                    self.get_trading_fee_discount(),                    #手續費折扣
                                                                    0,                                                  #手續費最低金額
                                                                    0,                                                  #手續費固定金額
                                                                    0,                                                  #每股股票股利
                                                                    0,                                                  #每股現金股利
                                                                    0,                                                  #每股減資金額
                                                                    self.ui.qtDayingTradingCheckBox.isChecked() )       #是否為當沖交易                                          
            self.accept()
        else:
            self.reject()
    
    def cancel( self ):
        self.reject()

    def on_trading_type_changed( self ):
        if self.ui.qtCommonTradeRadioButton.isChecked():
            self.ui.qtCommonTradeCountSpinBox.setEnabled( True )
            self.ui.qtOddTradeCountSpinBox.setEnabled( False )
        else:
            self.ui.qtCommonTradeCountSpinBox.setEnabled( False )
            self.ui.qtOddTradeCountSpinBox.setEnabled( True )

        self.compute_cost()

    def get_trading_type( self ):
        if self.ui.qtBuyRadioButton.isChecked():
            return TradingType.BUY
        else:
            return TradingType.SELL

    def get_trading_count( self ):
        if self.ui.qtCommonTradeRadioButton.isChecked():
            return self.ui.qtCommonTradeCountSpinBox.value() * 1000
        else:
            return self.ui.qtOddTradeCountSpinBox.value()
        
    def get_trading_fee_discount( self ):
        if self.ui.qtDiscountCheckBox.isChecked():
            return self.ui.qtDiscountRateDoubleSpinBox.value() / 10
        else:
            return 1

    def compute_cost( self ):
        e_trading_type = self.get_trading_type()
        f_trading_price = self.ui.qtPriceDoubleSpinBox.value()
        n_trading_count = self.get_trading_count()
        f_trading_fee_discount = self.get_trading_fee_discount() 
        b_bond = True if '債' in self.str_stock_name else False
        dict_result = Utility.compute_cost( e_trading_type, f_trading_price, n_trading_count, f_trading_fee_discount, self.b_etf, False, b_bond )

        if e_trading_type == TradingType.BUY:
            self.ui.qtTradingValueLineEdit.setText( format( dict_result[ TradingCost.TRADING_VALUE ], ',' ) )
            self.ui.qtTotalCostLineEdit.setText( format( dict_result[ TradingCost.TRADING_TOTAL_COST ], ',' ) )
        elif e_trading_type == TradingType.SELL:
            self.ui.qtTradingValueLineEdit.setText( format( -dict_result[ TradingCost.TRADING_VALUE ], ',' ) )
            self.ui.qtTotalCostLineEdit.setText( format( -dict_result[ TradingCost.TRADING_TOTAL_COST ], ',' ) )
        self.ui.qtFeeLineEdit.setText( format( dict_result[ TradingCost.TRADING_FEE ], ',' ) )
        self.ui.qtTaxLineEdit.setText( format( dict_result[ TradingCost.TRADING_TAX ], ',' ) )

class StockRegularTradingEditDialog( QDialog ):
    def __init__( self, str_stock_number, str_stock_name, e_trading_fee_type, b_discount, f_discount_value, n_trading_fee_minimum, n_trading_fee_constant, parent = None ):
        super().__init__( parent )

        self.ui = Ui_StockRegularTradingDialog()
        self.ui.setupUi( self )
        
        window_icon = QIcon( window_icon_file_path ) 
        self.setWindowIcon( window_icon )

        self.ui.qtStockNumberLabel.setText( str_stock_number )
        self.ui.qtStockNameLabel.setText( str_stock_name )
        obj_current_date = datetime.datetime.today()
        self.ui.qtDateEdit.setDate( obj_current_date.date() )
        self.ui.qtDateEdit.setCalendarPopup( True )


        self.setup_trading_fee_type( e_trading_fee_type )
        self.ui.qtDiscountRateDoubleSpinBox.setEnabled( b_discount )
        self.ui.qtDiscountCheckBox.setChecked( b_discount )
        self.ui.qtDiscountRateDoubleSpinBox.setValue( f_discount_value * 10 )

        self.setup_trading_fee_minimum( n_trading_fee_minimum )
        self.setup_trading_fee_constant( n_trading_fee_constant )

        button_group_1 = QButtonGroup( self )
        button_group_1.addButton( self.ui.qtVariableFeeRadioButton )
        button_group_1.addButton( self.ui.qtConstantFeeRadioButton )

        self.ui.qtVariableFeeRadioButton.toggled.connect( self.update_ui )
        self.ui.qtConstantFeeRadioButton.toggled.connect( self.update_ui )
        self.ui.qtDiscountCheckBox.stateChanged.connect( self.update_ui )
        self.ui.qtDiscountRateDoubleSpinBox.valueChanged.connect( self.compute_cost )
        self.ui.qtTradingFeeMinimumSpinBox.valueChanged.connect( self.compute_cost )
        self.ui.qtTradingFeeConstantSpinBox.valueChanged.connect( self.compute_cost )
        self.ui.qtPriceDoubleSpinBox.valueChanged.connect( self.compute_cost )
        self.ui.qtOddTradeCountSpinBox.valueChanged.connect( self.compute_cost )
        self.ui.qtOkPushButton.clicked.connect( self.accept_data )
        self.ui.qtCancelPushButton.clicked.connect( self.cancel )
        # self.load_stylesheet("style.css")
        self.dict_trading_data = {}
        self.update_ui()

    def load_stylesheet( self, file_path ):
        try:
            with open(file_path, "r", encoding="utf-8") as file:  # 指定 UTF-8 編碼
                stylesheet = file.read()
                self.setStyleSheet(stylesheet)
        except FileNotFoundError:
            print(f"CSS 檔案 {file_path} 找不到")
        except Exception as e:
            print(f"讀取 CSS 檔案時發生錯誤: {e}")

    def update_ui( self ):
        with ( QSignalBlocker( self.ui.qtVariableFeeRadioButton ),
               QSignalBlocker( self.ui.qtConstantFeeRadioButton ),
               QSignalBlocker( self.ui.qtDiscountCheckBox ),
               QSignalBlocker( self.ui.qtDiscountRateDoubleSpinBox ),
               QSignalBlocker( self.ui.qtTradingFeeMinimumSpinBox ),
               QSignalBlocker( self.ui.qtTradingFeeConstantSpinBox ) ):

            if self.ui.qtVariableFeeRadioButton.isChecked():
                self.ui.qtDiscountCheckBox.setEnabled( True )

                if self.ui.qtDiscountCheckBox.isChecked():
                    self.ui.qtDiscountRateDoubleSpinBox.setEnabled( True )
                    self.ui.label_10.setEnabled( True )
                else:
                    self.ui.qtDiscountRateDoubleSpinBox.setEnabled( False )
                    self.ui.label_10.setEnabled( False )

                self.ui.label_8.setEnabled( True )
                self.ui.qtTradingFeeMinimumSpinBox.setEnabled( True )
                self.ui.label_11.setEnabled( True )

                self.ui.qtTradingFeeConstantSpinBox.setEnabled( False )
                self.ui.label_4.setEnabled( False )
            else:
                self.ui.qtDiscountCheckBox.setEnabled( False )
                self.ui.qtDiscountRateDoubleSpinBox.setEnabled( False )
                self.ui.label_10.setEnabled( False )
                self.ui.label_8.setEnabled( False )
                self.ui.qtTradingFeeMinimumSpinBox.setEnabled( False )
                self.ui.label_11.setEnabled( False )
                self.ui.qtTradingFeeConstantSpinBox.setEnabled( True )
                self.ui.label_4.setEnabled( True )

            self.compute_cost()

    def setup_trading_date( self, str_date ):
        self.ui.qtDateEdit.setDate( datetime.datetime.strptime( str_date, "%Y-%m-%d" ).date() )

    def setup_trading_fee_type( self, e_trading_fee_type ):
        if e_trading_fee_type == TradingFeeType.VARIABLE:
            self.ui.qtVariableFeeRadioButton.setChecked( True )
        else:
            self.ui.qtConstantFeeRadioButton.setChecked( True )

        self.update_ui()

    def setup_trading_discount( self, f_discount_value ):
        if f_discount_value != 1:
            self.ui.qtDiscountCheckBox.setChecked( True )
            self.ui.qtDiscountRateDoubleSpinBox.setValue( f_discount_value * 10 )
            self.ui.qtDiscountRateDoubleSpinBox.setEnabled( True )
        else:
            self.ui.qtDiscountCheckBox.setChecked( False )
            self.ui.qtDiscountRateDoubleSpinBox.setValue( 6 )
            self.ui.qtDiscountRateDoubleSpinBox.setEnabled( False )

        self.update_ui()

    def setup_trading_fee_minimum( self, n_trading_fee_minimum ):
        self.ui.qtTradingFeeMinimumSpinBox.setValue( n_trading_fee_minimum )

    def setup_trading_fee_constant( self, n_trading_fee_constant ):
        self.ui.qtTradingFeeConstantSpinBox.setValue( n_trading_fee_constant )

    def setup_trading_price( self, f_price ):
        self.ui.qtPriceDoubleSpinBox.setValue( f_price )

    def setup_trading_count( self, n_count ):
        self.ui.qtOddTradeCountSpinBox.setValue( n_count )

    def accept_data( self ):

        if float( self.ui.qtTotalCostLineEdit.text().replace( ',', '' ) ) != 0:
            self.dict_trading_data = Utility.generate_trading_data( self.ui.qtDateEdit.date().toString( "yyyy-MM-dd" ), #交易日期
                                                                    TradingType.REGULAR_BUY,                            #交易種類
                                                                    self.ui.qtPriceDoubleSpinBox.value(),               #交易價格
                                                                    self.ui.qtOddTradeCountSpinBox.value(),             #交易股數
                                                                    self.get_trading_fee_type(),                        #手續費種類
                                                                    self.get_trading_fee_discount(),                    #手續費折扣
                                                                    self.ui.qtTradingFeeMinimumSpinBox.value(),         #手續費最低金額
                                                                    self.ui.qtTradingFeeConstantSpinBox.value(),        #手續費固定金額
                                                                    0,                                                  #每股股票股利
                                                                    0,                                                  #每股現金股利
                                                                    0,                                                  #每股減資金額
                                                                    False )                                             #是否為當沖交易
            self.accept()
        else:
            self.reject()
    
    def cancel( self ):
        self.reject()

    def get_trading_fee_type( self ):
        if self.ui.qtVariableFeeRadioButton.isChecked():
            return TradingFeeType.VARIABLE
        else:
            return TradingFeeType.CONSTANT

    def get_trading_fee_discount( self ):
        if self.ui.qtDiscountCheckBox.isChecked():
            return self.ui.qtDiscountRateDoubleSpinBox.value() / 10
        else:
            return 1

    def compute_cost( self ):
        e_trading_fee_type = self.get_trading_fee_type()
        f_trading_price = Decimal( str( self.ui.qtPriceDoubleSpinBox.value() ) )
        n_trading_count = int( self.ui.qtOddTradeCountSpinBox.value() ) 
        f_trading_fee_discount = Decimal( str( self.get_trading_fee_discount() ) )
        n_trading_fee_minimum = int( self.ui.qtTradingFeeMinimumSpinBox.value() ) 
        n_trading_fee_constant = int( self.ui.qtTradingFeeConstantSpinBox.value() )
        
        n_trading_value = int( f_trading_price * n_trading_count )
        if e_trading_fee_type == TradingFeeType.VARIABLE:
            n_trading_fee = int( n_trading_value * Decimal( '0.001425' ) * f_trading_fee_discount )
            n_trading_fee = max( n_trading_fee_minimum, n_trading_fee )
        else:
            n_trading_fee = int( n_trading_fee_constant )

        if n_trading_value == 0:
            n_trading_fee = 0

        n_trading_total_cost = n_trading_value + n_trading_fee

        self.ui.qtTradingValueLineEdit.setText( format( n_trading_value, ',' ) )
        self.ui.qtFeeLineEdit.setText( format( n_trading_fee, ',' ) )
        self.ui.qtTotalCostLineEdit.setText( format( n_trading_total_cost, ',' ) )

class StockCapitalIncreaseEditDialog( QDialog ):
    def __init__( self, str_stock_number, str_stock_name, parent = None ):
        super().__init__( parent )

        self.ui = Ui_StockCapitalIncreaseDialog()
        self.ui.setupUi( self )
        window_icon = QIcon( window_icon_file_path ) 
        self.setWindowIcon( window_icon )

        self.ui.qtStockNumberLabel.setText( str_stock_number )
        self.ui.qtStockNameLabel.setText( str_stock_name )
        obj_current_date = datetime.datetime.today()
        self.ui.qtDateEdit.setDate( obj_current_date.date() )
        self.ui.qtDateEdit.setCalendarPopup( True )

        self.ui.qtPriceDoubleSpinBox.valueChanged.connect( self.compute_cost )
        self.ui.qtOddTradeCountSpinBox.valueChanged.connect( self.compute_cost )
        self.ui.qtOkPushButton.clicked.connect( self.accept_data )
        self.ui.qtCancelPushButton.clicked.connect( self.cancel )
        self.dict_trading_data = {}

    def setup_trading_date( self, str_date ):
        self.ui.qtDateEdit.setDate( datetime.datetime.strptime( str_date, "%Y-%m-%d" ).date() )

    def setup_trading_price( self, f_price ):
        self.ui.qtPriceDoubleSpinBox.setValue( f_price )

    def setup_trading_count( self, f_count ):
        self.ui.qtOddTradeCountSpinBox.setValue( f_count )

    def accept_data( self ):
        if float( self.ui.qtTotalCostLineEdit.text().replace( ',', '' ) ) != 0:
            self.dict_trading_data = Utility.generate_trading_data( self.ui.qtDateEdit.date().toString( "yyyy-MM-dd" ), #交易日期
                                                                    TradingType.CAPITAL_INCREASE,                       #交易種類
                                                                    self.ui.qtPriceDoubleSpinBox.value(),               #交易價格
                                                                    self.ui.qtOddTradeCountSpinBox.value(),             #交易股數
                                                                    TradingFeeType.VARIABLE,                            #手續費種類
                                                                    1,                                                  #手續費折扣
                                                                    0,                                                  #手續費最低金額
                                                                    0,                                                  #手續費固定金額
                                                                    0,                                                  #每股股票股利
                                                                    0,                                                  #每股現金股利
                                                                    0,                                                  #每股減資金額
                                                                    False )                                             #是否為當沖交易
            self.accept()
        else:
            self.reject()
    
    def cancel( self ):
        self.reject()

    def get_trading_count( self ):
        return self.ui.qtOddTradeCountSpinBox.value()
        
    def compute_cost( self ):
        f_trading_price = self.ui.qtPriceDoubleSpinBox.value()
        n_trading_count = self.get_trading_count()
        self.ui.qtTotalCostLineEdit.setText( format( int( f_trading_price * n_trading_count ), ',' ) )

class SaveCheckDialog( QDialog ):
    def __init__( self, str_title = '', parent = None ):
        super().__init__( parent )

        self.ui = Ui_SaveCheckDialog()
        self.ui.setupUi( self )
        self.setWindowTitle( str_title )
        window_icon = QIcon( window_icon_file_path ) 
        self.setWindowIcon( window_icon )

        self.ui.qtSavePushButton.clicked.connect( self.save )
        self.ui.qtNoSavePushButton.clicked.connect( self.no_save )
        self.ui.qtAbortPushButton.clicked.connect( self.abort )
        self.n_return = 0

    def save( self ):
        self.n_return = 1
        self.accept()

    def no_save( self ):
        self.n_return = 2
        self.accept()
    
    def abort( self ):
        self.n_return = 3
        self.accept()

class Worker( QObject ):
    progress = Signal( int )  # Signal to emit progress updates
    finished = Signal( dict )  # Signal to emit the result when done

    def __init__( self, main_window ):
        super().__init__()
        self.main_window = main_window  # 傳入 MainWindow 的實例

    def run( self ):
        # Perform the time-consuming operation (e.g. loading stock data)
        self.main_window.initialize( False, self.update_progress )
        self.finished.emit({})  # Emit the result when done

    def update_progress( self, value ):
        self.progress.emit( value )  # Emit progress updates

class CustomSortModel( QStandardItemModel ):
    def sort(self, column, order):

        numeric_data = []
        for row in range(self.rowCount()):
            item = self.item(row, column)
            try:
                # 嘗試轉換為整數，去掉逗號
                value = float(item.text().replace(",", ""))
            except ( ValueError, AttributeError ):
                # 如果轉換失敗，跳過此項目
                value = float('-inf') if order.value == 1 else float('inf')
            numeric_data.append(( value, row ) )

        numeric_data.sort(
            key = lambda x: x[ 0 ],
            reverse = ( order.value == 1 )
        )

        # 保存整個行的 QStandardItem 物件
        data = [
            [self.item( row, col ).clone() for col in range( self.columnCount() ) ]
            for row in range(self.rowCount())
        ]
        list_header = []
        for row in range( self.rowCount() ):
            header_text = self.verticalHeaderItem( row ).text()
            list_header.append( header_text )

        self.setRowCount( 0 )
        for row_data in numeric_data:
            self.appendRow( data[ row_data[1] ] )

        self.setVerticalHeaderLabels( list_header )

class MainWindow( QMainWindow ):
    def __init__( self, b_unit_test = False, 
                  str_initial_data_file = 'TradingData.json', 
                  str_UI_setting_file = 'UISetting.config', 
                  str_stock_number_file = 'StockNumber.txt',
                  str_stock_price_file = 'StockPrice.txt'  ):
        super( MainWindow, self ).__init__()
        self.ui = Ui_MainWindow()
        self.ui.setupUi( self )  # 設置 UI
        window_icon = QIcon( window_icon_file_path ) 
        self.setWindowIcon( window_icon )
        
        if not b_unit_test:
            self.progress_bar = QProgressBar( self )
            self.progress_bar.setGeometry( 400, 350, 300, 25 )  # Adjust position and size as needed
            self.progress_bar.setMaximum( 100 )
            self.progress_bar.setVisible( False )

        self.ui.qtTabWidget.currentChanged.connect( self.on_tab_current_changed )
        self.ui.qtTabWidget.tabBarDoubleClicked.connect( self.on_tab_widget_double_clicked )
        self.ui.qtTabWidget.tabCloseRequested.connect( self.on_tab_widget_close )
        self.ui.qtTabWidget.tabBar().tabMoved.connect( self.on_tab_moved )
        self.ui.qtTabWidget.tabBar().setTabButton( 0, QTabBar.RightSide, None )  # 隱藏最後一個 tab 的 close 按鈕

        delegate = CenterIconDelegate()

        self.per_stock_trading_data_model = QStandardItemModel( 0, 0 ) 
        self.per_stock_trading_data_model.setVerticalHeaderLabels( self.get_trading_data_header() )
        self.ui.qtTradingDataTableView.setModel( self.per_stock_trading_data_model )
        self.ui.qtTradingDataTableView.setItemDelegate( delegate )
        self.ui.qtTradingDataTableView.horizontalHeader().hide()
        self.ui.qtTradingDataTableView.verticalHeader().setSectionResizeMode( QHeaderView.Fixed )
        self.ui.qtTradingDataTableView.clicked.connect( lambda index: self.on_trading_data_table_item_clicked( index, self.per_stock_trading_data_model ) )
        for row in range( len( self.get_trading_data_header() ) ):
            if row == 10 or row == 11:
                self.ui.qtTradingDataTableView.setRowHeight( row, 40 )
            else:
                self.ui.qtTradingDataTableView.setRowHeight( row, 25 )

        self.ui.qtHideTradingDataTableToolButton.clicked.connect( self.on_hide_trading_data_table_tool_button_clicked )

        self.ui.qtAddTradingDataPushButton.clicked.connect( self.on_add_trading_data_push_button_clicked )
        self.ui.qtAddRegularTradingDataPushButton.clicked.connect( self.on_add_regular_trading_data_push_button_clicked )
        self.ui.qtAddDividendDataPushButton.clicked.connect( self.on_add_dividend_data_push_button_clicked )
        self.ui.qtAddLimitBuyingDataPushButton.clicked.connect( self.on_add_limit_buying_data_push_button_clicked )
        self.ui.qtAddCapitalReductionDataPushButton.clicked.connect( self.on_add_capital_reduction_data_push_button_clicked )
        self.ui.qtExportAllStockTradingDataPushButton.clicked.connect( self.on_export_all_to_excell_button_clicked )
        self.ui.qtExportSelectedStockTradingDataPushButton.clicked.connect( self.on_export_selected_to_excell_button_clicked )

        self.ui.qtHideTradingDataTableToolButton.setIcon( down_icon )

        self.ui.qtNewFileAction.setShortcut( "Ctrl+N" )
        self.ui.qtNewFileAction.triggered.connect( self.on_new_file_action_triggered )
        self.ui.qtOpenFileAction.setShortcut( "Ctrl+O" )
        self.ui.qtOpenFileAction.triggered.connect( self.on_open_file_action_triggered )
        self.ui.qtSaveAsAction.setShortcut( "Ctrl+A" )
        self.ui.qtSaveAsAction.triggered.connect( self.on_save_as_action_triggered )
        self.ui.qtSaveAction.setShortcut( "Ctrl+S" )
        self.ui.qtSaveAction.triggered.connect( self.on_save_action_triggered )
        self.ui.qtExportCurrentGroupAction.setShortcut( "Ctrl+E" )
        self.ui.qtExportCurrentGroupAction.triggered.connect( self.on_export_current_group_action_triggered )
        self.ui.qtImportFullAction.setShortcut( "Ctrl+U" )
        self.ui.qtImportFullAction.triggered.connect( self.on_import_full_action_triggered )
        self.ui.qtImportSingleStockAction.setShortcut( "Ctrl+I" )
        self.ui.qtImportSingleStockAction.triggered.connect( self.on_import_single_stock_action_triggered )

        self.ui.qtFromNewToOldAction.setChecked( True )
        self.ui.qtFromOldToNewAction.setChecked( False )
        self.ui.qtFromNewToOldAction.triggered.connect( self.on_trigger_from_new_to_old )
        self.ui.qtFromOldToNewAction.triggered.connect( self.on_trigger_from_old_to_new )

        self.ui.qtShowAllAction.setChecked( True )
        self.ui.qtShow10Action.setChecked( False )
        self.ui.qtShowAllAction.triggered.connect( self.on_trigger_show_all )
        self.ui.qtShow10Action.triggered.connect( self.on_trigger_show_10 )

        self.ui.qtUse1ShareUnitAction.setChecked( True )
        self.ui.qtUse1000ShareUnitAction.setChecked( False )
        self.ui.qtUse1ShareUnitAction.triggered.connect( self.on_trigger_use_1_share_unit )
        self.ui.qtUse1000ShareUnitAction.triggered.connect( self.on_trigger_use_1000_share_unit )

        self.ui.qtADYearAction.setChecked( True )
        self.ui.qtROCYearAction.setChecked( False )
        self.ui.qtADYearAction.triggered.connect( self.on_trigger_AD_year )
        self.ui.qtROCYearAction.triggered.connect( self.on_trigger_ROC_year )
        
        self.trading_data_json_file_path = os.path.join( g_data_dir, 'StockInventory', str_initial_data_file )
        self.UISetting_file_path = os.path.join( g_data_dir, 'StockInventory', str_UI_setting_file )
        self.stock_number_file_path = os.path.join( g_data_dir, 'StockInventory', str_stock_number_file )
        self.stock_price_file_path = os.path.join( g_data_dir, 'StockInventory', str_stock_price_file )

        self.list_stock_list_table_horizontal_header = [ '總成本', '庫存股數', '平均成本', ' 收盤價', '現值', '總手續費', '總交易稅', '損益', '股利所得', '自動帶入股利', '匯出', '刪除' ]
        self.pick_up_stock( None )
        self.dict_all_account_ui_state = {}
        self.dict_all_account_cash_transfer_data = {}
        self.dict_all_account_all_stock_trading_data = {}
        self.dict_all_account_all_stock_trading_data_INITIAL = {}
        self.list_stock_list_column_width = [ 85 ] * len( self.list_stock_list_table_horizontal_header )
        self.list_stock_list_column_width[ len( self.list_stock_list_table_horizontal_header ) - 2 ] = 40
        self.list_stock_list_column_width[ len( self.list_stock_list_table_horizontal_header ) - 1 ] = 40
        self.n_current_tab = 0
        self.n_tab_index = 0
        self.str_current_save_file_path = None
        self.load_stylesheet( styles_css_path )
        if b_unit_test:
            self.initialize( True, None )
            self.load_initialize_data()
        else:
            self.start_loading_stock_data()

    def download_all_required_data( self, str_date, update_progress_callback ):
        self.set_progress_value( update_progress_callback, 0 )
        self.download_all_company_stock_number( str_date )
        self.set_progress_value( update_progress_callback, 10 )
        self.download_day_stock_price( str_date )
        self.set_progress_value( update_progress_callback, 20 )
        self.download_general_company_all_yearly_dividend_data( 2010, str_date )
        self.set_progress_value( update_progress_callback, 50 )
        self.download_listed_etf_all_yearly_dividend_data( 2010, str_date )
        self.set_progress_value( update_progress_callback, 80 )
        self.download_OTC_etf_all_yearly_dividend_data( 2010, str_date )

    def initialize( self, b_unit_test, update_progress_callback ):
        self.setEnabled( False ) # 資料下載前先Disable整個視窗
        obj_current_date = datetime.datetime.today() - datetime.timedelta( days = 1 )
        str_date = obj_current_date.strftime('%Y%m%d')
        
        if not b_unit_test:
            self.download_all_required_data( str_date, update_progress_callback )
        
        self.dict_all_company_number_to_name_and_type = self.load_all_company_stock_number()
        self.dict_all_company_number_to_price_info = self.load_day_stock_price()
        self.dict_auto_stock_yearly_dividned = self.load_general_company_all_yearly_dividend_data( 2010, b_unit_test )
        self.dict_auto_stock_listed_etf_yearly_dividned = self.load_listed_etf_all_yearly_dividend_data( 2010, b_unit_test )
        self.dict_auto_stock_OTC_etf_yearly_dividned = self.load_OTC_etf_all_yearly_dividend_data( 2010, b_unit_test )

        for key_stock_number, value in self.dict_auto_stock_OTC_etf_yearly_dividned.items():
            # 00687B 這支 ETF 應該是櫃買中心的資料，但是出現在證交所(重點是資料不完整)，所以要排除
            # 為了避免有相同的問題，我們就直接把 self.dict_auto_stock_listed_etf_yearly_dividned 排除所有櫃買中心的 ETF
            if key_stock_number in self.dict_auto_stock_listed_etf_yearly_dividned:
                del self.dict_auto_stock_listed_etf_yearly_dividned[ key_stock_number ]

        # common_keys = set(self.dict_auto_stock_yearly_dividned.keys()) & set(self.dict_auto_stock_listed_etf_yearly_dividned.keys())
        self.dict_auto_stock_yearly_dividned.update( self.dict_auto_stock_listed_etf_yearly_dividned )
        for key, value in self.dict_auto_stock_OTC_etf_yearly_dividned.items():
            if key not in self.dict_auto_stock_yearly_dividned:
                self.dict_auto_stock_yearly_dividned[ key ] = value

        n_retry = 0
        while len( self.dict_all_company_number_to_price_info ) == 0:
            #因為我們要下載前一天的股價資訊，但有時候遇到前一天是假日，就要再往前，若是連續假日，就要一直往前直到可以下載
            obj_current_date = obj_current_date - datetime.timedelta( days = 1 )
            n_weekday = obj_current_date.weekday()
            if n_weekday == 5 or n_weekday == 6:
                continue
            str_date = obj_current_date.strftime('%Y%m%d')
            self.download_day_stock_price( str_date )
            self.dict_all_company_number_to_price_info = self.load_day_stock_price()
            n_retry += 1
            if n_retry > 30:
                break
        str_valid_month_date = obj_current_date.strftime('%m/%d')
        self.list_stock_list_table_horizontal_header[ 3 ] = str_valid_month_date + ' 收盤價'
        self.set_progress_value( update_progress_callback, 100 )
        # self.load_initialize_data()
        self.setEnabled( True ) # 資料下載完後就會Enable

    def set_progress_value( self, update_progress_callback, f_progress ):
        if update_progress_callback:
            update_progress_callback( f_progress )

    def start_loading_stock_data( self ):
        # Show the progress bar
        self.progress_bar.setVisible( True )

        # Create and start a QThread for the worker
        self.thread = QThread()
        self.worker = Worker(self)

        # Move the worker to the new thread
        self.worker.moveToThread( self.thread )

        # Connect signals and slots
        self.worker.progress.connect( self.update_progress )
        self.worker.finished.connect( self.on_loading_finished )
        self.thread.started.connect( self.worker.run )
        self.worker.finished.connect( self.thread.quit )

        # Start the thread
        self.thread.start()

    def update_progress( self, value ):
        self.progress_bar.setValue( value )

    def on_loading_finished( self ):
        self.progress_bar.setVisible( False )
        self.load_initialize_data()

    def load_stylesheet( self, file_path ):
        try:
            with open(file_path, "r", encoding="utf-8") as file:  # 指定 UTF-8 編碼
                stylesheet = file.read()
                self.setStyleSheet(stylesheet)
        except FileNotFoundError:
            print(f"CSS 檔案 {file_path} 找不到")
        except Exception as e:
            print(f"讀取 CSS 檔案時發生錯誤: {e}")

    def keyReleaseEvent(self, event):
        key = event.key()
        modifiers = QApplication.keyboardModifiers()

        if key == Qt.Key_T:
            if self.ui.qtAddTradingDataPushButton.isEnabled():
                self.on_add_trading_data_push_button_clicked()
        elif key == Qt.Key_E:
            if self.ui.qtAddRegularTradingDataPushButton.isEnabled():
                self.on_add_regular_trading_data_push_button_clicked()
        elif key == Qt.Key_D:
            if self.ui.qtAddDividendDataPushButton.isEnabled():
                self.on_add_dividend_data_push_button_clicked()
        elif key == Qt.Key_A:
            if self.ui.qtAddLimitBuyingDataPushButton.isEnabled():
                self.on_add_limit_buying_data_push_button_clicked()
        elif key == Qt.Key_R:
            if self.ui.qtAddCapitalReductionDataPushButton.isEnabled():
                self.on_add_capital_reduction_data_push_button_clicked()
        elif key == Qt.Key_Enter:
            qt_line_edit = self.ui.qtTabWidget.currentWidget().findChild( QLineEdit, "StockInputLineEdit" )
            if qt_line_edit and qt_line_edit.hasFocus():
                self.on_add_stock_push_button_clicked()

    def add_new_tab_and_table( self, str_tab_title = None ): 
        str_tab_name = f"TabIndex{ self.n_tab_index }"
        increased_tab = QWidget()
        increased_tab.setObjectName( str_tab_name )
               
        uiqt_vertical_layout_main = QVBoxLayout( increased_tab )
        uiqt_vertical_layout_main.setSpacing(0)
        uiqt_vertical_layout_main.setContentsMargins(-1, 0, -1, 0)

        uiqt_stock_inventory_and_cash_transfer_tab_widget = QTabWidget( increased_tab )
        uiqt_stock_inventory_and_cash_transfer_tab_widget.setObjectName(u"qtStockInventoryAndCashTransferTabWidget")
        uiqt_stock_inventory_and_cash_transfer_tab_widget.setTabPosition( QTabWidget.West )

        stock_inventory_tab = QWidget()
        stock_inventory_tab.setObjectName( str_tab_name + "StockInventoryTab" )
        cash_transfer_tab = QWidget()
        cash_transfer_tab.setObjectName( str_tab_name + "CashTransferTab" )
        uiqt_stock_inventory_and_cash_transfer_tab_widget.insertTab( 0, stock_inventory_tab, "庫存總覽" )
        uiqt_stock_inventory_and_cash_transfer_tab_widget.insertTab( 1, cash_transfer_tab, "銀行" )
        
        uiqt_vertical_layout_main.addWidget( uiqt_stock_inventory_and_cash_transfer_tab_widget )

        # stock inventory tab
        uiqt_stock_inventory_tab_vertical_layout = QVBoxLayout( stock_inventory_tab )
        uiqt_stock_inventory_tab_vertical_layout.setSpacing( 0 )
        uiqt_stock_inventory_tab_vertical_layout.setContentsMargins( -1, 0, -1, 0 )
        uiqt_horizontal_layout_1 = QHBoxLayout()
        uiqt_stock_input_line_edit = QLineEdit( stock_inventory_tab )
        sizePolicy = QSizePolicy( QSizePolicy.Policy.Maximum, QSizePolicy.Policy.Fixed )
        sizePolicy.setHorizontalStretch( 0 )
        sizePolicy.setVerticalStretch( 0 )
        sizePolicy.setHeightForWidth( uiqt_stock_input_line_edit.sizePolicy().hasHeightForWidth() )

        uiqt_stock_input_line_edit.setSizePolicy( sizePolicy )
        uiqt_stock_input_line_edit.setMinimumSize( QSize( 150, 0 ) )
        uiqt_stock_input_line_edit.setMaximumSize( QSize( 150, 16777215 ) )
        uiqt_stock_input_line_edit.setObjectName( "StockInputLineEdit" )
        uiqt_horizontal_layout_1.addWidget( uiqt_stock_input_line_edit )

        uiqt_horizontal_spacer_1_0 = QSpacerItem( 10, 20, QSizePolicy.Policy.Minimum, QSizePolicy.Policy.Minimum )
        uiqt_horizontal_layout_1.addItem( uiqt_horizontal_spacer_1_0 )

        uiqt_add_stock_push_button = QPushButton( stock_inventory_tab )
        uiqt_add_stock_push_button.setMinimumSize( QSize( 75, 0 ) )
        uiqt_add_stock_push_button.setMaximumSize( QSize( 75, 16777215 ) )
        uiqt_add_stock_push_button.setText( "新增股票" )
        uiqt_add_stock_push_button.setObjectName( "AddStockPushButton" )
        uiqt_horizontal_layout_1.addWidget( uiqt_add_stock_push_button )

        uiqt_horizontal_spacer_1_1 = QSpacerItem( 40, 20, QSizePolicy.Policy.Minimum, QSizePolicy.Policy.Minimum )
        uiqt_horizontal_layout_1.addItem( uiqt_horizontal_spacer_1_1 )

        uiqt_extra_insurance_fee_check_box = QCheckBox( stock_inventory_tab )
        uiqt_extra_insurance_fee_check_box.setText( "補充保費" )
        uiqt_extra_insurance_fee_check_box.setObjectName( "ExtraInsuranceFeeCheckBox" )
        uiqt_horizontal_layout_1.addWidget( uiqt_extra_insurance_fee_check_box )

        uiqt_horizontal_spacer_1_2 = QSpacerItem( 40, 20, QSizePolicy.Policy.Minimum, QSizePolicy.Policy.Minimum )
        uiqt_horizontal_layout_1.addItem( uiqt_horizontal_spacer_1_2 )

        uiqt_display_type_combobox = QComboBox( stock_inventory_tab )
        uiqt_display_type_combobox.addItem( "顯示所有交易" )
        uiqt_display_type_combobox.addItem( "僅顯示目前庫存" )
        uiqt_display_type_combobox.addItem( "僅顯示已無庫存" )
        uiqt_display_type_combobox.setObjectName( "DisplayTypeComboBox" )
        uiqt_horizontal_layout_1.addWidget( uiqt_display_type_combobox )

        uiqt_horizontal_spacer_1_3 = QSpacerItem( 40, 20, QSizePolicy.Policy.Minimum, QSizePolicy.Policy.Minimum )
        uiqt_horizontal_layout_1.addItem( uiqt_horizontal_spacer_1_3 )

        uiqt_total_inventory_label = QLabel( stock_inventory_tab )
        uiqt_total_inventory_label.setText( "總庫存: " )
        uiqt_total_inventory_label.setObjectName( "TotalInventoryLabel" )
        uiqt_total_inventory_value_label = QLabel( stock_inventory_tab )
        uiqt_total_inventory_value_label.setText( "" )
        uiqt_total_inventory_value_label.setObjectName( "TotalInventoryValueLabel" )
        
        uiqt_horizontal_layout_1.addWidget( uiqt_total_inventory_label )
        uiqt_horizontal_layout_1.addWidget( uiqt_total_inventory_value_label )

        uiqt_horizontal_spacer_1_4 = QSpacerItem( 40, 20, QSizePolicy.Policy.Minimum, QSizePolicy.Policy.Minimum )
        uiqt_horizontal_layout_1.addItem( uiqt_horizontal_spacer_1_4 )

        uiqt_total_profit_label = QLabel( stock_inventory_tab )
        uiqt_total_profit_label.setText( "總損益: " )
        uiqt_total_profit_label.setObjectName( "TotalProfitLabel" )
        uiqt_total_profit_value_label = QLabel( stock_inventory_tab )
        uiqt_total_profit_value_label.setText( "" )
        uiqt_total_profit_value_label.setObjectName( "TotalProfitValueLabel" )
        
        uiqt_horizontal_layout_1.addWidget( uiqt_total_profit_label )
        uiqt_horizontal_layout_1.addWidget( uiqt_total_profit_value_label )

        uiqt_horizontal_spacer_1_5 = QSpacerItem( 40, 20, QSizePolicy.Policy.Minimum, QSizePolicy.Policy.Minimum )
        uiqt_horizontal_layout_1.addItem( uiqt_horizontal_spacer_1_5 )

        uiqt_xirr_label = QLabel( stock_inventory_tab )
        uiqt_xirr_label.setText( "年化報酬率: " )
        uiqt_xirr_label.setObjectName( "XIRRLabel" )
        uiqt_xirr_value_label = QLabel( stock_inventory_tab )
        uiqt_xirr_value_label.setText( "" )
        uiqt_xirr_value_label.setObjectName( "XIRRValueLabel" )

        uiqt_horizontal_layout_1.addWidget( uiqt_xirr_label )
        uiqt_horizontal_layout_1.addWidget( uiqt_xirr_value_label )

        uiqt_horizontal_spacer_1_6 = QSpacerItem(40, 20, QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Minimum)
        uiqt_horizontal_layout_1.addItem( uiqt_horizontal_spacer_1_6 )

        uiqt_stock_inventory_tab_vertical_layout.addLayout( uiqt_horizontal_layout_1 )

        uiqt_horizontal_layout_2 = QHBoxLayout()
        uiqt_horizontal_layout_2.setSpacing( 0 )

        uiqt_stock_select_combo_box = QComboBox( stock_inventory_tab )
        uiqt_stock_select_combo_box.setMinimumSize( QSize( 200, 0 ) )
        uiqt_stock_select_combo_box.setObjectName( "StockSelectComboBox" )
        uiqt_horizontal_layout_2.addWidget( uiqt_stock_select_combo_box )

        uiqt_horizontal_spacer_2_1 = QSpacerItem( 40, 0, QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Minimum )
        uiqt_horizontal_layout_2.addItem( uiqt_horizontal_spacer_2_1 )

        uiqt_stock_inventory_tab_vertical_layout.addLayout( uiqt_horizontal_layout_2 )

        uiqt_horizontal_layout_3 = QHBoxLayout()

        uiqt_stock_list_table_view = QTableView( stock_inventory_tab )
        uiqt_stock_list_table_view.setMinimumSize( QSize( 0, 100 ) )
        uiqt_stock_list_table_view.setObjectName( "StockListTableView" )
        uiqt_horizontal_layout_3.addWidget( uiqt_stock_list_table_view )

        uiqt_stock_inventory_tab_vertical_layout.addLayout( uiqt_horizontal_layout_3 )

        delegate = CenterIconDelegate()
        stock_list_model = CustomSortModel( 0, 0 )
        stock_list_model.setHorizontalHeaderLabels( self.list_stock_list_table_horizontal_header )
        uiqt_stock_list_table_view.verticalHeader().setSectionsMovable( True )
        uiqt_stock_list_table_view.verticalHeader().sectionMoved.connect( self.on_stock_list_table_vertical_header_section_moved )
        uiqt_stock_list_table_view.verticalHeader().sectionClicked.connect( self.on_stock_list_table_vertical_section_clicked )
        uiqt_stock_list_table_view.verticalHeader().setSectionResizeMode( QHeaderView.Fixed )
        uiqt_stock_list_table_view.horizontalHeader().sectionResized.connect( self.on_stock_list_table_horizontal_section_resized )
        uiqt_stock_list_table_view.setSortingEnabled( True )
        uiqt_stock_list_table_view.setModel( stock_list_model )
        uiqt_stock_list_table_view.setItemDelegate( delegate )
        uiqt_stock_list_table_view.clicked.connect( lambda index: self.on_stock_list_table_item_clicked( index, stock_list_model ) )
        uiqt_stock_list_table_view.horizontalHeader().sortIndicatorChanged.connect( self.update_stock_list_vertical_header )

        uiqt_stock_input_line_edit.textChanged.connect( self.on_stock_input_text_changed ) 

        uiqt_stock_select_combo_box.setVisible( False )
        uiqt_stock_select_combo_box.activated.connect( self.on_stock_select_combo_box_current_index_changed )
        uiqt_stock_select_combo_box.setStyleSheet( "QComboBox { combobox-popup: 0; }" )
        uiqt_stock_select_combo_box.setMaxVisibleItems( 10 )

        uiqt_add_stock_push_button.clicked.connect( self.on_add_stock_push_button_clicked )
        uiqt_extra_insurance_fee_check_box.stateChanged.connect( self.on_extra_insurance_fee_check_box_state_changed )
        uiqt_display_type_combobox.activated.connect( self.on_display_type_combo_box_current_index_changed )
        
        # Cash Transfer Tab
        uiqt_cash_transfer_tab_vertical_layout = QVBoxLayout( cash_transfer_tab )
        uiqt_cash_transfer_tab_vertical_layout.setSpacing( 0 )
        uiqt_cash_transfer_tab_vertical_layout.setContentsMargins( -1, 0, -1, 0 )

        uiqt_vertical_spacer_2_0 = QSpacerItem(20, 5, QSizePolicy.Policy.Minimum, QSizePolicy.Policy.Minimum)
        uiqt_cash_transfer_tab_vertical_layout.addItem( uiqt_vertical_spacer_2_0 )

        uiqt_horizontal_layout_4 = QHBoxLayout()
        uiqt_horizontal_layout_4.setSpacing( 0 )

        uiqt_cash_transfer_table_view = QTableView( cash_transfer_tab )
        uiqt_cash_transfer_table_view.setMinimumSize( QSize( 0, 166 ) )
        uiqt_cash_transfer_table_view.setMaximumSize( QSize( 16777215, 166 ) )
        uiqt_cash_transfer_table_view.setObjectName( "CashTransferTableView" )
        uiqt_horizontal_layout_4.addWidget( uiqt_cash_transfer_table_view )

        uiqt_cash_transfer_tab_vertical_layout.addLayout( uiqt_horizontal_layout_4 )

        uiqt_horizontal_layout_5 = QHBoxLayout()
        uiqt_horizontal_layout_5.setSpacing( 0 )

        # uiqt_horizontal_spacer_5_1 = QSpacerItem(3, 20, QSizePolicy.Policy.Minimum, QSizePolicy.Policy.Minimum)
        # uiqt_horizontal_layout_5.addItem( uiqt_horizontal_spacer_5_1 )

        uiqt_add_cash_transfer_push_button = QPushButton( cash_transfer_tab )
        uiqt_add_cash_transfer_push_button.setMaximumSize( QSize( 16777215, 16777215 ) )
        uiqt_add_cash_transfer_push_button.setMinimumWidth( 200 )
        uiqt_add_cash_transfer_push_button.setText( "新增入金/出金資料" )
        uiqt_add_cash_transfer_push_button.setObjectName( "AddCashTransferPushButton" )
        uiqt_add_cash_transfer_push_button.clicked.connect( self.on_add_cash_transfer_push_button_clicked )
        uiqt_horizontal_layout_5.addWidget( uiqt_add_cash_transfer_push_button )

        uiqt_horizontal_spacer_5_2 = QSpacerItem(40, 20, QSizePolicy.Policy.MinimumExpanding, QSizePolicy.Policy.Minimum)
        uiqt_horizontal_layout_5.addItem( uiqt_horizontal_spacer_5_2 )

        uiqt_vertical_spacer_2_1 = QSpacerItem(20, 5, QSizePolicy.Policy.Minimum, QSizePolicy.Policy.Minimum)
        uiqt_cash_transfer_tab_vertical_layout.addItem( uiqt_vertical_spacer_2_1 )

        uiqt_cash_transfer_tab_vertical_layout.addLayout( uiqt_horizontal_layout_5 )
        uiqt_vertical_spacer_2_2 = QSpacerItem(20, 40, QSizePolicy.Policy.Minimum, QSizePolicy.Policy.Expanding)
        uiqt_cash_transfer_tab_vertical_layout.addItem( uiqt_vertical_spacer_2_2 )

        cash_transfer_model = QStandardItemModel( 0, 0 )
        cash_transfer_model.setVerticalHeaderLabels( self.get_cash_transfer_header() )
        uiqt_cash_transfer_table_view.verticalHeader().setSectionResizeMode( QHeaderView.Fixed )
        uiqt_cash_transfer_table_view.horizontalHeader().setSectionResizeMode( QHeaderView.Fixed )
        uiqt_cash_transfer_table_view.setModel( cash_transfer_model )
        uiqt_cash_transfer_table_view.setItemDelegate( delegate )
        uiqt_cash_transfer_table_view.horizontalHeader().hide()
        uiqt_cash_transfer_table_view.clicked.connect( lambda index: self.on_cash_transfer_table_item_clicked( index, cash_transfer_model ) )


        if not str_tab_title:
            str_tab_title = "新群組"
        with QSignalBlocker( self.ui.qtTabWidget ):
            n_ori_count = self.ui.qtTabWidget.count()
            self.ui.qtTabWidget.insertTab( n_ori_count - 1, increased_tab, str_tab_title )
            self.ui.qtTabWidget.setCurrentIndex( n_ori_count - 1 )
        self.n_tab_index += 1
        return str_tab_name

    def on_tab_current_changed( self, index ): 
        n_tab_count = self.ui.qtTabWidget.count()
        if index == n_tab_count - 1:
            self.ui.qtTabWidget.setCurrentIndex( self.n_current_tab )
        else:
            self.n_current_tab = index
            self.pick_up_stock( None )
            self.refresh_stock_list_table()
            self.refresh_transfer_data_table()

            self.clear_per_stock_trading_table()
            self.update_button_enable_disable_status()

    def on_tab_widget_double_clicked( self, index ): 
        n_tab_count = self.ui.qtTabWidget.count()
        if index == n_tab_count - 1:
            str_tab_name = self.add_new_tab_and_table()
            self.dict_all_account_all_stock_trading_data[ str_tab_name ] = {}
            self.dict_all_account_ui_state[ str_tab_name ] = { "discount_checkbox": True, "discount_value": 0.6, "insurance_checkbox": False, "trading_fee_type": TradingFeeType.VARIABLE, "trading_fee_minimum": 1, "trading_fee_constant": 1 }
            self.dict_all_account_cash_transfer_data[ str_tab_name ] = []
        else:
            current_title = self.ui.qtTabWidget.tabText( index )
            dialog = EditTabTitleDialog( current_title, self )
            if dialog.exec() == QDialog.Accepted:
                new_title = dialog.get_new_title()
                self.ui.qtTabWidget.setTabText( index, new_title )
        self.auto_save_trading_data()

    def on_tab_widget_close( self, index ):

        str_tab_title = self.ui.qtTabWidget.tabText( index )
        tab_widget = self.ui.qtTabWidget.widget( index )
        result = self.show_warning_message_box_with_ok_cancel_button( "警告", f"確定要刪掉『{str_tab_title}』的所有資料嗎?\n建議先從「檔案」=>「匯出目前帳號資料」，匯出檔案做備份" )
        if result:
            str_tab_widget_name = tab_widget.objectName()
            value = self.dict_all_account_all_stock_trading_data.pop( str_tab_widget_name, None )
            self.ui.qtTabWidget.removeTab( index )
            self.clear_per_stock_trading_table()
            self.update_button_enable_disable_status()
            self.auto_save_trading_data()

    def on_tab_moved( self, n_from, n_to ): 
        str_tab_name = self.ui.qtTabWidget.widget( self.ui.qtTabWidget.count() - 1 ).objectName()
        tab_bar = self.ui.qtTabWidget.tabBar()
        if str_tab_name != 'tab_add':
            tab_bar.moveTab( n_to, n_from )
        else:
            self.auto_save_trading_data()

    def on_stock_input_text_changed( self ): 
        qt_combo_box = self.ui.qtTabWidget.currentWidget().findChild( QComboBox, "StockSelectComboBox" )
        qt_line_edit = self.ui.qtTabWidget.currentWidget().findChild( QLineEdit, "StockInputLineEdit" )

        with QSignalBlocker( qt_combo_box ), QSignalBlocker( qt_line_edit ):
            qt_combo_box.clear()
            str_stock_input = qt_line_edit.text()
            if len( str_stock_input ) == 0:
                qt_combo_box.setVisible( False )
                return
            qt_combo_box.setVisible( True )

            for stock_number, list_stock_name_and_type in self.dict_all_company_number_to_name_and_type.items():
                str_stock_name = list_stock_name_and_type[ 0 ]
                if str_stock_input in stock_number or str_stock_input in str_stock_name:
                    qt_combo_box.addItem( f"{stock_number} {str_stock_name}" )
            # self.ui.qtStockSelectComboBox.showPopup() #showPopup的話，focus會被搶走

            qt_line_edit.setFocus()

    def on_stock_select_combo_box_current_index_changed( self, index ): 
        qt_combo_box = self.ui.qtTabWidget.currentWidget().findChild( QComboBox, "StockSelectComboBox" )
        qt_line_edit = self.ui.qtTabWidget.currentWidget().findChild( QLineEdit, "StockInputLineEdit" )

        str_stock_input = qt_combo_box.currentText()
        qt_line_edit.setText( str_stock_input )
        qt_combo_box.setVisible( False )
        qt_line_edit.setFocus()

    def on_add_stock_push_button_clicked( self ): 
        str_tab_widget_name = self.ui.qtTabWidget.currentWidget().objectName()
        dict_per_account_all_stock_trading_data = self.dict_all_account_all_stock_trading_data[ str_tab_widget_name ]
        qt_line_edit = self.ui.qtTabWidget.currentWidget().findChild( QLineEdit, "StockInputLineEdit" )
        display_type_combobox = self.ui.qtTabWidget.currentWidget().findChild( QComboBox, "DisplayTypeComboBox")

        str_stock_input = qt_line_edit.text()
        qt_line_edit.clear()
        str_first_four_chars = str_stock_input.split(" ")[0]
        if str_first_four_chars not in self.dict_all_company_number_to_name_and_type:
            b_find = False
            for stock_number, list_stock_name_and_type in self.dict_all_company_number_to_name_and_type.items():
                str_stock_name = list_stock_name_and_type[ 0 ]
                if str_first_four_chars == str_stock_name:
                    str_first_four_chars = stock_number
                    b_find = True
                    break
            if not b_find:
                QMessageBox.warning( self, "警告", "輸入的股票代碼不存在", QMessageBox.Ok )
                return
        
        if str_first_four_chars not in dict_per_account_all_stock_trading_data:
            dict_trading_data = Utility.generate_trading_data( "0001-01-01",              #交易日期
                                                               TradingType.TEMPLATE,      #交易種類
                                                               0,                         #交易價格
                                                               0,                         #交易股數
                                                               TradingFeeType.VARIABLE,   #手續費種類
                                                               1,                         #手續費折扣
                                                               0,                         #手續費最低金額
                                                               0,                         #手續費固定金額
                                                               0,                         #每股股票股利
                                                               0,                         #每股現金股利
                                                               0,                         #每股減資金額
                                                               False )                    #是否為當沖交易
            dict_trading_data[ TradingData.USE_AUTO_DIVIDEND_DATA ] = True
            dict_per_account_all_stock_trading_data[ str_first_four_chars ] = [ dict_trading_data ]
            sorted_list = self.process_single_trading_data( str_tab_widget_name, str_first_four_chars )
            display_type_combobox.setCurrentIndex( 0 )
            self.refresh_stock_list_table()
            self.auto_save_trading_data()
        else:
            if display_type_combobox.currentIndex() != 0:
                display_type_combobox.setCurrentIndex( 0 )
                self.refresh_stock_list_table()

        table_view = self.ui.qtTabWidget.currentWidget().findChild( QTableView, "StockListTableView" )
        if table_view:
            table_model = table_view.model()
            for row in range( table_model.rowCount() ):
                header_text = table_model.verticalHeaderItem( row ).text()
                str_stock_number = header_text.split(" ")[0]
                if str_stock_number == str_first_four_chars:
                    index = table_model.index( row, 0 )
                    table_view.scrollTo( index, QTableView.PositionAtTop )
                    self.on_stock_list_table_item_clicked( index, table_model )
                    break

    def on_extra_insurance_fee_check_box_state_changed( self, state ): 
        str_tab_widget_name = self.ui.qtTabWidget.currentWidget().objectName()
        if state == 2:
            self.dict_all_account_ui_state[ str_tab_widget_name ][ "insurance_checkbox"] = True
        else:
            self.dict_all_account_ui_state[ str_tab_widget_name ][ "insurance_checkbox"] = False
        
        dict_per_company_trading_data = self.dict_all_account_all_stock_trading_data[ str_tab_widget_name ]
        for key_stock_name, value_list_trading_data in dict_per_company_trading_data.items():
            self.process_single_trading_data( str_tab_widget_name, key_stock_name )

        self.refresh_stock_list_table()
        if self.str_picked_stock_number != None:
            self.refresh_trading_data_table( dict_per_company_trading_data[ self.str_picked_stock_number ] )
        self.auto_save_trading_data()

    def on_display_type_combo_box_current_index_changed( self, index ):
        self.pick_up_stock( None )
        self.refresh_stock_list_table()
        self.clear_per_stock_trading_table()
        self.update_button_enable_disable_status()

    def on_stock_list_table_vertical_header_section_moved( self, n_logical_index, n_old_visual_index, n_new_visual_index ): 
        str_tab_widget_name = self.ui.qtTabWidget.currentWidget().objectName()

        #取得所有股票代碼
        dict_per_account_all_stock_trading_data = self.dict_all_account_all_stock_trading_data[ str_tab_widget_name ].copy()
        list_stock_number_all = []
        for index_row,( key_stock_number, value ) in enumerate( dict_per_account_all_stock_trading_data.items() ):
            list_stock_number_all.append( key_stock_number )
        
        
        table_view = self.ui.qtTabWidget.currentWidget().findChild( QTableView, "StockListTableView" )
        list_stock_number_visible = []
        list_visible_stock_index_of_all = []
        if table_view:
            table_model = table_view.model()

            for row in range( table_model.rowCount() ):
                header_text = table_model.verticalHeaderItem( row ).text()
                str_stock_number = header_text.split(" ")[0]
                if str_stock_number in list_stock_number_all:
                    list_stock_number_visible.append( str_stock_number )
                    list_visible_stock_index_of_all.append( list_stock_number_all.index( str_stock_number ) )


        element = list_stock_number_visible.pop( n_old_visual_index )
        list_stock_number_visible.insert( n_new_visual_index, element )


        for idx, val in zip( list_visible_stock_index_of_all, list_stock_number_visible ):
            list_stock_number_all[ idx ] = val

        dict_all_stock_trading_data_new = {}
        for index_row, str_stock_number in enumerate( list_stock_number_all ):
            dict_all_stock_trading_data_new[ str_stock_number ] = dict_per_account_all_stock_trading_data[ str_stock_number ]


        self.dict_all_account_all_stock_trading_data[ str_tab_widget_name ] = dict_all_stock_trading_data_new
        self.refresh_stock_list_table()
        self.auto_save_trading_data()

    def on_stock_list_table_vertical_section_clicked( self, n_logical_index ): 
        table_view = self.ui.qtTabWidget.currentWidget().findChild( QTableView, "StockListTableView" )
        if table_view:
            table_model = table_view.model()

            str_tab_widget_name = self.ui.qtTabWidget.currentWidget().objectName()
            dict_per_account_all_stock_trading_data = self.dict_all_account_all_stock_trading_data[ str_tab_widget_name ]
            header_text = table_model.verticalHeaderItem( n_logical_index ).text()
            str_stock_number = header_text.split(" ")[0]

            if str_stock_number in dict_per_account_all_stock_trading_data:
                if str_stock_number != self.str_picked_stock_number:
                    self.pick_up_stock( str_stock_number )
                    list_trading_data = dict_per_account_all_stock_trading_data[ str_stock_number ]
                    self.refresh_trading_data_table( list_trading_data )
            self.update_button_enable_disable_status()

    def on_stock_list_table_horizontal_section_resized( self, n_logical_index, n_old_size, n_new_size ): 
        self.list_stock_list_column_width[ n_logical_index ] = n_new_size
        self.save_share_UI_state()

    def on_stock_list_table_item_clicked( self, index: QModelIndex, table_model ):
        str_tab_widget_name = self.ui.qtTabWidget.currentWidget().objectName()
        dict_per_account_all_stock_trading_data = self.dict_all_account_all_stock_trading_data[ str_tab_widget_name ]
        item = table_model.itemFromIndex( index )
        if item is not None:
            n_column = index.column()  # 獲取列索引
            n_row = index.row()  # 獲取行索引
            header_text = table_model.verticalHeaderItem( index.row() ).text()
            str_stock_number = header_text.split(" ")[0]
            
            if n_column == len( self.list_stock_list_table_horizontal_header ) - 3:#自動帶入股利按鈕
                list_trading_data = dict_per_account_all_stock_trading_data[ str_stock_number ]
                list_trading_data[ 0 ][ TradingData.USE_AUTO_DIVIDEND_DATA ] = not list_trading_data[ 0 ][ TradingData.USE_AUTO_DIVIDEND_DATA ]
                self.pick_up_stock( str_stock_number )
                sorted_list = self.process_single_trading_data( str_tab_widget_name, str_stock_number )
                self.refresh_stock_list_table()
                self.refresh_trading_data_table( sorted_list )
                self.auto_save_trading_data()
            elif n_column == len( self.list_stock_list_table_horizontal_header ) - 2:#匯出按鈕
                file_path = self.open_save_json_file_dialog()
                if file_path:
                    list_save_tab_widget = [ self.ui.qtTabWidget.currentIndex() ]
                    self.manual_save_trading_data( list_save_tab_widget, file_path, str_stock_number )
                
                if str_stock_number != self.str_picked_stock_number:
                    self.pick_up_stock( str_stock_number )
                    list_trading_data = dict_per_account_all_stock_trading_data[ str_stock_number ]
                    self.refresh_trading_data_table( list_trading_data )
            elif n_column == len( self.list_stock_list_table_horizontal_header ) - 1:#刪除按鈕
                result = self.show_warning_message_box_with_ok_cancel_button( "警告", f"確定要刪掉『{header_text}』的所有資料嗎?" )
                if result:
                    del dict_per_account_all_stock_trading_data[ str_stock_number ]
                    self.pick_up_stock( None )
                    self.refresh_stock_list_table()
                    self.clear_per_stock_trading_table()
                    self.auto_save_trading_data()
            elif str_stock_number in dict_per_account_all_stock_trading_data:
                if str_stock_number != self.str_picked_stock_number:
                    self.pick_up_stock( str_stock_number )
                    list_trading_data = dict_per_account_all_stock_trading_data[ str_stock_number ]
                    self.refresh_trading_data_table( list_trading_data )

        self.update_button_enable_disable_status()

    def update_stock_list_vertical_header( self ):
        str_tab_widget_name = self.ui.qtTabWidget.currentWidget().objectName()
        dict_per_account_all_stock_trading_data = self.dict_all_account_all_stock_trading_data[ str_tab_widget_name ].copy()
        list_stock_number_all = []
        for index_row,( key_stock_number, value ) in enumerate( dict_per_account_all_stock_trading_data.items() ):
            list_stock_number_all.append( key_stock_number )

        table_view = self.ui.qtTabWidget.currentWidget().findChild( QTableView, "StockListTableView" )
        list_stock_number_visible = []
        list_visible_stock_index_of_all = []
        if table_view:
            table_model = table_view.model()
            dict_stock_name = {}
            for row in range( table_model.rowCount() ):
                header_text = table_model.verticalHeaderItem( row ).text()
                str_stock_number = header_text.split(" ")[0]
                dict_stock_name[ str_stock_number ] = header_text
                if str_stock_number in list_stock_number_all:
                    list_stock_number_visible.append( str_stock_number )


            for row in range( table_model.rowCount() ):
                index = table_model.index( row, 0 )
                hidden_data = table_model.data( index, Qt.UserRole )
                str_stock_number_in_hidden_data = str( hidden_data )
                table_model.setHeaderData( row, Qt.Vertical, dict_stock_name[ str_stock_number_in_hidden_data ] )
                if str_stock_number_in_hidden_data in list_stock_number_all:
                    list_visible_stock_index_of_all.append( list_stock_number_all.index( str_stock_number_in_hidden_data ) )

            for idx, val in zip( list_visible_stock_index_of_all, list_stock_number_visible ):
                list_stock_number_all[ idx ] = val


            dict_all_stock_trading_data_new = {}
            for index_row, str_stock_number in enumerate( list_stock_number_all ):
                dict_all_stock_trading_data_new[ str_stock_number ] = dict_per_account_all_stock_trading_data[ str_stock_number ]

            self.dict_all_account_all_stock_trading_data[ str_tab_widget_name ] = dict_all_stock_trading_data_new
            self.auto_save_trading_data()

    def on_add_cash_transfer_push_button_clicked( self ):
        str_tab_widget_name = self.ui.qtTabWidget.currentWidget().objectName()
        if str_tab_widget_name == 'tab_add':
            return

        list_per_account_all_cash_trasfer_data = self.dict_all_account_cash_transfer_data[ str_tab_widget_name ]

        n_current_index = self.ui.qtTabWidget.currentIndex()
        current_title = self.ui.qtTabWidget.tabText( n_current_index )

        dialog = CashTransferEditDialog( current_title, self )

        if dialog.exec():
            dict_transfer_data = dialog.dict_cash_transfer_data
            list_per_account_all_cash_trasfer_data.append( dict_transfer_data )
            self.process_single_transfer_data( str_tab_widget_name )
            self.refresh_transfer_data_table()
            self.auto_save_trading_data()

    def on_cash_transfer_table_item_clicked( self, index: QModelIndex, table_model ):
        item = table_model.itemFromIndex( index )
        if item is not None:
            n_row = index.row()  # 獲取行索引
            str_tab_widget_name = self.ui.qtTabWidget.currentWidget().objectName()
            list_per_account_cash_transfer_data = self.dict_all_account_cash_transfer_data[ str_tab_widget_name ]

            hidden_data = table_model.data( index, Qt.UserRole )
            n_findindex = -1
            for index, dict_selected_data in enumerate( list_per_account_cash_transfer_data ):
                if dict_selected_data[ TransferData.SORTED_INDEX_NON_SAVE ] == hidden_data:
                    n_findindex = index
                    break
            if n_findindex == -1:
                return
            dict_selected_data = list_per_account_cash_transfer_data[ n_findindex ]

            if n_row == len( self.get_cash_transfer_header() ) - 2: #編輯

                n_current_index = self.ui.qtTabWidget.currentIndex()
                current_title = self.ui.qtTabWidget.tabText( n_current_index )
                dialog = CashTransferEditDialog( current_title, self )
                dialog.setup_transfer_date( dict_selected_data[ TransferData.TRANSFER_DATE ] )
                dialog.setup_transfer_type( dict_selected_data[ TransferData.TRANSFER_TYPE ] )
                dialog.setup_transfer_value( dict_selected_data[ TransferData.TRANSFER_VALUE ] )

                if dialog.exec():
                    dict_transfer_data = dialog.dict_cash_transfer_data
                    list_per_account_cash_transfer_data[ n_findindex ] = dict_transfer_data
                    self.process_single_transfer_data( str_tab_widget_name )
                    self.refresh_transfer_data_table()
                    self.auto_save_trading_data()

            elif n_row == len( self.get_cash_transfer_header() ) - 1: #刪除
                result = self.show_warning_message_box_with_ok_cancel_button( "警告", f"確定要刪掉這筆匯款資料嗎?" )
                if result:
                    del list_per_account_cash_transfer_data[ n_findindex ]
                    self.process_single_transfer_data( str_tab_widget_name )
                    self.refresh_transfer_data_table()
                    self.auto_save_trading_data()

    def on_trigger_from_new_to_old( self ):
        with ( QSignalBlocker( self.ui.qtFromNewToOldAction ),
               QSignalBlocker( self.ui.qtFromOldToNewAction ) ):
                self.ui.qtFromNewToOldAction.setChecked( True )
                self.ui.qtFromOldToNewAction.setChecked( False )
                self.on_change_display_mode()

    def on_trigger_from_old_to_new( self ):
        with ( QSignalBlocker( self.ui.qtFromNewToOldAction ),
               QSignalBlocker( self.ui.qtFromOldToNewAction ) ):
                self.ui.qtFromNewToOldAction.setChecked( False )
                self.ui.qtFromOldToNewAction.setChecked( True )
                self.on_change_display_mode()

    def on_trigger_show_all( self ):
        with ( QSignalBlocker( self.ui.qtShowAllAction ),
               QSignalBlocker( self.ui.qtShow10Action ) ):
                self.ui.qtShowAllAction.setChecked( True )
                self.ui.qtShow10Action.setChecked( False )
                self.on_change_display_mode()

    def on_trigger_show_10( self ):
        with ( QSignalBlocker( self.ui.qtShowAllAction ),
               QSignalBlocker( self.ui.qtShow10Action ) ):
                self.ui.qtShowAllAction.setChecked( False )
                self.ui.qtShow10Action.setChecked( True )
                self.on_change_display_mode()

    def on_trigger_use_1_share_unit( self ):
        with ( QSignalBlocker( self.ui.qtUse1ShareUnitAction ),
               QSignalBlocker( self.ui.qtUse1000ShareUnitAction ) ):
                self.ui.qtUse1ShareUnitAction.setChecked( True )
                self.ui.qtUse1000ShareUnitAction.setChecked( False )
                self.on_change_display_mode()
                self.refresh_stock_list_table()

    def on_trigger_use_1000_share_unit( self ):
        with ( QSignalBlocker( self.ui.qtUse1ShareUnitAction ),
               QSignalBlocker( self.ui.qtUse1000ShareUnitAction ) ):
                self.ui.qtUse1ShareUnitAction.setChecked( False )
                self.ui.qtUse1000ShareUnitAction.setChecked( True )
                self.on_change_display_mode()
                self.refresh_stock_list_table()

    def on_trigger_AD_year( self ):
        with ( QSignalBlocker( self.ui.qtADYearAction ),
               QSignalBlocker( self.ui.qtROCYearAction ) ):
                self.ui.qtADYearAction.setChecked( True )
                self.ui.qtROCYearAction.setChecked( False )
                self.on_change_display_mode()

    def on_trigger_ROC_year( self ):
        with ( QSignalBlocker( self.ui.qtADYearAction ),
               QSignalBlocker( self.ui.qtROCYearAction ) ):
                self.ui.qtADYearAction.setChecked( False )
                self.ui.qtROCYearAction.setChecked( True )
                self.on_change_display_mode()

    def on_change_display_mode( self ): 
        if self.str_picked_stock_number != None:
            str_tab_widget_name = self.ui.qtTabWidget.currentWidget().objectName()
            dict_per_account_all_stock_trading_data = self.dict_all_account_all_stock_trading_data[ str_tab_widget_name ]
            self.refresh_trading_data_table( dict_per_account_all_stock_trading_data[ self.str_picked_stock_number ] )

        self.refresh_transfer_data_table()
        self.save_share_UI_state()

    def on_hide_trading_data_table_tool_button_clicked( self ):
        if self.ui.qtTradingDataTableView.isVisible():
            self.ui.qtHideTradingDataTableToolButton.setIcon( up_icon )
            self.ui.qtTradingDataTableView.setHidden( True )
        else:
            self.ui.qtHideTradingDataTableToolButton.setIcon( down_icon )
            self.ui.qtTradingDataTableView.setHidden( False )

    def on_add_trading_data_push_button_clicked( self ): 
        if self.str_picked_stock_number is None:
            return

        str_tab_widget_name = self.ui.qtTabWidget.currentWidget().objectName()
        dict_per_account_all_stock_trading_data = self.dict_all_account_all_stock_trading_data[ str_tab_widget_name ]
        b_discount = self.dict_all_account_ui_state[ str_tab_widget_name ][ "discount_checkbox"]
        f_discount_value = self.dict_all_account_ui_state[ str_tab_widget_name ][ "discount_value"]

        str_stock_number = self.str_picked_stock_number
        list_stock_name_and_type = self.dict_all_company_number_to_name_and_type[ str_stock_number ]
        str_stock_name = list_stock_name_and_type[ 0 ]
        str_b_etf = self.dict_all_company_number_to_name_and_type[ str_stock_number ][ 1 ]
        b_etf = True if str_b_etf == "True" else False
        dialog = StockTradingEditDialog( str_stock_number, str_stock_name, b_etf, b_discount, f_discount_value, self )

        if dialog.exec():
            dict_trading_data = dialog.dict_trading_data
            if dict_trading_data[ TradingData.TRADING_FEE_DISCOUNT ] == 1:
                self.dict_all_account_ui_state[ str_tab_widget_name ][ "discount_checkbox"] = False
                self.dict_all_account_ui_state[ str_tab_widget_name ][ "discount_value"] = 0.6
            else:
                self.dict_all_account_ui_state[ str_tab_widget_name ][ "discount_checkbox"] = True
                self.dict_all_account_ui_state[ str_tab_widget_name ][ "discount_value"] = dict_trading_data[ TradingData.TRADING_FEE_DISCOUNT ]


            dict_per_account_all_stock_trading_data[ str_stock_number ].append( dict_trading_data )
            sorted_list = self.process_single_trading_data( str_tab_widget_name, str_stock_number )
            self.refresh_stock_list_table()
            self.refresh_trading_data_table( sorted_list )
            self.auto_save_trading_data()

    def on_add_regular_trading_data_push_button_clicked( self ):
        if self.str_picked_stock_number is None:
            return

        str_tab_widget_name = self.ui.qtTabWidget.currentWidget().objectName()
        dict_per_account_all_stock_trading_data = self.dict_all_account_all_stock_trading_data[ str_tab_widget_name ]
        b_discount = self.dict_all_account_ui_state[ str_tab_widget_name ][ "discount_checkbox"]
        f_discount_value = self.dict_all_account_ui_state[ str_tab_widget_name ][ "discount_value"]
        e_trading_fee_type = self.dict_all_account_ui_state[ str_tab_widget_name ][ "trading_fee_type"]
        n_trading_fee_minimum = self.dict_all_account_ui_state[ str_tab_widget_name ][ "trading_fee_minimum"]
        n_trading_fee_constant = self.dict_all_account_ui_state[ str_tab_widget_name ][ "trading_fee_constant"]

        str_stock_number = self.str_picked_stock_number
        list_stock_name_and_type = self.dict_all_company_number_to_name_and_type[ str_stock_number ]
        str_stock_name = list_stock_name_and_type[ 0 ]
        dialog = StockRegularTradingEditDialog( str_stock_number, str_stock_name, e_trading_fee_type, b_discount, f_discount_value, n_trading_fee_minimum, n_trading_fee_constant, self )

        if dialog.exec():
            dict_trading_data = dialog.dict_trading_data
            if dict_trading_data[ TradingData.TRADING_FEE_DISCOUNT ] == 1:
                self.dict_all_account_ui_state[ str_tab_widget_name ][ "discount_checkbox"] = False
                self.dict_all_account_ui_state[ str_tab_widget_name ][ "discount_value"] = 0.6
            else:
                self.dict_all_account_ui_state[ str_tab_widget_name ][ "discount_checkbox"] = True
                self.dict_all_account_ui_state[ str_tab_widget_name ][ "discount_value"] = dict_trading_data[ TradingData.TRADING_FEE_DISCOUNT ]

            self.dict_all_account_ui_state[ str_tab_widget_name ][ "trading_fee_type"] = dict_trading_data[ TradingData.TRADING_FEE_TYPE ]
            self.dict_all_account_ui_state[ str_tab_widget_name ][ "trading_fee_minimum"] = dict_trading_data[ TradingData.TRADING_FEE_MINIMUM ]
            self.dict_all_account_ui_state[ str_tab_widget_name ][ "trading_fee_constant"] = dict_trading_data[ TradingData.TRADING_FEE_CONSTANT ]


            dict_per_account_all_stock_trading_data[ str_stock_number ].append( dict_trading_data )
            sorted_list = self.process_single_trading_data( str_tab_widget_name, str_stock_number )
            self.refresh_stock_list_table()
            self.refresh_trading_data_table( sorted_list )
            self.auto_save_trading_data()

    def on_add_dividend_data_push_button_clicked( self ): 
        if self.str_picked_stock_number is None:
            return

        str_tab_widget_name = self.ui.qtTabWidget.currentWidget().objectName()
        dict_per_account_all_stock_trading_data = self.dict_all_account_all_stock_trading_data[ str_tab_widget_name ]

        str_stock_number = self.str_picked_stock_number
        list_stock_name_and_type = self.dict_all_company_number_to_name_and_type[ str_stock_number ]
        str_stock_name = list_stock_name_and_type[ 0 ]
        dialog = StockDividendEditDialog( str_stock_number, str_stock_name, self )

        if dialog.exec():
            dict_trading_data = dialog.dict_trading_data
            dict_per_account_all_stock_trading_data[ str_stock_number ].append( dict_trading_data )
            sorted_list = self.process_single_trading_data( str_tab_widget_name, str_stock_number )
            self.refresh_stock_list_table()
            self.refresh_trading_data_table( sorted_list )
            self.auto_save_trading_data()

    def on_add_limit_buying_data_push_button_clicked( self ): 
        if self.str_picked_stock_number is None:
            return

        str_tab_widget_name = self.ui.qtTabWidget.currentWidget().objectName()
        dict_per_account_all_stock_trading_data = self.dict_all_account_all_stock_trading_data[ str_tab_widget_name ]

        str_stock_number = self.str_picked_stock_number
        list_stock_name_and_type = self.dict_all_company_number_to_name_and_type[ str_stock_number ]
        str_stock_name = list_stock_name_and_type[ 0 ]
        dialog = StockCapitalIncreaseEditDialog( str_stock_number, str_stock_name, self )

        if dialog.exec():
            pass
            dict_trading_data = dialog.dict_trading_data

            dict_per_account_all_stock_trading_data[ str_stock_number ].append( dict_trading_data )
            sorted_list = self.process_single_trading_data( str_tab_widget_name, str_stock_number )
            self.refresh_stock_list_table()
            self.refresh_trading_data_table( sorted_list )
            self.auto_save_trading_data()

    def on_add_capital_reduction_data_push_button_clicked( self ): 
        if self.str_picked_stock_number is None:
            return
        str_tab_widget_name = self.ui.qtTabWidget.currentWidget().objectName()
        dict_per_account_all_stock_trading_data = self.dict_all_account_all_stock_trading_data[ str_tab_widget_name ]

        str_stock_number = self.str_picked_stock_number
        list_stock_name_and_type = self.dict_all_company_number_to_name_and_type[ str_stock_number ]
        str_stock_name = list_stock_name_and_type[ 0 ]
        dialog = StockCapitalReductionEditDialog( str_stock_number, str_stock_name, self )

        if dialog.exec():
            dict_trading_data = dialog.dict_trading_data
            dict_per_account_all_stock_trading_data[ str_stock_number ].append( dict_trading_data )
            sorted_list = self.process_single_trading_data( str_tab_widget_name, str_stock_number )
            self.refresh_stock_list_table()
            self.refresh_trading_data_table( sorted_list )
            self.auto_save_trading_data()

    def on_trading_data_table_item_clicked( self, index: QModelIndex, table_model ): 
        item = table_model.itemFromIndex( index )
        if item is not None:
            n_row = index.row()  # 獲取行索引
            str_tab_widget_name = self.ui.qtTabWidget.currentWidget().objectName()
            dict_per_account_all_stock_trading_data = self.dict_all_account_all_stock_trading_data[ str_tab_widget_name ]
            list_trading_data = dict_per_account_all_stock_trading_data[ self.str_picked_stock_number ]

            if ( n_row == len( self.get_trading_data_header() ) - 2 or #編輯
                n_row == len( self.get_trading_data_header() ) - 1 ): #刪除

                if self.str_picked_stock_number is None:
                    return
                str_stock_number = self.str_picked_stock_number
                list_stock_name_and_type = self.dict_all_company_number_to_name_and_type[ str_stock_number ]
                str_stock_name = list_stock_name_and_type[ 0 ]

                hidden_data = table_model.data( index, Qt.UserRole )
                n_findindex = -1
                for index, dict_selected_data in enumerate( list_trading_data ):
                    if dict_selected_data[ TradingData.SORTED_INDEX_NON_SAVE ] == hidden_data:
                        n_findindex = index
                        break
                if n_findindex == -1:
                    return
                dict_selected_data = list_trading_data[ n_findindex ]

                if n_row == len( self.get_trading_data_header() ) - 2: #編輯
                    if dict_selected_data[ TradingData.TRADING_TYPE ] == TradingType.TEMPLATE:
                        return
                    if dict_selected_data[ TradingData.TRADING_TYPE ] == TradingType.BUY or dict_selected_data[ TradingData.TRADING_TYPE ] == TradingType.SELL:
                        str_b_etf = self.dict_all_company_number_to_name_and_type[ str_stock_number ][ 1 ]
                        b_etf = True if str_b_etf == "True" else False
                        dialog = StockTradingEditDialog( str_stock_number, str_stock_name, b_etf, True, 0, self )
                        dialog.setup_trading_date( dict_selected_data[ TradingData.TRADING_DATE ] )
                        dialog.setup_trading_type( dict_selected_data[ TradingData.TRADING_TYPE ] )
                        dialog.setup_trading_discount( dict_selected_data[ TradingData.TRADING_FEE_DISCOUNT ] )
                        dialog.setup_trading_price( dict_selected_data[ TradingData.TRADING_PRICE ] )
                        dialog.setup_trading_count( dict_selected_data[ TradingData.TRADING_COUNT ] )
                        dialog.setup_daying_trading( dict_selected_data[ TradingData.DAYING_TRADING ] )
                    elif dict_selected_data[ TradingData.TRADING_TYPE ] == TradingType.REGULAR_BUY:
                        dialog = StockRegularTradingEditDialog( str_stock_number, str_stock_name, TradingFeeType.VARIABLE, True, 0, 0, 0, self )
                        dialog.setup_trading_date( dict_selected_data[ TradingData.TRADING_DATE ] )
                        dialog.setup_trading_fee_type( dict_selected_data[ TradingData.TRADING_FEE_TYPE ] )
                        dialog.setup_trading_discount( dict_selected_data[ TradingData.TRADING_FEE_DISCOUNT ] )
                        dialog.setup_trading_fee_minimum( dict_selected_data[ TradingData.TRADING_FEE_MINIMUM ] )
                        dialog.setup_trading_fee_constant( dict_selected_data[ TradingData.TRADING_FEE_CONSTANT ] )
                        dialog.setup_trading_price( dict_selected_data[ TradingData.TRADING_PRICE ] )
                        dialog.setup_trading_count( dict_selected_data[ TradingData.TRADING_COUNT ] )
                    elif dict_selected_data[ TradingData.TRADING_TYPE ] == TradingType.DIVIDEND:
                        dialog = StockDividendEditDialog( str_stock_number, str_stock_name, self )
                        dialog.setup_trading_date( dict_selected_data[ TradingData.TRADING_DATE ] )
                        dialog.setup_stock_dividend( dict_selected_data[ TradingData.STOCK_DIVIDEND_PER_SHARE ] )
                        dialog.setup_cash_dividend( dict_selected_data[ TradingData.CASH_DIVIDEND_PER_SHARE ] )
                    elif dict_selected_data[ TradingData.TRADING_TYPE ] == TradingType.CAPITAL_INCREASE:
                        dialog = StockCapitalIncreaseEditDialog( str_stock_number, str_stock_name, self )
                        dialog.setup_trading_date( dict_selected_data[ TradingData.TRADING_DATE ] )
                        dialog.setup_trading_price( dict_selected_data[ TradingData.TRADING_PRICE ] )
                        dialog.setup_trading_count( dict_selected_data[ TradingData.TRADING_COUNT ] )
                    elif dict_selected_data[ TradingData.TRADING_TYPE ] == TradingType.CAPITAL_REDUCTION:
                        dialog = StockCapitalReductionEditDialog( str_stock_number, str_stock_name, self )
                        dialog.setup_trading_date( dict_selected_data[ TradingData.TRADING_DATE ] )
                        dialog.setup_stock_capital_reduction( dict_selected_data[ TradingData.CAPITAL_REDUCTION_PER_SHARE ] )

                    if dialog.exec():
                        dict_trading_data = dialog.dict_trading_data
                        dict_per_account_all_stock_trading_data[ str_stock_number ][ n_findindex ] = dict_trading_data
                        sorted_list = self.process_single_trading_data( str_tab_widget_name, str_stock_number )
                        self.refresh_stock_list_table()
                        self.refresh_trading_data_table( sorted_list )
                        self.auto_save_trading_data()

                elif n_row == len( self.get_trading_data_header() ) - 1: #刪除
                    result = self.show_warning_message_box_with_ok_cancel_button( "警告", f"確定要刪掉這筆交易資料嗎?" )
                    if result:
                        del dict_per_account_all_stock_trading_data[ str_stock_number ][ n_findindex ]
                        sorted_list = self.process_single_trading_data( str_tab_widget_name, str_stock_number )
                        self.refresh_stock_list_table()
                        self.refresh_trading_data_table( sorted_list )
                        self.auto_save_trading_data()

    def export_trading_data_to_excel( self, worksheet, str_stock_number, str_stock_name, list_trading_data ): 
        all_thin_border = Border(
        left = Side( style = 'thin', color = "000000" ),  # 左邊框：細線，黑色
        right = Side( style = 'thin', color = "000000" ), # 右邊框：細線，黑色
        top = Side( style = 'thin', color = "000000" ),   # 上邊框：細線，黑色
        bottom = Side( style = 'thin', color = "000000" ) # 下邊框：細線，黑色
        )

        top_thick_border = Border(
        left = Side( style = 'thin', color = "000000" ),  # 左邊框：細線，黑色
        right = Side( style = 'thin', color = "000000" ), # 右邊框：細線，黑色
        top = Side( style = 'thick', color = "000000" ),   # 上邊框：細線，黑色
        bottom = Side( style = 'thin', color = "000000" ) # 下邊框：細線，黑色
        )

        bottom_thick_border = Border(
        left = Side( style = 'thin', color = "000000" ),  # 左邊框：細線，黑色
        right = Side( style = 'thin', color = "000000" ), # 右邊框：細線，黑色
        top = Side( style = 'thin', color = "000000" ),   # 上邊框：細線，黑色
        bottom = Side( style = 'thick', color = "000000" ) # 下邊框：細線，黑色
        )

        right_thick_border = Border(
        left = Side( style = 'thin', color = "000000" ),  # 左邊框：細線，黑色
        right = Side( style = 'thick', color = "000000" ), # 右邊框：細線，黑色
        top = Side( style = 'thin', color = "000000" ),   # 上邊框：細線，黑色
        bottom = Side( style = 'thin', color = "000000" ) # 下邊框：細線，黑色
        )

        str_title = str_stock_number + " " + str_stock_name
        worksheet.column_dimensions['A'].width = 22

        b_use_auto_dividend = list_trading_data[ 0 ][ TradingData.USE_AUTO_DIVIDEND_DATA ]
        data_index = 0
        n_row_start = 0
        for dict_per_trading_data in list_trading_data:
            
            e_trading_type = dict_per_trading_data[ TradingData.TRADING_TYPE ]
            if e_trading_type == TradingType.TEMPLATE:
                continue

            if e_trading_type == TradingType.DIVIDEND:
                if b_use_auto_dividend:#使用自動帶入股利資料，但是這筆資料不是自動股利資料，就跳過，反之亦然
                    if TradingData.IS_AUTO_DIVIDEND_DATA_NON_SAVE not in dict_per_trading_data or not dict_per_trading_data[TradingData.IS_AUTO_DIVIDEND_DATA_NON_SAVE]:
                        continue
                else:
                    if TradingData.IS_AUTO_DIVIDEND_DATA_NON_SAVE in dict_per_trading_data and dict_per_trading_data[TradingData.IS_AUTO_DIVIDEND_DATA_NON_SAVE]:
                        continue
            if data_index % 10 == 0:
                list_data_header = self.get_trading_data_header()
                list_data_header.insert( 0, str_title )

                n_row_start = int( ( len( list_data_header ) -2 + 1 ) * int( data_index / 10 ) )
                for index_row, str_header in enumerate( list_data_header ):
                    if index_row == len( list_data_header ) - 2:
                        break
                    worksheet.cell( row = n_row_start + index_row + 1, column = 1, value = str_header ).border = all_thin_border
                    if index_row == 0:
                        worksheet.cell( row = n_row_start + index_row + 1, column = 1 ).font = Font( bold = True )
                index_column = 0

            worksheet.column_dimensions[ get_column_letter( index_column + 2 ) ].width = 12

            list_data = self.get_per_trading_data_text_list( dict_per_trading_data )
            list_data.insert( 0, "" )
            for index_row, str_data in enumerate( list_data ):
                str_data = str_data.replace( ',', '' )
                n_cell_row = n_row_start + index_row + 1
                n_cell_column = index_column + 2
                cell = worksheet.cell( row = n_cell_row, column = n_cell_column )
                if str_data == "買進":
                    color_fill = PatternFill( start_color = "DA9694", end_color = "DA9694", fill_type="solid")
                    cell.fill = color_fill
                elif str_data == "賣出":
                    color_fill = PatternFill( start_color = "76933C", end_color = "76933C", fill_type="solid")
                    cell.fill = color_fill
                elif str_data == "股利分配":
                    color_fill = PatternFill( start_color = "8DB4E2", end_color = "8DB4E2", fill_type="solid")
                    cell.fill = color_fill
                elif str_data == "減資":
                    color_fill = PatternFill( start_color = "B1A0C7", end_color = "B1A0C7", fill_type="solid")
                    cell.fill = color_fill
                elif str_data == "增資":
                    color_fill = PatternFill( start_color = "FABF8F", end_color = "FABF8F", fill_type="solid")
                    cell.fill = color_fill

                str_cell = get_column_letter( n_cell_column ) + str( n_cell_row )
                if index_row == 1:
                    worksheet.cell( row = n_cell_row, column = n_cell_column, value = str_data ).border = all_thin_border
                else:
                    try:
                        f_data= float( str_data )
                        if f_data.is_integer():
                            f_data = int( f_data )
                            worksheet[ str_cell ].number_format = "#,##0"  #顯示千位逗號
                        elif ( f_data * 10 ).is_integer():
                            worksheet[ str_cell ].number_format = "#,##0.0"
                        elif ( f_data * 100 ).is_integer():
                            worksheet[ str_cell ].number_format = "#,##0.00"
                        elif ( f_data * 1000 ).is_integer():
                            worksheet[ str_cell ].number_format = "#,##0.000"
                        worksheet.cell( row = n_cell_row, column = n_cell_column, value = f_data ).border = all_thin_border
                    except ValueError:
                        worksheet.cell( row = n_cell_row, column = n_cell_column, value = str_data ).border = all_thin_border

                worksheet[ str_cell ].alignment = Alignment( horizontal = "center", vertical = "center", shrink_to_fit = True )
            data_index += 1
            index_column += 1

    def on_export_selected_to_excell_button_clicked( self ): 
        if self.str_picked_stock_number is None:
            return
        str_tab_widget_name = self.ui.qtTabWidget.currentWidget().objectName()
        str_stock_number = self.str_picked_stock_number
        list_stock_name_and_type = self.dict_all_company_number_to_name_and_type[ str_stock_number ]
        str_stock_name = list_stock_name_and_type[ 0 ]
        dict_per_account_all_stock_trading_data = self.dict_all_account_all_stock_trading_data[ str_tab_widget_name ]
        list_trading_data = dict_per_account_all_stock_trading_data[ str_stock_number ]
        file_path = self.open_save_excel_file_dialog()
        if file_path:
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = str_stock_number + " " + str_stock_name
            worksheet.page_setup.orientation = "landscape"
            worksheet.page_setup.paperSize = 9
            worksheet.page_margins.left = 0.7    # 左邊界
            worksheet.page_margins.right = 0.7   # 右邊界
            worksheet.page_margins.top = 0.75    # 上邊界
            worksheet.page_margins.bottom = 0.75 # 下邊界
            worksheet.page_margins.header = 0.3  # 頁眉邊界
            worksheet.page_margins.footer = 0.3  # 頁腳邊界
            self.export_trading_data_to_excel( worksheet, str_stock_number, str_stock_name, list_trading_data )
            workbook.save( file_path )

    def on_export_all_to_excell_button_clicked( self ): 
        file_path = self.open_save_excel_file_dialog()
        if file_path:
            workbook = Workbook()
            str_tab_widget_name = self.ui.qtTabWidget.currentWidget().objectName()
            dict_per_account_all_stock_trading_data = self.dict_all_account_all_stock_trading_data[ str_tab_widget_name ]

            for index, ( key_stock_number, value_list_trading_data ) in enumerate( dict_per_account_all_stock_trading_data.items() ):
                list_stock_name_and_type = self.dict_all_company_number_to_name_and_type[ key_stock_number ]
                str_stock_name = list_stock_name_and_type[ 0 ]
                str_tab_title = key_stock_number + " " + str_stock_name
                if index == 0:
                    worksheet = workbook.active
                    worksheet.title = str_tab_title
                    worksheet.page_setup.orientation = "landscape"
                    worksheet.page_setup.paperSize = 9
                    worksheet.page_margins.left = 0.7    # 左邊界
                    worksheet.page_margins.right = 0.7   # 右邊界
                    worksheet.page_margins.top = 0.75    # 上邊界
                    worksheet.page_margins.bottom = 0.75 # 下邊界
                    worksheet.page_margins.header = 0.3  # 頁眉邊界
                    worksheet.page_margins.footer = 0.3  # 頁腳邊界
                else:
                    worksheet = workbook.create_sheet( str_tab_title, index )
                    worksheet.page_setup.orientation = "landscape"
                    worksheet.page_setup.paperSize = 9
                    worksheet.page_margins.left = 0.7    # 左邊界
                    worksheet.page_margins.right = 0.7   # 右邊界
                    worksheet.page_margins.top = 0.75    # 上邊界
                    worksheet.page_margins.bottom = 0.75 # 下邊界
                    worksheet.page_margins.header = 0.3  # 頁眉邊界
                    worksheet.page_margins.footer = 0.3  # 頁腳邊界
                self.export_trading_data_to_excel( worksheet, key_stock_number, str_stock_name, value_list_trading_data )
            workbook.save( file_path )

    def on_new_file_action_triggered( self ): 
        b_need_save = False
        if ( ( self.dict_all_account_all_stock_trading_data_INITIAL != self.dict_all_account_all_stock_trading_data ) and
             ( self.str_current_save_file_path is None or not self.compare_json_files( self.str_current_save_file_path, self.trading_data_json_file_path ) ) ):
            dialog = SaveCheckDialog( '開新檔案', self )
            if dialog.exec():
                if dialog.n_return == 1: #儲存
                    b_need_save = True
                elif dialog.n_return == 2: #不儲存
                    pass
                else: #取消
                    return
            else:
                return
        if b_need_save:
            if self.str_current_save_file_path:
                list_save_tab_widget = list( range( self.ui.qtTabWidget.count() - 1 ) )
                self.manual_save_trading_data( list_save_tab_widget, self.str_current_save_file_path )
            else:
                file_path = self.open_save_json_file_dialog()
                if file_path:
                    list_save_tab_widget = list( range( self.ui.qtTabWidget.count() - 1 ) )
                    self.manual_save_trading_data( list_save_tab_widget, file_path )
                    self.str_current_save_file_path = file_path
                else:
                    return

        with QSignalBlocker( self.ui.qtTabWidget ):
            for index in range( self.ui.qtTabWidget.count() - 2, -1, -1 ):
                self.ui.qtTabWidget.removeTab( index )

        self.dict_all_account_all_stock_trading_data.clear()
        self.dict_all_account_cash_transfer_data.clear()
        self.dict_all_account_ui_state.clear()
        str_tab_name = self.add_new_tab_and_table()
        self.dict_all_account_all_stock_trading_data[ str_tab_name ] = {}
        self.dict_all_account_cash_transfer_data[ str_tab_name ] = []
        self.dict_all_account_ui_state[ str_tab_name ] = { "discount_checkbox": True, "discount_value": 0.6, "insurance_checkbox": False, "trading_fee_type": TradingFeeType.VARIABLE, "trading_fee_minimum": 1, "trading_fee_constant": 1 }
        self.dict_all_account_all_stock_trading_data_INITIAL = copy.deepcopy( self.dict_all_account_all_stock_trading_data )
        self.ui.qtTabWidget.setCurrentIndex( 0 )
        self.pick_up_stock( None )
        self.refresh_stock_list_table()
        self.refresh_transfer_data_table()
        self.clear_per_stock_trading_table()
        self.update_button_enable_disable_status()
        self.auto_save_trading_data()
        self.str_current_save_file_path = None

    def on_open_file_action_triggered( self ): 
        file_path = self.open_load_json_file_dialog()
        if file_path:
            b_need_save = False
            if ( ( self.dict_all_account_all_stock_trading_data_INITIAL != self.dict_all_account_all_stock_trading_data ) and
                 ( self.str_current_save_file_path is None or not self.compare_json_files( self.str_current_save_file_path, self.trading_data_json_file_path ) ) ):
                dialog = SaveCheckDialog( '開啟舊檔', self )
                if dialog.exec():
                    if dialog.n_return == 1: #儲存
                        b_need_save = True
                    elif dialog.n_return == 2: #不儲存
                        pass
                    else: #取消
                        return
                else:
                    return
            if b_need_save:
                if self.str_current_save_file_path:
                    list_save_tab_widget = list( range( self.ui.qtTabWidget.count() - 1 ) )
                    self.manual_save_trading_data( list_save_tab_widget, self.str_current_save_file_path )
                else:
                    file_path = self.open_save_json_file_dialog()
                    if file_path:
                        list_save_tab_widget = list( range( self.ui.qtTabWidget.count() - 1 ) )
                        self.manual_save_trading_data( list_save_tab_widget, file_path )
                        self.str_current_save_file_path = file_path
                    else:
                        return

            with QSignalBlocker( self.ui.qtTabWidget ):
                for index in range( self.ui.qtTabWidget.count() - 2, -1, -1 ):
                    self.ui.qtTabWidget.removeTab( index )


                self.dict_all_account_all_stock_trading_data.clear()
                self.dict_all_account_cash_transfer_data.clear()
                self.dict_all_account_ui_state.clear()
                self.load_trading_data_and_create_tab( file_path, 
                                                    self.dict_all_account_all_stock_trading_data, 
                                                    self.dict_all_account_ui_state, 
                                                    self.dict_all_account_cash_transfer_data,
                                                    True )
                if len( self.dict_all_account_all_stock_trading_data ) == 0:
                    str_tab_name = self.add_new_tab_and_table()
                    self.dict_all_account_all_stock_trading_data[ str_tab_name ] = {}
                    self.dict_all_account_cash_transfer_data[ str_tab_name ] = []
                    self.dict_all_account_ui_state[ str_tab_name ] = { "discount_checkbox": True, "discount_value": 0.6, "insurance_checkbox": False, "trading_fee_type": TradingFeeType.VARIABLE, "trading_fee_minimum": 1, "trading_fee_constant": 1 }
                self.ui.qtTabWidget.setCurrentIndex( 0 )

            self.process_all_trading_data()
            self.pick_up_stock( None )
            self.refresh_stock_list_table()
            self.process_all_transfer_data()
            self.refresh_transfer_data_table()
            self.clear_per_stock_trading_table()
            self.update_button_enable_disable_status()
            self.auto_save_trading_data()
            self.str_current_save_file_path = file_path

    def on_save_as_action_triggered( self ): 
        file_path = self.open_save_json_file_dialog()
        if file_path:
            list_save_tab_widget = list( range( self.ui.qtTabWidget.count() - 1 ) )
            self.manual_save_trading_data( list_save_tab_widget, file_path )
            self.str_current_save_file_path = file_path

    def on_save_action_triggered( self ): 
        file_path = self.str_current_save_file_path or self.open_save_json_file_dialog()
        if file_path:
            list_save_tab_widget = list( range( self.ui.qtTabWidget.count() - 1 ) )
            self.manual_save_trading_data( list_save_tab_widget, file_path )
            self.str_current_save_file_path = file_path

    def on_export_current_group_action_triggered( self ): 
        file_path = self.open_save_json_file_dialog()
        if file_path:
            list_save_tab_widget = [ self.ui.qtTabWidget.currentIndex() ]
            self.manual_save_trading_data( list_save_tab_widget, file_path )

    def on_import_full_action_triggered( self ):
        file_path = self.open_load_json_file_dialog()
        if file_path:
            dict_account_to_tab_widget_name = {}
            for index in range( self.ui.qtTabWidget.count() - 1 ):
                tab_widget = self.ui.qtTabWidget.widget( index )
                str_tab_title = self.ui.qtTabWidget.tabText( index )
                str_tab_widget_name = tab_widget.objectName()
                dict_account_to_tab_widget_name[ str_tab_title ] = str_tab_widget_name

                # self.dict_all_account_all_stock_trading_data[ self.str_picked_stock_number ] 

            dict_all_account_all_stock_trading_data_LOAD = {}
            dict_all_account_ui_state_LOAD = {}
            dict_all_account_cash_transfer_data_LOAD = {}
            self.load_trading_data_and_create_tab( file_path, 
                                                   dict_all_account_all_stock_trading_data_LOAD, 
                                                   dict_all_account_ui_state_LOAD, 
                                                   dict_all_account_cash_transfer_data_LOAD, 
                                                   False )
            b_duplicate = False
            for str_account_name, dict_per_account_all_stock_trading_data_LOAD in dict_all_account_all_stock_trading_data_LOAD.items():
                if str_account_name in dict_account_to_tab_widget_name:
                    str_tab_widget_name = dict_account_to_tab_widget_name[ str_account_name ]
                    for key_stock_number, value in dict_per_account_all_stock_trading_data_LOAD.items():
                        if key_stock_number in self.dict_all_account_all_stock_trading_data[ str_tab_widget_name ]:
                            b_duplicate = True
                            break
                    if b_duplicate:
                        break

            if b_duplicate:
                dialog = ImportDataDuplicateOptionDialog( self )
                if dialog.exec():
                    if dialog.b_overwrite:
                        for str_account_name, dict_per_account_all_stock_trading_data_LOAD in dict_all_account_all_stock_trading_data_LOAD.items():
                            if str_account_name in dict_account_to_tab_widget_name:
                                str_tab_widget_name = dict_account_to_tab_widget_name[ str_account_name ]
                                self.dict_all_account_all_stock_trading_data[ str_tab_widget_name ].update( dict_per_account_all_stock_trading_data_LOAD )
                            else:
                                str_tab_widget_name = self.add_new_tab_and_table( str_account_name )
                                self.dict_all_account_all_stock_trading_data[ str_tab_widget_name ] = dict_per_account_all_stock_trading_data_LOAD
                                dict_per_account_ui_state_LOAD = dict_all_account_ui_state_LOAD[ str_account_name ]
                                self.dict_all_account_ui_state[ str_tab_widget_name ] = dict_per_account_ui_state_LOAD
                    else:
                        for str_account_name, dict_per_account_all_stock_trading_data_LOAD in dict_all_account_all_stock_trading_data_LOAD.items():
                            if str_account_name in dict_account_to_tab_widget_name:
                                str_tab_widget_name = dict_account_to_tab_widget_name[ str_account_name ]
                                for key_stock_number, value in dict_per_account_all_stock_trading_data_LOAD.items():
                                    if key_stock_number not in self.dict_all_account_all_stock_trading_data[ str_tab_widget_name ]:
                                        self.dict_all_account_all_stock_trading_data[ str_tab_widget_name ][ key_stock_number ] = value
                                    else:
                                        value.pop( 0 ) #移除第一筆資料 因為第一筆資料是虛的
                                        self.dict_all_account_all_stock_trading_data[ str_tab_widget_name ][ key_stock_number ].extend( value )
                            else:
                                str_tab_widget_name = self.add_new_tab_and_table( str_account_name )
                                self.dict_all_account_all_stock_trading_data[ str_tab_widget_name ] = dict_per_account_all_stock_trading_data_LOAD
                                dict_per_account_ui_state_LOAD = dict_all_account_ui_state_LOAD[ str_account_name ]
                                self.dict_all_account_ui_state[ str_tab_widget_name ] = dict_per_account_ui_state_LOAD
            else:
                for str_account_name, dict_per_account_all_stock_trading_data_LOAD in dict_all_account_all_stock_trading_data_LOAD.items():
                    if str_account_name in dict_account_to_tab_widget_name:
                        str_tab_widget_name = dict_account_to_tab_widget_name[ str_account_name ]
                        self.dict_all_account_all_stock_trading_data[ str_tab_widget_name ].update( dict_per_account_all_stock_trading_data_LOAD )
                    else:
                        str_tab_widget_name = self.add_new_tab_and_table( str_account_name )
                        self.dict_all_account_all_stock_trading_data[ str_tab_widget_name ] = dict_per_account_all_stock_trading_data_LOAD
                        dict_per_account_ui_state_LOAD = dict_all_account_ui_state_LOAD[ str_account_name ]
                        self.dict_all_account_ui_state[ str_tab_widget_name ] = dict_per_account_ui_state_LOAD
                    self.dict_all_account_cash_transfer_data[ str_tab_widget_name ] = dict_all_account_cash_transfer_data_LOAD[ str_account_name ]

            self.process_all_trading_data()
            self.pick_up_stock( None )
            self.refresh_stock_list_table()
            self.process_all_transfer_data()
            self.refresh_transfer_data_table()
            self.clear_per_stock_trading_table()
            self.update_button_enable_disable_status()
            self.auto_save_trading_data()

    def on_import_single_stock_action_triggered( self ):
        file_path = self.open_load_json_file_dialog()
        if file_path:
            dict_all_account_all_stock_trading_data_LOAD = {}
            dict_all_account_ui_state_LOAD = {}
            dict_all_account_cash_transfer_data_LOAD = {}
            self.load_trading_data_and_create_tab( file_path, 
                                                   dict_all_account_all_stock_trading_data_LOAD, 
                                                   dict_all_account_ui_state_LOAD, 
                                                   dict_all_account_cash_transfer_data_LOAD, 
                                                   False )
            if len( dict_all_account_all_stock_trading_data_LOAD ) == 0:
                return
            elif len( dict_all_account_all_stock_trading_data_LOAD ) > 1:
                self.show_error_message_box_with_ok_button( "錯誤", "請選擇只包含單一帳戶及單一個股的檔案" )
                return
            
            dict_per_account_all_stock_trading_data_LOAD = dict_all_account_all_stock_trading_data_LOAD.popitem()[ 1 ]
            if len( dict_per_account_all_stock_trading_data_LOAD ) == 0:
                return
            elif len( dict_per_account_all_stock_trading_data_LOAD ) > 1:
                self.show_error_message_box_with_ok_button( "錯誤", "請選擇只包含單一帳戶及單一個股的檔案" )
                return
            str_stock_number, list_trading_data = dict_per_account_all_stock_trading_data_LOAD.popitem()

            if self.ui.qtTabWidget.count() <= 1:
                self.show_error_message_box_with_ok_button( "錯誤", "請先建立群組" )
                return

            str_tab_widget_name = self.ui.qtTabWidget.currentWidget().objectName()
            dict_per_account_all_stock_trading_data = self.dict_all_account_all_stock_trading_data[ str_tab_widget_name ]

            b_duplicate = str_stock_number in dict_per_account_all_stock_trading_data

            if b_duplicate:
                dialog = ImportDataDuplicateOptionDialog( self )
                if dialog.exec():
                    if dialog.b_overwrite:
                        self.dict_all_account_all_stock_trading_data[ str_tab_widget_name ][ str_stock_number ] = list_trading_data
                    else:
                        list_trading_data.pop( 0 ) #移除第一筆資料 因為第一筆資料是虛的
                        self.dict_all_account_all_stock_trading_data[ str_tab_widget_name ][ str_stock_number ].extend( list_trading_data )
            else:
                self.dict_all_account_all_stock_trading_data[ str_tab_widget_name ][ str_stock_number ] = list_trading_data

            self.process_all_trading_data()
            self.pick_up_stock( None )
            self.refresh_stock_list_table()
            self.clear_per_stock_trading_table()
            self.update_button_enable_disable_status()
            self.auto_save_trading_data()

    def load_initialize_data( self ): 
        with QSignalBlocker( self.ui.qtTabWidget ):
            self.load_trading_data_and_create_tab( self.trading_data_json_file_path, 
                                                   self.dict_all_account_all_stock_trading_data, 
                                                   self.dict_all_account_ui_state, 
                                                   self.dict_all_account_cash_transfer_data, 
                                                   True )
            self.load_share_UI_state()
            if len( self.dict_all_account_all_stock_trading_data ) == 0:
                str_tab_name = self.add_new_tab_and_table()
                self.dict_all_account_all_stock_trading_data[ str_tab_name ] = {}
                self.dict_all_account_ui_state[ str_tab_name ] = { "discount_checkbox": True, "discount_value": 0.6, "insurance_checkbox": False, "trading_fee_type": TradingFeeType.VARIABLE, "trading_fee_minimum": 1, "trading_fee_constant": 1 }
                self.dict_all_account_cash_transfer_data[ str_tab_name ] = []
            self.dict_all_account_all_stock_trading_data_INITIAL = self.dict_all_account_all_stock_trading_data.copy()
            self.ui.qtTabWidget.setCurrentIndex( 0 )
            self.process_all_trading_data()
            self.refresh_stock_list_table()
            self.process_all_transfer_data()
            self.refresh_transfer_data_table()

    def load_trading_data_and_create_tab( self, file_path, 
                                                dict_all_account_all_stock_trading_data, 
                                                dict_all_account_ui_state, 
                                                dict_all_account_cash_transfer, 
                                                b_create_tab ): 
        if not os.path.exists( file_path ):
            return

        with open( file_path, 'r', encoding='utf-8' ) as f:
            version = f.readline().strip()
            data = json.load( f )

        if version == "v1.0.0":
            for item_account in data:
                if "account_name" in item_account and \
                   "trading_data" in item_account and \
                   "discount_checkbox" in item_account and \
                   "discount_value" in item_account and \
                   "insurance_checkbox" in item_account:
                    dict_per_account_all_stock_trading_data = item_account[ "trading_data" ]
                    dict_ui_state = {}
                    dict_ui_state[ "discount_checkbox" ] = item_account[ "discount_checkbox" ]
                    dict_ui_state[ "discount_value" ] = item_account[ "discount_value" ]
                    dict_ui_state[ "insurance_checkbox" ] = item_account[ "insurance_checkbox" ]
                    dict_ui_state[ "trading_fee_type" ] = TradingFeeType.VARIABLE
                    dict_ui_state[ "trading_fee_minimum" ] = 1
                    dict_ui_state[ "trading_fee_constant" ] = 1

                    dict_per_stock_trading_data = {} 
                    for key_stock_number, value_list_trading_data  in dict_per_account_all_stock_trading_data.items():
                        list_trading_data = []
                        for item_trading_data in value_list_trading_data:
                            if ( "trading_date" in item_trading_data and
                                 "trading_type" in item_trading_data and
                                 "trading_price" in item_trading_data and
                                 "trading_count" in item_trading_data and
                                 "trading_fee_discount" in item_trading_data and
                                 "stock_dividend_per_share" in item_trading_data and
                                 "cash_dividend_per_share" in item_trading_data and
                                 "capital_reduction_per_share" in item_trading_data ):
                                
                                if item_trading_data[ "trading_type" ] == 3:
                                    e_trading_type = TradingType.CAPITAL_INCREASE
                                elif item_trading_data[ "trading_type" ] == 4:
                                    e_trading_type = TradingType.DIVIDEND
                                elif item_trading_data[ "trading_type" ] == 5:
                                    e_trading_type = TradingType.CAPITAL_REDUCTION
                                else:
                                    e_trading_type = TradingType( item_trading_data[ "trading_type" ] )
                                dict_per_trading_data = Utility.generate_trading_data( item_trading_data[ "trading_date" ],                 #交易日期
                                                                                       e_trading_type,                                      #交易種類
                                                                                       item_trading_data[ "trading_price" ],                #交易價格
                                                                                       item_trading_data[ "trading_count" ],                #交易股數
                                                                                       TradingFeeType.VARIABLE,                             #手續費種類
                                                                                       item_trading_data[ "trading_fee_discount" ],         #手續費折扣
                                                                                       0,                                                   #手續費最低金額
                                                                                       0,                                                   #手續費固定金額
                                                                                       item_trading_data[ "stock_dividend_per_share" ],     #每股股票股利
                                                                                       item_trading_data[ "cash_dividend_per_share" ],      #每股現金股利
                                                                                       item_trading_data[ "capital_reduction_per_share" ],  #每股減資金額 
                                                                                       False )                                              #是否為當沖交易  
                                if item_trading_data[ "trading_date" ] == '0001-01-01':
                                    dict_per_trading_data[ TradingData.USE_AUTO_DIVIDEND_DATA ] = item_trading_data[ "use_auto_dividend_data" ]       
                                list_trading_data.append( dict_per_trading_data )
                        dict_per_stock_trading_data[ key_stock_number ] = list_trading_data
                    if b_create_tab:
                        str_tab_name = self.add_new_tab_and_table( item_account[ "account_name" ] )
                        dict_all_account_ui_state[ str_tab_name ] = dict_ui_state
                        dict_all_account_all_stock_trading_data[ str_tab_name ] = dict_per_stock_trading_data
                        dict_all_account_cash_transfer[ str_tab_name ] = []
                    else:
                        dict_all_account_ui_state[ item_account[ "account_name" ] ] = dict_ui_state
                        dict_all_account_all_stock_trading_data[ item_account[ "account_name" ] ] = dict_per_stock_trading_data
                        dict_all_account_cash_transfer[ item_account[ "account_name" ] ] = []
        else:
            for item_account in data:
                if "account_name" in item_account and \
                   "trading_data" in item_account and \
                   "discount_checkbox" in item_account and \
                   "discount_value" in item_account and \
                   "insurance_checkbox" in item_account and \
                    "trading_fee_type" in item_account and \
                    "trading_fee_minimum" in item_account and \
                    "trading_fee_constant" in item_account:
                    dict_per_account_all_stock_trading_data = item_account[ "trading_data" ]
                    dict_ui_state = {}
                    dict_ui_state[ "discount_checkbox" ] = item_account[ "discount_checkbox" ]
                    dict_ui_state[ "discount_value" ] = item_account[ "discount_value" ]
                    dict_ui_state[ "insurance_checkbox" ] = item_account[ "insurance_checkbox" ]
                    dict_ui_state[ "trading_fee_type" ] = TradingFeeType( item_account[ "trading_fee_type" ] )
                    dict_ui_state[ "trading_fee_minimum" ] = item_account[ "trading_fee_minimum" ]
                    dict_ui_state[ "trading_fee_constant" ] = item_account[ "trading_fee_constant" ]

                    dict_per_stock_trading_data = {} 
                    for key_stock_number, value_list_trading_data  in dict_per_account_all_stock_trading_data.items():
                        list_trading_data = []
                        for item_trading_data in value_list_trading_data:
                            if ( "trading_date" in item_trading_data and
                                 "trading_type" in item_trading_data and
                                 "trading_price" in item_trading_data and
                                 "trading_count" in item_trading_data and
                                 "trading_fee_discount" in item_trading_data and
                                 "stock_dividend_per_share" in item_trading_data and
                                 "cash_dividend_per_share" in item_trading_data and
                                 "capital_reduction_per_share" in item_trading_data ):
                                
                                e_trading_type = TradingType( item_trading_data[ "trading_type" ] )
                                e_trading_fee_type = TradingFeeType( item_trading_data[ "trading_fee_type" ] )
                                dict_per_trading_data = Utility.generate_trading_data( item_trading_data[ "trading_date" ],                 #交易日期
                                                                                       e_trading_type,                                      #交易種類
                                                                                       item_trading_data[ "trading_price" ],                #交易價格
                                                                                       item_trading_data[ "trading_count" ],                #交易股數
                                                                                       e_trading_fee_type,                                  #手續費種類
                                                                                       item_trading_data[ "trading_fee_discount" ],         #手續費折扣
                                                                                       item_trading_data[ "trading_fee_minimum" ],          #手續費最低金額
                                                                                       item_trading_data[ "trading_fee_constant" ],         #手續費固定金額
                                                                                       item_trading_data[ "stock_dividend_per_share" ],     #每股股票股利
                                                                                       item_trading_data[ "cash_dividend_per_share" ],      #每股現金股利
                                                                                       item_trading_data[ "capital_reduction_per_share" ],  #每股減資金額     
                                                                                       item_trading_data[ "daying_trading" ] )              #是否為當沖交易
                                if item_trading_data[ "trading_date" ] == '0001-01-01':
                                    dict_per_trading_data[ TradingData.USE_AUTO_DIVIDEND_DATA ] = item_trading_data[ "use_auto_dividend_data" ]       
                                list_trading_data.append( dict_per_trading_data )
                        dict_per_stock_trading_data[ key_stock_number ] = list_trading_data

                    list_transfer_data = item_account[ "transfer_data" ]
                    list_per_account_cash_transfer_data = []
                    for item_transfer_data in list_transfer_data:
                        if ( "transfer_date" in item_transfer_data and
                             "transfer_type" in item_transfer_data and
                             "transfer_value" in item_transfer_data ):
                            dict_per_transfer_data = {}
                            dict_per_transfer_data[ TransferData.TRANSFER_DATE ] = item_transfer_data[ "transfer_date" ]
                            dict_per_transfer_data[ TransferData.TRANSFER_TYPE ] = TransferType( item_transfer_data[ "transfer_type" ] )
                            dict_per_transfer_data[ TransferData.TRANSFER_VALUE ] = item_transfer_data[ "transfer_value" ]
                            list_per_account_cash_transfer_data.append( dict_per_transfer_data )

                    if b_create_tab:
                        str_tab_name = self.add_new_tab_and_table( item_account[ "account_name" ] )
                        for index in range( self.ui.qtTabWidget.count() - 1 ):
                            tab_widget = self.ui.qtTabWidget.widget( index )
                            if tab_widget.objectName() == str_tab_name:
                                qt_insurance_check_box = tab_widget.findChild( QCheckBox, "ExtraInsuranceFeeCheckBox" )
                                with QSignalBlocker( qt_insurance_check_box ):
                                    qt_insurance_check_box.setChecked( item_account[ "insurance_checkbox" ] )
                                    break
                        dict_all_account_ui_state[ str_tab_name ] = dict_ui_state
                        dict_all_account_all_stock_trading_data[ str_tab_name ] = dict_per_stock_trading_data
                        dict_all_account_cash_transfer[ str_tab_name ] = list_per_account_cash_transfer_data
                    else:
                        dict_all_account_ui_state[ item_account[ "account_name" ] ] = dict_ui_state
                        dict_all_account_all_stock_trading_data[ item_account[ "account_name" ] ] = dict_per_stock_trading_data
                        dict_all_account_cash_transfer[ item_account[ "account_name" ] ] = list_per_account_cash_transfer_data

    def auto_save_trading_data( self ): 
        list_save_tab_widget = list( range( self.ui.qtTabWidget.count() - 1 ) )
        self.manual_save_trading_data( list_save_tab_widget, self.trading_data_json_file_path )

    def manual_save_trading_data( self, list_save_tab_indice, file_path, str_stock_number = None ): 
        export_list_all_account_all_stock_trading_data = []

        for index in range( self.ui.qtTabWidget.count() - 1 ):
            if index not in list_save_tab_indice:
                continue
            tab_widget = self.ui.qtTabWidget.widget( index )

            qt_insurance_check_box = tab_widget.findChild( QCheckBox, "ExtraInsuranceFeeCheckBox" )

            str_tab_title = self.ui.qtTabWidget.tabText( index )
            str_tab_widget_name = tab_widget.objectName()
            value_dict_per_account_all_stock_trading_data = self.dict_all_account_all_stock_trading_data[ str_tab_widget_name ]
            export_dict_per_account_all_info = {}
            export_dict_per_account_all_stock_trading_data = {}
            for key_stock, value_list_per_stock_trading_data in value_dict_per_account_all_stock_trading_data.items():
                if str_stock_number is not None and key_stock != str_stock_number:
                    continue
                export_data = []
                for item in value_list_per_stock_trading_data:
                    if TradingData.IS_AUTO_DIVIDEND_DATA_NON_SAVE in item and item[ TradingData.IS_AUTO_DIVIDEND_DATA_NON_SAVE ] == True:
                        continue
                    dict_per_trading_data = {}
                    dict_per_trading_data[ "trading_date" ] = item[ TradingData.TRADING_DATE ]
                    dict_per_trading_data[ "trading_type" ] = int( item[ TradingData.TRADING_TYPE ].value )
                    if item[ TradingData.TRADING_TYPE ] == TradingType.CAPITAL_REDUCTION: #CAPITAL_REDUCTION 為了顯示，所以需要寫一些數值進去，但實際上不用存
                        dict_per_trading_data[ "trading_price" ] = 0
                        dict_per_trading_data[ "trading_count" ] = 0
                    else:
                        dict_per_trading_data[ "trading_price" ] = item[ TradingData.TRADING_PRICE ]
                        dict_per_trading_data[ "trading_count" ] = item[ TradingData.TRADING_COUNT ]
                    dict_per_trading_data[ "trading_fee_type" ] = int( item[ TradingData.TRADING_FEE_TYPE ].value )
                    dict_per_trading_data[ "trading_fee_discount" ] = item[ TradingData.TRADING_FEE_DISCOUNT ]
                    dict_per_trading_data[ "trading_fee_minimum" ] = item[ TradingData.TRADING_FEE_MINIMUM ]
                    dict_per_trading_data[ "trading_fee_constant" ] = item[ TradingData.TRADING_FEE_CONSTANT ]
                    dict_per_trading_data[ "stock_dividend_per_share" ] = item[ TradingData.STOCK_DIVIDEND_PER_SHARE ]
                    dict_per_trading_data[ "cash_dividend_per_share" ] = item[ TradingData.CASH_DIVIDEND_PER_SHARE ]
                    dict_per_trading_data[ "capital_reduction_per_share" ] = item[ TradingData.CAPITAL_REDUCTION_PER_SHARE ]
                    if dict_per_trading_data[ "trading_date" ] == '0001-01-01':
                        dict_per_trading_data[ "use_auto_dividend_data" ] = item[ TradingData.USE_AUTO_DIVIDEND_DATA ]
                    dict_per_trading_data[ "daying_trading" ] = item[ TradingData.DAYING_TRADING ]

                    export_data.append( dict_per_trading_data )
                export_dict_per_account_all_stock_trading_data[ key_stock ] = export_data

            list_transfer_data = []
            value_dict_per_account_all_cash_transfer_data = self.dict_all_account_cash_transfer_data[ str_tab_widget_name ]
            for dict_cash_transfer_data in value_dict_per_account_all_cash_transfer_data:
                dict_per_transfer_data = {}
                dict_per_transfer_data[ "transfer_date" ] = dict_cash_transfer_data[ TransferData.TRANSFER_DATE ]
                dict_per_transfer_data[ "transfer_type" ] = int( dict_cash_transfer_data[ TransferData.TRANSFER_TYPE ].value )
                dict_per_transfer_data[ "transfer_value" ] = dict_cash_transfer_data[ TransferData.TRANSFER_VALUE ]
                list_transfer_data.append( dict_per_transfer_data )

            export_dict_per_account_all_info[ "account_name" ] = str_tab_title
            export_dict_per_account_all_info[ "trading_data" ] = export_dict_per_account_all_stock_trading_data
            export_dict_per_account_all_info[ "discount_checkbox" ] = self.dict_all_account_ui_state[ str_tab_widget_name ][ "discount_checkbox"]
            export_dict_per_account_all_info[ "discount_value" ] = self.dict_all_account_ui_state[ str_tab_widget_name ][ "discount_value"]
            export_dict_per_account_all_info[ "insurance_checkbox" ] = qt_insurance_check_box.isChecked()
            export_dict_per_account_all_info[ "trading_fee_type" ] = int( self.dict_all_account_ui_state[ str_tab_widget_name ][ "trading_fee_type"].value )
            export_dict_per_account_all_info[ "trading_fee_minimum" ] = self.dict_all_account_ui_state[ str_tab_widget_name ][ "trading_fee_minimum"]
            export_dict_per_account_all_info[ "trading_fee_constant" ] = self.dict_all_account_ui_state[ str_tab_widget_name ][ "trading_fee_constant"]
            export_dict_per_account_all_info[ "transfer_data" ] = list_transfer_data


            export_list_all_account_all_stock_trading_data.append( export_dict_per_account_all_info )

        with open( file_path, 'w', encoding='utf-8' ) as f:
            f.write( "v1.0.1" '\n' )
            json.dump( export_list_all_account_all_stock_trading_data, f, ensure_ascii=False, indent=4 )
    
    def save_share_UI_state( self ): 
        # 確保目錄存在，若不存在則遞歸創建
        os.makedirs( os.path.dirname( self.UISetting_file_path ), exist_ok = True )

        with open( self.UISetting_file_path, 'w', encoding='utf-8' ) as f:
            f.write( "版本," + 'v1.0.0' + '\n' )
            f.write( "顯示排序," + str( self.ui.qtFromNewToOldAction.isChecked() ) + '\n' )
            f.write( "顯示數量," + str( self.ui.qtShowAllAction.isChecked() ) + '\n' )
            f.write( "顯示單位," + str( self.ui.qtUse1ShareUnitAction.isChecked() ) + '\n' )
            f.write( "年度顯示," + str( self.ui.qtADYearAction.isChecked() ) + '\n' )
            f.write( "欄寬" )
            for i in range( len( self.list_stock_list_column_width ) ):
                f.write( f",{ self.list_stock_list_column_width[ i ] }" )
            f.write( "\n" )

    def load_share_UI_state( self ): 
        with ( QSignalBlocker( self.ui.qtFromNewToOldAction ),
               QSignalBlocker( self.ui.qtFromOldToNewAction ), 
               QSignalBlocker( self.ui.qtShowAllAction ), 
               QSignalBlocker( self.ui.qtShow10Action ),
               QSignalBlocker( self.ui.qtUse1ShareUnitAction ), 
               QSignalBlocker( self.ui.qtUse1000ShareUnitAction ),
               QSignalBlocker( self.ui.qtADYearAction ), 
               QSignalBlocker( self.ui.qtROCYearAction ) 
               ):

            if os.path.exists( self.UISetting_file_path ):
                with open( self.UISetting_file_path, 'r', encoding='utf-8' ) as f:
                    data = f.readlines()
                    for i, row in enumerate( data ):
                        row = row.strip().split( ',' )
                        if row[0] == "版本":
                            continue
                        elif row[0] == "顯示排序":
                            if row[ 1 ] == 'True':
                                self.ui.qtFromNewToOldAction.setChecked( True )
                                self.ui.qtFromOldToNewAction.setChecked( False )
                            else:
                                self.ui.qtFromNewToOldAction.setChecked( False )
                                self.ui.qtFromOldToNewAction.setChecked( True )
                        elif row[0] == "顯示數量":
                            if row[ 1 ] == 'True':
                                self.ui.qtShowAllAction.setChecked( True )
                                self.ui.qtShow10Action.setChecked( False )
                            else:
                                self.ui.qtShowAllAction.setChecked( False )
                                self.ui.qtShow10Action.setChecked( True )
                        elif row[0] == "顯示單位":
                            if row[ 1 ] == 'True':
                                self.ui.qtUse1ShareUnitAction.setChecked( True )
                                self.ui.qtUse1000ShareUnitAction.setChecked( False )
                            else:
                                self.ui.qtUse1ShareUnitAction.setChecked( False )
                                self.ui.qtUse1000ShareUnitAction.setChecked( True )
                        elif row[0] == "年度顯示":
                            if row[ 1 ] == 'True':
                                self.ui.qtADYearAction.setChecked( True )
                                self.ui.qtROCYearAction.setChecked( False )
                            else:
                                self.ui.qtADYearAction.setChecked( False )
                                self.ui.qtROCYearAction.setChecked( True )
                        elif row[0] == '欄寬':
                            self.list_stock_list_column_width = []
                            for i in range( 1, len( row ) ):
                                self.list_stock_list_column_width.append( int( row[ i ] ) )

    def process_all_trading_data( self ): 
        for key_account_name, value_dict_per_company_trading_data in self.dict_all_account_all_stock_trading_data.items():
            for key_stock_name, value_list_trading_data in value_dict_per_company_trading_data.items():
                self.process_single_trading_data( key_account_name, key_stock_name )
    
    def process_single_trading_data( self, str_tab_widget_name, str_stock_number ): 
        list_trading_data = self.dict_all_account_all_stock_trading_data[ str_tab_widget_name ][ str_stock_number ].copy()
        #先拔掉所有的AUTO_DIVIDEND_DATA，下面有需要再重新插入，避免重複插入
        list_trading_data = [item for item in list_trading_data if TradingData.IS_AUTO_DIVIDEND_DATA_NON_SAVE not in item or not item[TradingData.IS_AUTO_DIVIDEND_DATA_NON_SAVE]]

        str_stock_name = self.dict_all_company_number_to_name_and_type[ str_stock_number ][ 0 ]
        b_bond = True if '債' in str_stock_name else False
        b_KY = True if 'KY' in str_stock_name else False
        b_extra_insurance_fee = self.dict_all_account_ui_state[ str_tab_widget_name ][ "insurance_checkbox"]
        if b_KY:
            b_extra_insurance_fee = False

        str_b_etf = self.dict_all_company_number_to_name_and_type[ str_stock_number ][ 1 ]
        b_etf = True if str_b_etf == "True" else False
        sorted_list = sorted( list_trading_data, key=lambda x: ( datetime.datetime.strptime( x[ TradingData.TRADING_DATE ], "%Y-%m-%d"), -x[ TradingData.TRADING_TYPE ] ) )

        str_current_date = datetime.datetime.today().strftime("%Y-%m-%d")
        b_use_auto_dividend = sorted_list[ 0 ][ TradingData.USE_AUTO_DIVIDEND_DATA ]
        if b_use_auto_dividend:
            if str_stock_number in self.dict_auto_stock_yearly_dividned:
                auto_list_dividend = copy.deepcopy( self.dict_auto_stock_yearly_dividned[ str_stock_number ] ) 
                auto_list_dividend = sorted( auto_list_dividend, key=lambda x: ( datetime.datetime.strptime( x[ TradingData.TRADING_DATE ], "%Y-%m-%d") ) )
                if len( sorted_list ) > 1:
                    first_data = sorted_list[ 1 ]
                    for index, auto_dividend_data in enumerate( auto_list_dividend ):
                        if auto_dividend_data[ TradingData.TRADING_DATE ] > first_data[ TradingData.TRADING_DATE ]:
                            if auto_dividend_data[ TradingData.TRADING_DATE ] > str_current_date:
                                break
                            sorted_list.append( auto_dividend_data )
                    sorted_list = sorted( sorted_list, key=lambda x: ( datetime.datetime.strptime( x[ TradingData.TRADING_DATE ], "%Y-%m-%d"), -x[ TradingData.TRADING_TYPE ] ) )

        n_accumulated_inventory = 0
        n_accumulated_cost = 0
        n_accumulated_stock_dividend = 0
        n_accumulated_cash_dividend = 0
        str_last_buying_date = ''
        n_last_buying_count = 0
        list_calibration_data = [] #因為若是已經沒有庫存股票，那麼股利分配或是減資的資料就不會被計算
        for index, item in enumerate( sorted_list ):
            item[ TradingData.SORTED_INDEX_NON_SAVE ] = index
            e_trading_type = item[ TradingData.TRADING_TYPE ]
            
            if e_trading_type == TradingType.TEMPLATE:
                list_calibration_data.append( item )
                continue
            elif e_trading_type == TradingType.BUY:
                f_trading_price = item[ TradingData.TRADING_PRICE ]
                n_trading_count = item[ TradingData.TRADING_COUNT ]
                f_trading_fee_discount = item[ TradingData.TRADING_FEE_DISCOUNT ]
                dict_result = Utility.compute_cost( e_trading_type, f_trading_price, n_trading_count, f_trading_fee_discount, b_etf, False, b_bond )
                item[ TradingData.TRADING_VALUE_NON_SAVE ] = dict_result[ TradingCost.TRADING_VALUE ]
                item[ TradingData.TRADING_FEE_NON_SAVE ] = dict_result[ TradingCost.TRADING_FEE ]
                item[ TradingData.TRADING_TAX_NON_SAVE ] = dict_result[ TradingCost.TRADING_TAX ]
                item[ TradingData.EXTRA_INSURANCE_FEE_NON_SAVE ] = 0 
                n_per_trading_total_cost = item[ TradingData.TRADING_COST_NON_SAVE ] = dict_result[ TradingCost.TRADING_TOTAL_COST ]
                n_accumulated_inventory += n_trading_count
                n_accumulated_cost += n_per_trading_total_cost

                item[ TradingData.STOCK_DIVIDEND_GAIN_NON_SAVE ] = 0
                item[ TradingData.CASH_DIVIDEND_GAIN_NON_SAVE ] = 0
                str_buying_date = item[ TradingData.TRADING_DATE ]
                if str_last_buying_date == str_buying_date:
                    if item[ TradingData.DAYING_TRADING ] == True:
                        n_last_buying_count += n_trading_count
                else:
                    if item[ TradingData.DAYING_TRADING ] == True:
                        str_last_buying_date = str_buying_date
                        n_last_buying_count = n_trading_count
            elif e_trading_type == TradingType.REGULAR_BUY:
                f_trading_price = item[ TradingData.TRADING_PRICE ]
                n_trading_count = item[ TradingData.TRADING_COUNT ]
                e_trading_fee_type = item[ TradingData.TRADING_FEE_TYPE ]
                f_trading_fee_discount = Decimal( str( item[ TradingData.TRADING_FEE_DISCOUNT ] ) )
                n_trading_fee_minimum = item[ TradingData.TRADING_FEE_MINIMUM ]
                n_trading_fee_constant = item[ TradingData.TRADING_FEE_CONSTANT ]
                n_trading_value = int( f_trading_price * n_trading_count )
                if e_trading_fee_type == TradingFeeType.VARIABLE:
                    n_trading_fee = int( n_trading_value * Decimal( '0.001425' ) * f_trading_fee_discount )
                    n_trading_fee = max( n_trading_fee_minimum, n_trading_fee )
                else:
                    n_trading_fee = int( n_trading_fee_constant )

                if n_trading_value == 0:
                    n_trading_fee = 0

                n_trading_total_cost = n_trading_value + n_trading_fee

                item[ TradingData.TRADING_VALUE_NON_SAVE ] = n_trading_value
                item[ TradingData.TRADING_FEE_NON_SAVE ] = n_trading_fee
                item[ TradingData.TRADING_TAX_NON_SAVE ] = 0
                item[ TradingData.EXTRA_INSURANCE_FEE_NON_SAVE ] = 0 
                item[ TradingData.TRADING_COST_NON_SAVE ] = n_trading_total_cost
                n_accumulated_inventory += n_trading_count
                n_accumulated_cost += n_trading_total_cost

                item[ TradingData.STOCK_DIVIDEND_GAIN_NON_SAVE ] = 0
                item[ TradingData.CASH_DIVIDEND_GAIN_NON_SAVE ] = 0
            elif e_trading_type == TradingType.SELL:
                str_selling_date = item[ TradingData.TRADING_DATE ]
                obj_selling_date = datetime.datetime.strptime( str_selling_date, "%Y-%m-%d")
                f_trading_price = item[ TradingData.TRADING_PRICE ]
                f_trading_fee_discount = item[ TradingData.TRADING_FEE_DISCOUNT ]
                n_trading_count = item[ TradingData.TRADING_COUNT ]

                if ( str_selling_date == str_last_buying_date and    #賣出與買入同一天
                     item[ TradingData.DAYING_TRADING ] == True and  #有勾選是當沖交易
                     obj_selling_date >= datetime.datetime.strptime( '2017-04-28', "%Y-%m-%d" ) ): #交易日期在2017-04-28之後。因為在這之後才通過當沖交易稅減半
                    item[ TradingData.IS_REALLY_DAYING_TRADING_NON_SAVE ] = True
                    if n_trading_count <= n_last_buying_count: #賣出數量小於或等於買入數量，表示全部賣出數量都可視為當沖
                        dict_result = Utility.compute_cost( e_trading_type, f_trading_price, n_trading_count, f_trading_fee_discount, b_etf, True, b_bond )
                        item[ TradingData.TRADING_VALUE_NON_SAVE ] = dict_result[ TradingCost.TRADING_VALUE ]
                        item[ TradingData.TRADING_FEE_NON_SAVE ] = dict_result[ TradingCost.TRADING_FEE ]
                        item[ TradingData.TRADING_TAX_NON_SAVE ] = dict_result[ TradingCost.TRADING_TAX ]
                        n_per_trading_total_cost = item[ TradingData.TRADING_COST_NON_SAVE ] = dict_result[ TradingCost.TRADING_TOTAL_COST ]
                        n_accumulated_cost -= n_per_trading_total_cost
                        n_accumulated_inventory -= n_trading_count
                        n_last_buying_count -= n_trading_count
                    else: #賣出數量大於買入數量，表示只有部分數量都可視為當沖
                        n_trading_count_1 = n_last_buying_count
                        dict_result = Utility.compute_cost( e_trading_type, f_trading_price, n_trading_count_1, f_trading_fee_discount, b_etf, True, b_bond )#這部分是當沖
                        n_trading_value_1 = dict_result[ TradingCost.TRADING_VALUE ]
                        n_trading_fee_1 = dict_result[ TradingCost.TRADING_FEE ]
                        n_trading_tax_1 = dict_result[ TradingCost.TRADING_TAX ]
                        n_trading_total_cost_1 = dict_result[ TradingCost.TRADING_TOTAL_COST ]

                        n_trading_count_2 = n_trading_count - n_last_buying_count
                        dict_result = Utility.compute_cost( e_trading_type, f_trading_price, n_trading_count_2, f_trading_fee_discount, b_etf, False, b_bond )#這部分不是當沖
                        n_trading_value_2 = dict_result[ TradingCost.TRADING_VALUE ]
                        n_trading_fee_2 = dict_result[ TradingCost.TRADING_FEE ]
                        n_trading_tax_2 = dict_result[ TradingCost.TRADING_TAX ]
                        n_trading_total_cost_2 = dict_result[ TradingCost.TRADING_TOTAL_COST ]

                        item[ TradingData.TRADING_VALUE_NON_SAVE ] = n_trading_value_1 + n_trading_value_2
                        item[ TradingData.TRADING_FEE_NON_SAVE ] = n_trading_fee_1 + n_trading_fee_2
                        item[ TradingData.TRADING_TAX_NON_SAVE ] = n_trading_tax_1 + n_trading_tax_2
                        n_per_trading_total_cost = item[ TradingData.TRADING_COST_NON_SAVE ] = n_trading_total_cost_1 + n_trading_total_cost_2
                        n_accumulated_cost -= n_per_trading_total_cost
                        n_accumulated_inventory -= n_trading_count

                        n_last_buying_count = 0
                else:
                    item[ TradingData.IS_REALLY_DAYING_TRADING_NON_SAVE ] = False
                    dict_result = Utility.compute_cost( e_trading_type, f_trading_price, n_trading_count, f_trading_fee_discount, b_etf, False, b_bond )
                    item[ TradingData.TRADING_VALUE_NON_SAVE ] = dict_result[ TradingCost.TRADING_VALUE ]
                    item[ TradingData.TRADING_FEE_NON_SAVE ] = dict_result[ TradingCost.TRADING_FEE ]
                    item[ TradingData.TRADING_TAX_NON_SAVE ] = dict_result[ TradingCost.TRADING_TAX ]
                    n_per_trading_total_cost = item[ TradingData.TRADING_COST_NON_SAVE ] = dict_result[ TradingCost.TRADING_TOTAL_COST ]
                    n_accumulated_cost -= n_per_trading_total_cost
                    n_accumulated_inventory -= n_trading_count

                item[ TradingData.EXTRA_INSURANCE_FEE_NON_SAVE ] = 0
                item[ TradingData.STOCK_DIVIDEND_GAIN_NON_SAVE ] = 0
                item[ TradingData.CASH_DIVIDEND_GAIN_NON_SAVE ] = 0
            elif e_trading_type == TradingType.CAPITAL_INCREASE:
                f_trading_price = item[ TradingData.TRADING_PRICE ]
                n_trading_count = item[ TradingData.TRADING_COUNT ]
                f_trading_fee_discount = item[ TradingData.TRADING_FEE_DISCOUNT ]
                
                n_per_trading_total_cost = item[ TradingData.TRADING_COST_NON_SAVE ] = int( f_trading_price * n_trading_count )
                item[ TradingData.TRADING_VALUE_NON_SAVE ] = n_per_trading_total_cost
                item[ TradingData.TRADING_FEE_NON_SAVE ] = 0
                item[ TradingData.TRADING_TAX_NON_SAVE ] = 0
                item[ TradingData.EXTRA_INSURANCE_FEE_NON_SAVE ] = 0 
                n_accumulated_inventory += n_trading_count
                n_accumulated_cost += n_per_trading_total_cost

                item[ TradingData.STOCK_DIVIDEND_GAIN_NON_SAVE ] = 0
                item[ TradingData.CASH_DIVIDEND_GAIN_NON_SAVE ] = 0
                str_buying_date = item[ TradingData.TRADING_DATE ]
            elif e_trading_type == TradingType.DIVIDEND:
                if n_accumulated_inventory <= 0: #沒有庫存就不用算股利了
                    continue
                n_stock_dividend_value_gain = int( Decimal( str( item[ TradingData.STOCK_DIVIDEND_PER_SHARE ] ) ) * Decimal( str( n_accumulated_inventory ) ) ) # n_stock_dividend_value_gain單位為元
                n_stock_dividend_share_gain = int( Decimal( str( item[ TradingData.STOCK_DIVIDEND_PER_SHARE ] ) ) * Decimal( str( n_accumulated_inventory ) ) / Decimal( '10' ) ) # n_stock_dividend_share_gain單位為股 除以10是因為票面額10元
                n_cash_dividend_gain = int( Decimal( str(item[ TradingData.CASH_DIVIDEND_PER_SHARE ] ) ) * Decimal( str( n_accumulated_inventory ) ) )

                if b_use_auto_dividend:
                    if TradingData.IS_AUTO_DIVIDEND_DATA_NON_SAVE not in item or not item[TradingData.IS_AUTO_DIVIDEND_DATA_NON_SAVE]:
                        n_stock_dividend_share_gain = 0
                        n_stock_dividend_value_gain = 0
                        n_cash_dividend_gain = 0
                else:
                    if TradingData.IS_AUTO_DIVIDEND_DATA_NON_SAVE in item and item[TradingData.IS_AUTO_DIVIDEND_DATA_NON_SAVE]:
                        n_stock_dividend_share_gain = 0
                        n_stock_dividend_value_gain = 0
                        n_cash_dividend_gain = 0

                item[ TradingData.TRADING_VALUE_NON_SAVE ] = 0
                item[ TradingData.TRADING_TAX_NON_SAVE ] = 0
                item[ TradingData.TRADING_COST_NON_SAVE ] = 0

                item[ TradingData.STOCK_DIVIDEND_GAIN_NON_SAVE ] = n_stock_dividend_share_gain
                n_accumulated_inventory += n_stock_dividend_share_gain
                
                n_extra_insurance_fee_for_cash_dividend = 0
                n_extra_insurance_fee_for_stock_dividend = 0
                if b_extra_insurance_fee and n_cash_dividend_gain + n_stock_dividend_value_gain > 20000:
                    obj_dividend_date = datetime.datetime.strptime( item[ TradingData.TRADING_DATE ], "%Y-%m-%d")
                    if obj_dividend_date.year >= 2013 and obj_dividend_date.year < 2021:
                        n_extra_insurance_fee_for_cash_dividend = int( math.ceil( Decimal( str( n_cash_dividend_gain ) ) * Decimal( str( '0.0191' ) ) ) )
                        n_extra_insurance_fee_for_stock_dividend = int( math.ceil( Decimal( str( n_stock_dividend_value_gain ) ) * Decimal( str( '0.0191' ) ) ) )
                    elif obj_dividend_date.year >= 2021:
                        n_extra_insurance_fee_for_cash_dividend = int( math.ceil( Decimal( str( n_cash_dividend_gain ) ) * Decimal( str( '0.0211' ) ) ) )
                        n_extra_insurance_fee_for_stock_dividend = int( math.ceil( Decimal( str( n_stock_dividend_value_gain ) ) * Decimal( str( '0.0211' ) ) ) )

                if n_cash_dividend_gain > 10:
                    item[ TradingData.CASH_DIVIDEND_GAIN_NON_SAVE ] = n_cash_dividend_gain
                    item[ TradingData.TRADING_FEE_NON_SAVE ] = 10

                    item[ TradingData.EXTRA_INSURANCE_FEE_NON_SAVE ] = n_extra_insurance_fee_for_cash_dividend + n_extra_insurance_fee_for_stock_dividend
                    n_accumulated_cost = n_accumulated_cost - n_cash_dividend_gain + 10 + n_extra_insurance_fee_for_cash_dividend + n_extra_insurance_fee_for_stock_dividend
                    n_accumulated_cash_dividend = n_accumulated_cash_dividend + n_cash_dividend_gain - 10 - n_extra_insurance_fee_for_cash_dividend - n_extra_insurance_fee_for_stock_dividend #要想一下要不要扣掉手續費
                else:
                    item[ TradingData.CASH_DIVIDEND_GAIN_NON_SAVE ] = 0
                    item[ TradingData.TRADING_FEE_NON_SAVE ] = 0
                    item[ TradingData.EXTRA_INSURANCE_FEE_NON_SAVE ] = 0 
                n_accumulated_stock_dividend += n_stock_dividend_share_gain
            elif e_trading_type == TradingType.CAPITAL_REDUCTION:
                if n_accumulated_inventory == 0: #沒有庫存就不用算減資了
                    continue
                item[ TradingData.TRADING_PRICE ] = -item[ TradingData.CAPITAL_REDUCTION_PER_SHARE ]
                item[ TradingData.TRADING_COUNT ] = n_accumulated_inventory
                item[ TradingData.TRADING_VALUE_NON_SAVE ] = -int( n_accumulated_inventory * item[ TradingData.CAPITAL_REDUCTION_PER_SHARE ] )
                item[ TradingData.TRADING_FEE_NON_SAVE ] = 0
                item[ TradingData.TRADING_TAX_NON_SAVE ] = 0
                item[ TradingData.EXTRA_INSURANCE_FEE_NON_SAVE ] = 0 
                item[ TradingData.TRADING_COST_NON_SAVE ] = 0
                item[ TradingData.STOCK_DIVIDEND_GAIN_NON_SAVE ] = 0
                item[ TradingData.CASH_DIVIDEND_GAIN_NON_SAVE ] = 0
                n_accumulated_cost = n_accumulated_cost - int( Decimal( str( n_accumulated_inventory ) ) * Decimal( str( item[ TradingData.CAPITAL_REDUCTION_PER_SHARE ] ) ) )
                n_accumulated_inventory = int( Decimal( str( n_accumulated_inventory ) ) * ( Decimal( str( '10' ) ) - Decimal( str( item[ TradingData.CAPITAL_REDUCTION_PER_SHARE ] ) ) ) / Decimal( str( '10' ) ) )

            item[ TradingData.ACCUMULATED_COST_NON_SAVE ] = n_accumulated_cost
            item[ TradingData.ACCUMULATED_INVENTORY_NON_SAVE ] = n_accumulated_inventory
            item[ TradingData.AVERAGE_COST_NON_SAVE ] = n_accumulated_cost / n_accumulated_inventory if n_accumulated_inventory != 0 else 0
            item[ TradingData.ALL_STOCK_DIVIDEND_GAIN_NON_SAVE ] = n_accumulated_stock_dividend
            item[ TradingData.ALL_CASH_DIVIDEND_GAIN_NON_SAVE ] = n_accumulated_cash_dividend
            list_calibration_data.append( item )

        self.dict_all_account_all_stock_trading_data[ str_tab_widget_name ][ str_stock_number ] = list_calibration_data.copy()
        return list_calibration_data
    
    def process_all_transfer_data( self ): 
        for key_account_name, value_dict_per_company_transfer_data in self.dict_all_account_cash_transfer_data.items():
            self.process_single_transfer_data( key_account_name )

    def process_single_transfer_data( self, str_tab_widget_name ):
        list_transfer_data = self.dict_all_account_cash_transfer_data[ str_tab_widget_name ]
        if len( list_transfer_data ) == 0:
            return
        sorted_list = sorted( list_transfer_data, key=lambda x: ( datetime.datetime.strptime( x[ TransferData.TRANSFER_DATE ], "%Y-%m-%d") ) )

        n_total_value = 0
        for index, item in enumerate( sorted_list ):
            item[ TransferData.SORTED_INDEX_NON_SAVE ] = index
            e_transfer_type = item[ TransferData.TRANSFER_TYPE ]
            if e_transfer_type == TransferType.TRANSFER_IN:
                n_total_value += item[ TransferData.TRANSFER_VALUE ]
            elif e_transfer_type == TransferType.TRANSFER_OUT:
                n_total_value -= item[ TransferData.TRANSFER_VALUE ]
            
            item[ TransferData.TOTAL_VALUE_NON_SAVE ] = n_total_value

        self.dict_all_account_cash_transfer_data[ str_tab_widget_name ] = sorted_list

    def refresh_stock_list_table( self ): 
        str_tab_widget_name = self.ui.qtTabWidget.currentWidget().objectName()
        dict_per_account_all_stock_trading_data = self.dict_all_account_all_stock_trading_data[ str_tab_widget_name ]
        table_view = self.ui.qtTabWidget.currentWidget().findChild( QTableView, "StockListTableView" )
        display_type_combobox = self.ui.qtTabWidget.currentWidget().findChild( QComboBox, "DisplayTypeComboBox")
        n_total_profit = 0
        n_total_inventory = 0
        if table_view:
            table_model = table_view.model()
            table_model.clear()

            if self.ui.qtUse1ShareUnitAction.isChecked():
                self.list_stock_list_table_horizontal_header[ 1 ] = "庫存股數"
            else:
                self.list_stock_list_table_horizontal_header[ 1 ] = "庫存張數"
            table_model.setHorizontalHeaderLabels( self.list_stock_list_table_horizontal_header )
            
            with QSignalBlocker( table_view.horizontalHeader() ):
                list_vertical_labels = []
                index_row = 0
                for key_stock_number, value_list_stock_trading_data in dict_per_account_all_stock_trading_data.items():
                    dict_trading_data_first = value_list_stock_trading_data[ 0 ] #取第一筆交易資料，因為第一筆交易資料有存是否使用自動帶入股利
                    dict_trading_data_last = value_list_stock_trading_data[ len( value_list_stock_trading_data ) - 1 ] #取最後一筆交易資料，因為最後一筆交易資料的庫存等內容才是所有累計的結果

                    n_accumulated_inventory = 0
                    if TradingData.ACCUMULATED_INVENTORY_NON_SAVE in dict_trading_data_last:
                        n_accumulated_inventory = dict_trading_data_last[ TradingData.ACCUMULATED_INVENTORY_NON_SAVE ]
                    if display_type_combobox.currentIndex() == 1:
                        if n_accumulated_inventory == 0:
                            continue
                    elif display_type_combobox.currentIndex() == 2:
                        if n_accumulated_inventory != 0:
                            continue

                    if self.ui.qtUse1ShareUnitAction.isChecked():
                        str_accumulated_inventory = format( n_accumulated_inventory, "," )
                    else:
                        f_accumulated_inventory_1000_share = n_accumulated_inventory / 1000
                        if f_accumulated_inventory_1000_share.is_integer():
                            str_accumulated_inventory = format( int( f_accumulated_inventory_1000_share ), "," )
                        else:
                            str_accumulated_inventory = format( f_accumulated_inventory_1000_share, "," )


                    list_stock_name_and_type = self.dict_all_company_number_to_name_and_type[ key_stock_number ]
                    str_stock_name = list_stock_name_and_type[ 0 ]
                    list_vertical_labels.append( f"{key_stock_number} {str_stock_name}" )
                    n_total_trading_fee = 0
                    n_total_trading_tax = 0
                    for index, item_trading_data in enumerate( value_list_stock_trading_data ):
                        if index == 0:
                            continue
                        n_total_trading_fee += item_trading_data[ TradingData.TRADING_FEE_NON_SAVE ]
                        n_total_trading_tax += item_trading_data[ TradingData.TRADING_TAX_NON_SAVE ]

                    n_accumulated_cost = 0
                    if TradingData.ACCUMULATED_COST_NON_SAVE in dict_trading_data_last:
                        n_accumulated_cost = dict_trading_data_last[ TradingData.ACCUMULATED_COST_NON_SAVE ]
                    f_average_cost = 0
                    if TradingData.AVERAGE_COST_NON_SAVE in dict_trading_data_last:
                        f_average_cost = round( dict_trading_data_last[ TradingData.AVERAGE_COST_NON_SAVE ], 3 )
                    n_accumulated_stock_dividend = 0 
                    if TradingData.ALL_STOCK_DIVIDEND_GAIN_NON_SAVE in dict_trading_data_last:
                        n_accumulated_stock_dividend = dict_trading_data_last[ TradingData.ALL_STOCK_DIVIDEND_GAIN_NON_SAVE ]
                    n_accumulated_cash_dividend = 0
                    if TradingData.ALL_CASH_DIVIDEND_GAIN_NON_SAVE in dict_trading_data_last:
                        n_accumulated_cash_dividend = dict_trading_data_last[ TradingData.ALL_CASH_DIVIDEND_GAIN_NON_SAVE ]

                    n_accumulated_dividend_profit = 0
                    if key_stock_number in self.dict_all_company_number_to_price_info:
                        try:
                            f_stock_price = float( self.dict_all_company_number_to_price_info[ key_stock_number ] )
                            str_stock_price = format( f_stock_price, "," )
                            n_net_value = int( n_accumulated_inventory * f_stock_price )
                            str_net_value = format( n_net_value, "," )
                            n_profit = n_net_value - n_accumulated_cost
                            n_total_profit += n_profit
                            n_total_inventory += n_net_value
                            n_accumulated_dividend_profit = n_accumulated_stock_dividend * f_stock_price + n_accumulated_cash_dividend
                            str_profit = format( n_profit, "," )
                            if n_profit > 0:
                                str_color = QBrush( '#FF0000' )
                            elif n_profit < 0:
                                str_color = QBrush( '#00AA00' )
                            else:
                                str_color = QBrush( '#FFFFFF' )
                        except ValueError:
                            str_stock_price = "N/A"
                            str_net_value = "N/A"
                            str_profit = "N/A"
                            str_color = QBrush( '#FFFFFF' )
                    else:
                        str_stock_price = "N/A"
                        str_net_value = "N/A"
                        str_profit = "N/A"
                        str_color = QBrush( '#FFFFFF' )

                    list_data = [ format( n_accumulated_cost, "," ),      #總成本
                                  str_accumulated_inventory,              #庫存股數
                                  format( f_average_cost, "," ),          #平均成本
                                  str_stock_price,                        #當前股價
                                  str_net_value,                          #淨值
                                  format( n_total_trading_fee, ","),      #總手續費
                                  format( n_total_trading_tax, ","),      #總交易稅
                                  str_profit,                             #損益
                                  format( n_accumulated_dividend_profit, ",") ] #股利所得
                                    
                    for column, data in enumerate( list_data ):
                        standard_item = QStandardItem( data )
                        standard_item.setTextAlignment( Qt.AlignHCenter | Qt.AlignVCenter )
                        standard_item.setFlags( standard_item.flags() & ~Qt.ItemIsEditable )
                        standard_item.setData( key_stock_number, Qt.UserRole )
                        if column == len( list_data ) - 2:
                            standard_item.setForeground( QBrush( str_color ) )
                        table_model.setItem( index_row, column, standard_item ) 

                    use_auto_dividend_item = QStandardItem()
                    if dict_trading_data_first[ TradingData.USE_AUTO_DIVIDEND_DATA ]:
                        use_auto_dividend_item.setIcon( check_icon )
                    else:
                        use_auto_dividend_item.setIcon( uncheck_icon )
                    use_auto_dividend_item.setFlags( use_auto_dividend_item.flags() & ~Qt.ItemIsEditable )
                    table_model.setItem( index_row, len( self.list_stock_list_table_horizontal_header ) - 3, use_auto_dividend_item)
                        
                    export_icon_item = QStandardItem("")
                    export_icon_item.setIcon( export_icon )
                    export_icon_item.setFlags( export_icon_item.flags() & ~Qt.ItemIsEditable )
                    table_model.setItem( index_row, len( self.list_stock_list_table_horizontal_header ) - 2, export_icon_item )
                    delete_icon_item = QStandardItem("")
                    delete_icon_item.setIcon( delete_icon )
                    delete_icon_item.setFlags( delete_icon_item.flags() & ~Qt.ItemIsEditable )
                    table_model.setItem( index_row, len( self.list_stock_list_table_horizontal_header ) - 1, delete_icon_item )
                    index_row += 1

                for column in range( len( self.list_stock_list_table_horizontal_header ) ):
                    if column < len( self.list_stock_list_column_width ):
                        table_view.setColumnWidth( column, self.list_stock_list_column_width[ column ] )
                    else:
                        table_view.setColumnWidth( column, 100 )
                        self.list_stock_list_column_width.append( 100 )

                for index_row in range( len( dict_per_account_all_stock_trading_data ) ):
                    table_view.setRowHeight( index_row, 25 )
                table_model.setVerticalHeaderLabels( list_vertical_labels )
        total_profit_value_label = self.ui.qtTabWidget.currentWidget().findChild( QLabel, "TotalProfitValueLabel")
        total_profit_value_label.setText( format( n_total_profit, "," ) )
        total_inventory_value_label = self.ui.qtTabWidget.currentWidget().findChild( QLabel, "TotalInventoryValueLabel")
        total_inventory_value_label.setText( format( n_total_inventory, "," ) )

        #以下為計算年化報酬率
        list_total_trading_date = []
        list_total_trading_flows = []
        for key_stock_number, value_list_trading_data in dict_per_account_all_stock_trading_data.items():
            dict_trading_data_last = value_list_trading_data[ len( value_list_trading_data ) - 1 ] #取最後一筆交易資料，因為最後一筆交易資料的庫存等內容才是所有累計的結果
            if key_stock_number not in self.dict_all_company_number_to_price_info:
                continue

            n_accumulated_inventory = 0
            if TradingData.ACCUMULATED_INVENTORY_NON_SAVE in dict_trading_data_last:
                n_accumulated_inventory = dict_trading_data_last[ TradingData.ACCUMULATED_INVENTORY_NON_SAVE ]
            if display_type_combobox.currentIndex() == 1:
                if n_accumulated_inventory == 0:
                    continue
            elif display_type_combobox.currentIndex() == 2:
                if n_accumulated_inventory != 0:
                    continue


            for trading_data in value_list_trading_data:
                e_trading_type = trading_data[ TradingData.TRADING_TYPE ]
                if e_trading_type == TradingType.TEMPLATE:
                    continue
                if ( e_trading_type == TradingType.BUY or
                     e_trading_type == TradingType.REGULAR_BUY or 
                     e_trading_type == TradingType.CAPITAL_INCREASE ):
                    n_cost = -trading_data[ TradingData.TRADING_COST_NON_SAVE ]
                elif e_trading_type == TradingType.SELL:
                    n_cost = trading_data[ TradingData.TRADING_COST_NON_SAVE ]
                elif e_trading_type == TradingType.DIVIDEND:
                    n_cost = trading_data[ TradingData.CASH_DIVIDEND_GAIN_NON_SAVE ] - trading_data[ TradingData.TRADING_FEE_NON_SAVE ] - trading_data[ TradingData.EXTRA_INSURANCE_FEE_NON_SAVE ] 
                elif e_trading_type == TradingType.CAPITAL_REDUCTION:
                    #trading_data[ TradingData.TRADING_VALUE_NON_SAVE ]本身是負值，但因為是退款，所以在算XIRR時應該要用正數，因此取負號
                    n_cost = -trading_data[ TradingData.TRADING_VALUE_NON_SAVE ] 
                list_total_trading_flows.append( n_cost )

                str_trading_date = trading_data[ TradingData.TRADING_DATE ]
                obj_trading_date = datetime.datetime.strptime( str_trading_date, "%Y-%m-%d" )
                list_total_trading_date.append( obj_trading_date )

        list_total_trading_flows.append( n_total_inventory )
        obj_current_date = datetime.datetime.today()
        list_total_trading_date.append( obj_current_date )
        if len( list_total_trading_flows ) > 1:
            try:
                xirr_result = Utility.xirr( list_total_trading_flows, list_total_trading_date )
                xirr_value_label = self.ui.qtTabWidget.currentWidget().findChild( QLabel, "XIRRValueLabel")
                xirr_value_label.setText( f"{xirr_result:.3%}" )
            except ValueError as e:
                pass
    
    def refresh_transfer_data_table( self ):
        str_tab_widget_name = self.ui.qtTabWidget.currentWidget().objectName()
        list_per_account_all_cash_transfer_data = self.dict_all_account_cash_transfer_data[ str_tab_widget_name ]

        table_view = self.ui.qtTabWidget.currentWidget().findChild( QTableView, "CashTransferTableView" )
        if table_view:
            table_model = table_view.model()
            table_model.clear()
            table_model.setVerticalHeaderLabels( self.get_cash_transfer_header() )

            if self.ui.qtFromNewToOldAction.isChecked():
                loop_list = list_per_account_all_cash_transfer_data[::-1]
            else:
                loop_list = list_per_account_all_cash_transfer_data

            for column, item in enumerate( loop_list ):
                str_date = item[ TransferData.TRANSFER_DATE ]
                str_year = str_date.split( '-' )[ 0 ]
                if self.ui.qtROCYearAction.isChecked():
                    str_year = str( int( str_year ) - 1911 )

                str_month_date = str_date[ 5: ].replace( '-', '/' )
                str_date = f"{str_year}/{str_month_date}"


                str_total_value = format( item[ TransferData.TOTAL_VALUE_NON_SAVE ], "," )

                e_transfer_type = item[ TransferData.TRANSFER_TYPE ]
                if e_transfer_type == TransferType.TRANSFER_IN:
                    str_color = QBrush( '#FF0000' )
                    str_transfer_value = format( item[ TransferData.TRANSFER_VALUE ], "," )
                else:
                    str_color = QBrush( '#00AA00' )
                    str_transfer_value = format( -item[ TransferData.TRANSFER_VALUE ], "," )

                date_standard_item = QStandardItem( str_date )
                date_standard_item.setTextAlignment( Qt.AlignHCenter | Qt.AlignVCenter )
                date_standard_item.setFlags( date_standard_item.flags() & ~Qt.ItemIsEditable )
                table_model.setItem( 0, column, date_standard_item )

                transfer_value_standard_item = QStandardItem( str_transfer_value )
                transfer_value_standard_item.setTextAlignment( Qt.AlignHCenter | Qt.AlignVCenter )
                transfer_value_standard_item.setFlags( transfer_value_standard_item.flags() & ~Qt.ItemIsEditable )
                transfer_value_standard_item.setForeground( QBrush( str_color ) )
                table_model.setItem( 1, column, transfer_value_standard_item )

                total_value_standard_item = QStandardItem( str_total_value )
                total_value_standard_item.setTextAlignment( Qt.AlignHCenter | Qt.AlignVCenter )
                total_value_standard_item.setFlags( total_value_standard_item.flags() & ~Qt.ItemIsEditable )
                table_model.setItem( 2, column, total_value_standard_item )

                edit_icon_item = QStandardItem("")
                edit_icon_item.setIcon( edit_icon )
                edit_icon_item.setFlags( edit_icon_item.flags() & ~Qt.ItemIsEditable )
                edit_icon_item.setData( item[ TransferData.SORTED_INDEX_NON_SAVE ], Qt.UserRole )
                delete_icon_item = QStandardItem("")
                delete_icon_item.setIcon( delete_icon )
                delete_icon_item.setFlags( delete_icon_item.flags() & ~Qt.ItemIsEditable )
                delete_icon_item.setData( item[ TransferData.SORTED_INDEX_NON_SAVE ], Qt.UserRole )

                table_model.setItem( 3, column, edit_icon_item )
                table_model.setItem( 4, column, delete_icon_item )

                if self.ui.qtShow10Action.isChecked():
                    if column == 9:
                        break

    def clear_per_account_transfer_table( self ):
        self.per_stock_trading_data_model.clear()
        self.per_stock_trading_data_model.setVerticalHeaderLabels( self.get_trading_data_header() )

    def refresh_trading_data_table( self, sorted_list ):
        self.clear_per_stock_trading_table()
        self.per_stock_trading_data_model.setVerticalHeaderLabels( self.get_trading_data_header() )
        self.ui.qtTradingDataTableView.horizontalHeader().hide()

        if self.ui.qtFromNewToOldAction.isChecked():
            loop_list = sorted_list[::-1]
        else:
            loop_list = sorted_list

        b_use_auto_dividend = sorted_list[ 0 ][ TradingData.USE_AUTO_DIVIDEND_DATA ]
        column = 0
        for dict_per_trading_data in loop_list:
            e_trading_type = dict_per_trading_data[ TradingData.TRADING_TYPE ]
            if e_trading_type == TradingType.TEMPLATE:
                continue

            if e_trading_type == TradingType.DIVIDEND:
                if b_use_auto_dividend:#使用自動帶入股利資料，但是這筆資料不是自動股利資料，就跳過，反之亦然
                    if TradingData.IS_AUTO_DIVIDEND_DATA_NON_SAVE not in dict_per_trading_data or not dict_per_trading_data[TradingData.IS_AUTO_DIVIDEND_DATA_NON_SAVE]:
                        continue
                else:
                    if TradingData.IS_AUTO_DIVIDEND_DATA_NON_SAVE in dict_per_trading_data and dict_per_trading_data[TradingData.IS_AUTO_DIVIDEND_DATA_NON_SAVE]:
                        continue

            list_data = self.get_per_trading_data_text_list( dict_per_trading_data )

            if TradingData.IS_REALLY_DAYING_TRADING_NON_SAVE in dict_per_trading_data:
                if dict_per_trading_data[ TradingData.IS_REALLY_DAYING_TRADING_NON_SAVE ]:
                    list_data[ 2 ] = "當沖賣"

            for row, data in enumerate( list_data ):
                standard_item = QStandardItem( data )
                if data == "買進":
                    standard_item.setBackground( QBrush( '#DA9694' ) )
                elif data == "定期定額買進":
                    standard_item.setBackground( QBrush( '#FF99FF' ) )
                elif data == "賣出" or data == "當沖賣":
                    standard_item.setBackground( QBrush( '#76933C' ) )
                elif data == "股利分配":
                    standard_item.setBackground( QBrush( '#8DB4E2' ) )
                elif data == "減資":
                    standard_item.setBackground( QBrush( '#B1A0C7' ) )
                elif data == "增資":
                    standard_item.setBackground( QBrush( '#FABF8F' ) )
                standard_item.setTextAlignment( Qt.AlignHCenter | Qt.AlignVCenter )
                standard_item.setFlags( standard_item.flags() & ~Qt.ItemIsEditable )
                self.per_stock_trading_data_model.setItem( row, column, standard_item ) 

            if e_trading_type != TradingType.DIVIDEND or not b_use_auto_dividend:
                edit_icon_item = QStandardItem("")
                edit_icon_item.setIcon( edit_icon )
                edit_icon_item.setFlags( edit_icon_item.flags() & ~Qt.ItemIsEditable )
                edit_icon_item.setData( dict_per_trading_data[ TradingData.SORTED_INDEX_NON_SAVE ], Qt.UserRole )
                delete_icon_item = QStandardItem("")
                delete_icon_item.setIcon( delete_icon )
                delete_icon_item.setFlags( delete_icon_item.flags() & ~Qt.ItemIsEditable )
                delete_icon_item.setData( dict_per_trading_data[ TradingData.SORTED_INDEX_NON_SAVE ], Qt.UserRole )

                self.per_stock_trading_data_model.setItem( len( list_data ), column, edit_icon_item )
                self.per_stock_trading_data_model.setItem( len( list_data ) + 1, column, delete_icon_item )

                if self.ui.qtShow10Action.isChecked():
                    if column == 9:
                        break
            column += 1

        for row in range( len( self.get_trading_data_header() ) ):
            if row == 10 or row == 11:
                self.ui.qtTradingDataTableView.setRowHeight( row, 40 )
            else:
                self.ui.qtTradingDataTableView.setRowHeight( row, 25 )

    def clear_per_stock_trading_table( self ):
        self.per_stock_trading_data_model.clear()
        self.per_stock_trading_data_model.setVerticalHeaderLabels( self.get_trading_data_header() )
        for row in range( len( self.get_trading_data_header() ) ):
            if row == 10 or row == 11:
                self.ui.qtTradingDataTableView.setRowHeight( row, 40 )
            else:
                self.ui.qtTradingDataTableView.setRowHeight( row, 25 )

    def get_cash_transfer_header( self ):
        return [ '日期', '入金/出金', '餘額', '編輯', '刪除' ]

    def get_trading_data_header( self ):
        if self.ui.qtUse1ShareUnitAction.isChecked():
            return ['年度', '日期', '交易種類', '交易價格', '交易股數', '交易金額', '手續費', '交易稅', '補充保費', '單筆總成本', '全部股票股利 /\n每股股票股利', '全部現金股利 /\n每股現金股利',
                    '累計總成本', '庫存股數', '平均成本', '編輯', '刪除' ]
        else:
            return ['年度', '日期', '交易種類', '交易價格', '交易張數', '交易金額', '手續費', '交易稅', '補充保費', '單筆總成本', '全部股票股利 /\n每股股票股利', '全部現金股利 /\n每股現金股利',
                    '累計總成本', '庫存張數', '平均成本', '編輯', '刪除' ]

    def get_per_trading_data_text_list( self, dict_per_trading_data ):
        e_trading_type = dict_per_trading_data[ TradingData.TRADING_TYPE ]
        if e_trading_type == TradingType.TEMPLATE:
            return []
        str_date = dict_per_trading_data[ TradingData.TRADING_DATE ]
        str_year = str_date.split( '-' )[ 0 ]
        if self.ui.qtROCYearAction.isChecked():
            str_year = str( int( str_year ) - 1911 )

        str_month_date = str_date[ 5: ].replace( '-', '/' )
        obj_date = datetime.datetime.strptime( str_date, "%Y-%m-%d" )
        n_weekday = obj_date.weekday()
        if n_weekday == 0:
            str_weekday = "(一)"
        elif n_weekday == 1:
            str_weekday = "(二)"
        elif n_weekday == 2:
            str_weekday = "(三)"
        elif n_weekday == 3:
            str_weekday = "(四)"
        elif n_weekday == 4:
            str_weekday = "(五)"
        elif n_weekday == 5:
            str_weekday = "(六)"
        else:
            str_weekday = "(日)"
        f_trading_price = dict_per_trading_data[ TradingData.TRADING_PRICE ]
        n_trading_count = dict_per_trading_data[ TradingData.TRADING_COUNT ]
        n_trading_value = dict_per_trading_data[ TradingData.TRADING_VALUE_NON_SAVE ]
        n_trading_fee = dict_per_trading_data[ TradingData.TRADING_FEE_NON_SAVE ]
        n_trading_tax = dict_per_trading_data[ TradingData.TRADING_TAX_NON_SAVE ]
        n_extra_insurance_fee = dict_per_trading_data[ TradingData.EXTRA_INSURANCE_FEE_NON_SAVE ]
        n_per_trading_total_cost = dict_per_trading_data[ TradingData.TRADING_COST_NON_SAVE ]
        f_stock_dividend_per_share = dict_per_trading_data[ TradingData.STOCK_DIVIDEND_PER_SHARE ]
        f_cash_dividend_per_share = dict_per_trading_data[ TradingData.CASH_DIVIDEND_PER_SHARE ]
        n_stock_dividend_gain = dict_per_trading_data[ TradingData.STOCK_DIVIDEND_GAIN_NON_SAVE ]
        n_cash_dividend_gain = dict_per_trading_data[ TradingData.CASH_DIVIDEND_GAIN_NON_SAVE ]
        n_accumulated_cost = dict_per_trading_data[ TradingData.ACCUMULATED_COST_NON_SAVE ]
        n_accumulated_inventory = dict_per_trading_data[ TradingData.ACCUMULATED_INVENTORY_NON_SAVE ]
        f_average_cost = round( dict_per_trading_data[ TradingData.AVERAGE_COST_NON_SAVE ], 3 )
        if self.ui.qtUse1ShareUnitAction.isChecked():
            str_trading_count = format( n_trading_count, "," )
            str_stock_dividend_gain = format( n_stock_dividend_gain, "," )
            str_accumulated_inventory = format( n_accumulated_inventory, "," )
        else:
            f_trading_count = n_trading_count / 1000
            f_stock_dividend_gain = n_stock_dividend_gain / 1000
            f_accumulated_inventory = n_accumulated_inventory / 1000
            if f_trading_count.is_integer():
                str_trading_count = format( int( f_trading_count ), "," )
            else:
                str_trading_count = format( f_trading_count, "," )
            if f_stock_dividend_gain.is_integer():
                str_stock_dividend_gain = format( int( f_stock_dividend_gain ), "," )
            else:
                str_stock_dividend_gain = format( f_stock_dividend_gain, "," )
            if f_accumulated_inventory.is_integer():
                str_accumulated_inventory = format( int( f_accumulated_inventory ), "," )
            else:
                str_accumulated_inventory = format( f_accumulated_inventory, "," )


        str_trading_price = format( f_trading_price, "," )
        str_trading_value = format( n_trading_value, "," )
        str_trading_fee = format( n_trading_fee, "," )
        str_trading_tax = format( n_trading_tax, "," )
        str_extra_insurance_fee = format( n_extra_insurance_fee, "," )
        str_per_trading_total_cost = format( n_per_trading_total_cost, "," )
        str_stock_dividend = str_stock_dividend_gain + ' / ' + str( f_stock_dividend_per_share )
        str_cash_dividend = format( n_cash_dividend_gain, "," ) + ' / ' + str( f_cash_dividend_per_share )
        str_accumulated_cost = format( n_accumulated_cost, "," )
        str_average_cost = format( f_average_cost, "," )

        if e_trading_type == TradingType.BUY:
            str_trading_type = "買進"
            str_trading_tax = "N/A"
            str_extra_insurance_fee = "N/A"
            str_stock_dividend = "N/A"
            str_cash_dividend = "N/A"
        elif e_trading_type == TradingType.REGULAR_BUY:
            str_trading_type = "定期定額買進"
            str_trading_tax = "N/A"
            str_extra_insurance_fee = "N/A"
            str_stock_dividend = "N/A"
            str_cash_dividend = "N/A"
        elif e_trading_type == TradingType.SELL:
            n_trading_value = -n_trading_value
            str_trading_value = format( n_trading_value, "," )
            str_per_trading_total_cost = format( n_per_trading_total_cost, "," )
            n_per_trading_total_cost = -n_per_trading_total_cost
            str_trading_type = "賣出"
            str_extra_insurance_fee = "N/A"
            str_stock_dividend = "N/A"
            str_cash_dividend = "N/A"
        elif e_trading_type == TradingType.CAPITAL_INCREASE:
            str_trading_type = "增資"
            str_trading_fee = "N/A"
            str_trading_tax = "N/A"
            str_extra_insurance_fee = "N/A"
            str_stock_dividend = "N/A"
            str_cash_dividend = "N/A"
        elif e_trading_type == TradingType.DIVIDEND:
            str_trading_type = "股利分配"
            str_trading_price = "N/A"
            str_trading_count = "N/A"
            str_trading_value = "N/A"
            str_trading_tax = "N/A"
            str_per_trading_total_cost = "N/A"
        elif e_trading_type == TradingType.CAPITAL_REDUCTION:
            str_trading_type = "減資"
            str_trading_fee = "N/A"
            str_trading_tax = "N/A"
            str_extra_insurance_fee = "N/A"
            str_per_trading_total_cost = "N/A"
            str_stock_dividend = "N/A"
            str_cash_dividend = "N/A"

        list_data = [ str_year,                     #交易年度
                      str_month_date + str_weekday, #交易日期
                      str_trading_type,             #交易種類
                      str_trading_price,            #交易價格
                      str_trading_count,            #交易股數
                      str_trading_value,            #交易金額
                      str_trading_fee,              #手續費
                      str_trading_tax,              #交易稅
                      str_extra_insurance_fee,      #補充保費
                      str_per_trading_total_cost,   #單筆總成本
                      str_stock_dividend,           #總獲得股數 / 每股股票股利
                      str_cash_dividend,            #總獲得現金 / 每股現金股利
                      str_accumulated_cost,         #累計總成本
                      str_accumulated_inventory,    #庫存股數
                      str_average_cost ]            #均價
        return list_data

    def update_button_enable_disable_status( self ): 
        if self.str_picked_stock_number is None or self.ui.qtTabWidget.currentWidget().objectName() not in self.dict_all_account_all_stock_trading_data:
            self.ui.qtAddTradingDataPushButton.setEnabled( False )
            self.ui.qtAddRegularTradingDataPushButton.setEnabled( False )
            self.ui.qtAddDividendDataPushButton.setEnabled( False )
            self.ui.qtAddLimitBuyingDataPushButton.setEnabled( False )
            self.ui.qtAddCapitalReductionDataPushButton.setEnabled( False )
            self.ui.qtExportSelectedStockTradingDataPushButton.setEnabled( False )
        else:
            self.ui.qtAddTradingDataPushButton.setEnabled( True )
            self.ui.qtAddRegularTradingDataPushButton.setEnabled( True )
            self.ui.qtAddLimitBuyingDataPushButton.setEnabled( True )
            self.ui.qtAddCapitalReductionDataPushButton.setEnabled( True )
            self.ui.qtExportSelectedStockTradingDataPushButton.setEnabled( True )
            str_tab_widget_name = self.ui.qtTabWidget.currentWidget().objectName()

            dict_per_account_all_stock_trading_data = self.dict_all_account_all_stock_trading_data[ str_tab_widget_name ]
            if self.str_picked_stock_number in dict_per_account_all_stock_trading_data:
                list_trading_data = dict_per_account_all_stock_trading_data[ self.str_picked_stock_number ]
                b_use_auto_dividend = list_trading_data[ 0 ][ TradingData.USE_AUTO_DIVIDEND_DATA ]
                self.ui.qtAddDividendDataPushButton.setEnabled( not b_use_auto_dividend )
            else:
                self.ui.qtAddDividendDataPushButton.setEnabled( False )

    def open_load_json_file_dialog( self ): 
        # 彈出讀取檔案對話框
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "匯入交易資料",     # 對話框標題
            "",                # 預設路徑
            "JSON (*.json);;"  # 檔案類型過濾
        )
        return file_path

    def open_save_json_file_dialog( self ): 
        # 彈出儲存檔案對話框
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "匯出交易資料",     # 對話框標題
            "",                # 預設路徑
            "JSON (*.json);;"  # 檔案類型過濾
        )
        return file_path
    
    def open_save_excel_file_dialog( self ): 
        # 彈出儲存檔案對話框
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Save File",             # 對話框標題
            "",                      # 預設路徑
            "Excel 活頁簿 (*.xlsx);;All Files (*)"  # 檔案類型過濾
        )
        return file_path

    def show_warning_message_box_with_ok_cancel_button( self, str_title, str_message ): 
        message_box = QMessageBox( self )
        message_box.setIcon( QMessageBox.Warning )  # 設置為警告圖示
        message_box.setWindowTitle( str_title )
        message_box.setText( str_message )

        # 添加自訂按鈕
        button_ok = message_box.addButton("確定", QMessageBox.AcceptRole)
        button_cancel = message_box.addButton("取消", QMessageBox.RejectRole)

        message_box.exec()

        if message_box.clickedButton() == button_ok:
            return True
        elif message_box.clickedButton() == button_cancel:
            return False
        
    def show_error_message_box_with_ok_button( self, str_title, str_message ): 
        message_box = QMessageBox( self )
        message_box.setIcon( QMessageBox.Critical )  # 設置為警告圖示
        message_box.setWindowTitle( str_title )
        message_box.setText( str_message )

        # 添加自訂按鈕
        button_ok = message_box.addButton("確定", QMessageBox.AcceptRole)

        message_box.exec()

        if message_box.clickedButton() == button_ok:
            return True

    def compare_json_files( self, file1, file2 ):
        """比較兩個 JSON 文件內容是否一致"""
        if os.path.exists( file1 ) and os.path.exists( file2 ):
            with open( file1, 'r', encoding='utf-8') as f1, open( file2, 'r', encoding='utf-8') as f2:
                f1.readline()  # 跳過第一行
                data1 = json.load(f1)  # 解析第一個 JSON 文件
                
                f2.readline()  # 跳過第一行
                data2 = json.load(f2)  # 解析第二個 JSON 文件

            # 返回比較結果
            return data1 == data2
        else:
            return False

    def pick_up_stock( self, str_stock_number ):
        self.str_picked_stock_number = str_stock_number
        if str_stock_number is not None:
            str_stock_name = self.dict_all_company_number_to_name_and_type[ str_stock_number ][ 0 ]
            self.ui.qtCurrentSelectCompanyLabel.setText( f"目前選擇公司：{ str_stock_number } { str_stock_name }" )
            self.ui.qtCurrentSelectCompanyLabel.setStyleSheet("color: yellow; ")
        else:
            self.ui.qtCurrentSelectCompanyLabel.setText( "" )

    #region 從網上下載資料
    def check_internet_via_http( self, url="https://www.google.com", timeout=3):
        """
        檢測是否有網路連線（透過 HTTP 請求）
        :param url: 用於測試的 URL
        :param timeout: 超時時間（秒）
        :return: True（有網路連線）或 False（無網路連線）
        """
        try:
            response = requests.get(url, timeout=timeout)
            return response.status_code == 200
        except requests.RequestException:
            return False

    def send_get_request( self, url ):
        retries = 0
        while retries < 3:
            try:
                headers = { 'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36' }
                res = requests.get( url )
                if res.status_code == 200:
                    return res
                else:
                    print("\033[1;31mRequest failed\033[0m")
                    print(f"Status code {res.status_code}. Retrying...")
            except requests.exceptions.Timeout:
                print("\033[1;31mTimeout\033[0m")
            except requests.exceptions.TooManyRedirects:
                print("\033[1;31mTooManyRedirects\033[0m")
            
            retries += 1
            time.sleep(2)  # 等待2秒後重試
    
        raise Exception("Max retries exceeded. Failed to get a successful response.")

    def send_post_request( self, url, payload, max_retries = 3, timeout = 10 ):
        retries = 0
        while retries < max_retries:

            try:
                # 發送 POST 請求
                res = requests.post( url, data = payload, timeout=timeout )
                # 檢查回應的狀態碼，確保是成功的 2xx 系列
                if res.status_code == 200:
                    return res
                else:
                    print("\033[1;31mRequest failed\033[0m")
                    print(f"Status code {res.status_code}. Retrying...")
            except requests.exceptions.Timeout:
                print("\033[1;31mTimeout\033[0m")
            except requests.exceptions.TooManyRedirects:
                print("\033[1;31mTooManyRedirects\033[0m")

            retries += 1
            time.sleep(2)  # 等待2秒後重試
        
        raise Exception("Max retries exceeded. Failed to get a successful response.")

    def download_all_company_stock_number( self, str_date ): 
        dict_company_number_to_name = {}
        b_need_to_download = False
        if os.path.exists( self.stock_number_file_path ):
            with open( self.stock_number_file_path, 'r', encoding='utf-8' ) as f:
                data = f.readlines()
                for i, row in enumerate( data ):
                    if i == 0:
                        if row.strip() != str_date:
                            if self.check_internet_via_http(): #日期不一樣，且又有網路時才重新下載，不然就用舊的
                                b_need_to_download = True
                        else:
                            break
                    else:
                        ele = row.strip().split( ',' )
                        if len( ele ) == 2:
                            b_need_to_download = True
                            break

                        dict_company_number_to_name[ ele[ 0 ] ] = [ ele[ 1 ], ele[ 2 ] ]
        else:
            b_need_to_download = True

        if b_need_to_download:
            tds = []
            # 上市公司股票代碼
            companyNymUrl = "https://isin.twse.com.tw/isin/C_public.jsp?strMode=2"
            try:
                res = self.send_get_request( companyNymUrl )
                soup = BeautifulSoup( res.text, "lxml" )
                tr = soup.findAll( 'tr' )

                total_company_count = 0
                for raw in tr:
                    data = [ td.get_text() for td in raw.findAll("td" )]
                    if len( data ) == 7 and ( data[ 5 ] == 'ESVUFR' or 
                                              data[ 5 ] == 'CEOGBU' or
                                              data[ 5 ] == 'CEOGCU' or
                                              data[ 5 ] == 'CEOGDU' or 
                                              data[ 5 ] == 'CEOGEU' or 
                                              data[ 5 ] == 'CEOGMU' or
                                              data[ 5 ] == 'CEOJBU' or 
                                              data[ 5 ] == 'CEOJEU' or
                                              data[ 5 ] == 'CEOIBU' or
                                              data[ 5 ] == 'CEOIEU' or
                                              data[ 5 ] == 'CEOIRU' ): 
                        b_ETF = False if data[ 5 ] == 'ESVUFR' else True
                        total_company_count += 1
                        if '\u3000' in data[ 0 ]:
                            modified_data = data[ 0 ].split("\u3000")
                            if '-創' in modified_data[ 1 ]:
                                continue
                            modified_data_after_strip = [ modified_data[ 0 ].strip(), modified_data[ 1 ].strip(), b_ETF ]
                            tds.append( modified_data_after_strip )
            except Exception as e:
                pass                

            # 上櫃公司股票代碼
            companyNymUrl = "https://isin.twse.com.tw/isin/C_public.jsp?strMode=4"
            try:
                res = self.send_get_request( companyNymUrl )
                soup = BeautifulSoup( res.text, "lxml" )
                tr = soup.findAll( 'tr' )
                for raw in tr:
                    data = [ td.get_text() for td in raw.findAll("td") ]
                    if len( data ) == 7 and ( data[ 5 ] == 'ESVUFR' or 
                                              data[ 5 ] == 'CEOGBU' or
                                              data[ 5 ] == 'CEOGEU' or 
                                              data[ 5 ] == 'CEOJBU' or 
                                              data[ 5 ] == 'CEOIBU' or
                                              data[ 5 ] == 'CEOIEU' or
                                              data[ 5 ] == 'CEOIRU' ): 
                        b_ETF = False if data[ 5 ] == 'ESVUFR' else True
                        total_company_count += 1
                        if '\u3000' in data[ 0 ]:
                            modified_data = data[ 0 ].split("\u3000")
                            if '-創' in modified_data[ 1 ]:
                                continue
                            modified_data_after_strip = [ modified_data[ 0 ].strip(), modified_data[ 1 ].strip(), b_ETF ]
                            tds.append( modified_data_after_strip )
            except Exception as e:
                pass

            # 興櫃公司股票代碼
            companyNymUrl = "https://isin.twse.com.tw/isin/C_public.jsp?strMode=5"
            try:
                res = self.send_get_request( companyNymUrl )
                soup = BeautifulSoup( res.text, "lxml" )
                tr = soup.findAll( 'tr' )
                for raw in tr:
                    data = [ td.get_text() for td in raw.findAll("td") ]
                    if len( data ) == 7 and data[ 5 ] == 'ESVUFR': 
                        total_company_count += 1
                        if '\u3000' in data[ 0 ]:
                            modified_data = data[ 0 ].split("\u3000")
                            if '-創' in modified_data[ 1 ]:
                                continue
                            modified_data_after_strip = [ modified_data[ 0 ].strip(), modified_data[ 1 ].strip(), False ]
                            tds.append( modified_data_after_strip )
            except Exception as e:
                pass

            if len( tds ) == 0:
                return
            
            # 確保目錄存在，若不存在則遞歸創建
            os.makedirs( os.path.dirname( self.stock_number_file_path ), exist_ok = True )
            with open( self.stock_number_file_path, 'w', encoding='utf-8' ) as f:
                f.write( str_date + '\n' )
                for row in tds:
                    f.write( str( row[ 0 ] ) + ',' + str( row[ 1 ] ) + ',' + str( row[ 2 ] ) + '\n' )
                    dict_company_number_to_name.pop( row[ 0 ], None )
    
                for key_number, value_name_and_etf in dict_company_number_to_name.items():
                    f.write( str( key_number ) + ',' + str( value_name_and_etf[ 0 ] ) + ',' + str( value_name_and_etf[ 1 ] ) + '\n' )

    def load_all_company_stock_number( self ):
        dict_company_number_to_name = {}
        if os.path.exists( self.stock_number_file_path ):
            with open( self.stock_number_file_path, 'r', encoding='utf-8' ) as f:
                data = f.readlines()
                for i, row in enumerate( data ):
                    if i == 0:
                        continue
                    else:
                        ele = row.strip().split( ',' )
                        dict_company_number_to_name[ ele[ 0 ] ] = [ ele[ 1 ], ele[ 2 ] ]
        return dict_company_number_to_name
        
    def download_day_stock_price( self, str_date ):
        b_need_to_download = False
        if os.path.exists( self.stock_price_file_path ):
            with open( self.stock_price_file_path, 'r', encoding='utf-8' ) as f:
                data = f.readlines()
                for i, row in enumerate( data ):
                    if i == 0:
                        if row.strip() != str_date:
                            if self.check_internet_via_http(): #日期不一樣，且又有網路時才重新下載，不然就用舊的
                                b_need_to_download = True
                    else:
                        ele = row.strip().split( ',' )
        else:
            b_need_to_download = True

        if b_need_to_download:
            # 上市公司股價從證交所取得
            # https://www.twse.com.tw/rwd/zh/afterTrading/MI_INDEX?date=20240912&type=ALLBUT0999&response=json&_=1726121461234
            url = 'https://www.twse.com.tw/rwd/zh/afterTrading/MI_INDEX?date=' + str_date + '&type=ALLBUT0999&response=json&_=1726121461234'
            try:
                res = self.send_get_request( url )
                soup = BeautifulSoup( res.content, 'html.parser' )

                all_stock_price = []
                json_str = soup.get_text()
                json_data = json.loads(json_str)
                if 'tables' in json_data:
                    for item in json_data['tables']:
                        if 'title' in item:
                            if '每日收盤行情' in item['title']:
                                for data in item['data']:
                                    #index 0 證券代號    "0050",
                                    #index 1 證券名稱    "元大台灣50",
                                    #index 2 成交股數    "16,337,565",
                                    #index 3 成交筆數    "15,442",
                                    #index 4 成交金額    "2,900,529,886",
                                    #index 5 開盤價      "176.10",
                                    #index 6 最高價      "178.65",
                                    #index 7 最低價      "176.10",
                                    #index 8 收盤價      "178.30",
                                    #index 9 漲跌(+/-)   "<p style= color:red>+<\u002fp>",
                                    #index 10 漲跌價差    "6.45",
                                    #index 11 最後揭示買價 "178.20",
                                    #index 12 最後揭示買量 "5",
                                    #index 13 最後揭示賣價 "178.30",
                                    #index 14 最後揭示賣量 "103",
                                    #index 15 本益比 

                                    list_stock_price = [ data[ 0 ], data[ 1 ], data[ 8 ].replace( ',', '' ) ] 
                                    all_stock_price.append( list_stock_price )
            except Exception as e:
                pass

            # 上櫃公司股價從櫃買中心取得
            # https://www.tpex.org.tw/www/zh-tw/afterTrading/dailyQuotes?date=2024%2F12%2F09&id=&response=html
            formatted_date = f"{str_date[:4]}%2F{str_date[4:6]}%2F{str_date[6:]}"
            url = 'https://www.tpex.org.tw/www/zh-tw/afterTrading/dailyQuotes?date=' + formatted_date + '&id=&response=html'
            try:
                res = self.send_get_request( url )
                
                soup = BeautifulSoup( res.text, "lxml" )
                tr = soup.findAll( 'tr' )
                for raw in tr:
                    if not raw.find( 'th' ):
                        td_elements = raw.findAll( "td" )
                        if len( td_elements ) == 19:
                            # index 0 證券代號	
                            # index 1 證券名稱	
                            # index 2 收盤	
                            # index 3 漲跌	
                            # index 4 開盤
                            # index 5 最高	
                            # index 6 最低
                            # index 7 均價	
                            # index 8 成交股數
                            # index 9 成交金額(元)
                            # index 10 成交筆數	
                            # index 11 最後買價	
                            # index 12 最後買量(千股)	2020/4/29 開始才有這筆資訊
                            # index 13 最後賣價	
                            # index 14 最後賣量(千股)   2020/4/29 開始才有這筆資訊
                            # index 15 發行股數	次日
                            # index 16 參考價	次日
                            # index 17 漲停價	次日
                            # index 18 跌停價
                            str_stock_number = td_elements[ 0 ].get_text().strip()
                            str_stock_name = td_elements[ 1 ].get_text().strip()
                            str_stock_price = td_elements[ 2 ].get_text().strip()
                            list_stock_price = [ str_stock_number, str_stock_name, str_stock_price.replace( ',', '' ) ] 
                            all_stock_price.append( list_stock_price )
            except Exception as e:
                pass    

            if len( all_stock_price ) == 0:
                print( "no data" )
                return 
            
            # 確保目錄存在，若不存在則遞歸創建
            os.makedirs( os.path.dirname( self.stock_price_file_path ), exist_ok = True )
            with open( self.stock_price_file_path, 'w', encoding='utf-8' ) as f:
                f.write( str_date + '\n' )
                for row in all_stock_price:
                    f.write( str( row[ 0 ] ) + ',' + str( row[ 1 ] ) + ',' + str( row[ 2 ] ) + '\n' )
    
    def load_day_stock_price( self ):
        dict_company_number_to_price_info = {}
        if os.path.exists( self.stock_price_file_path ):
            with open( self.stock_price_file_path, 'r', encoding='utf-8' ) as f:
                data = f.readlines()
                for i, row in enumerate( data ):
                    if i == 0:
                        continue
                    else:
                        ele = row.strip().split( ',' )
                        dict_company_number_to_price_info[ ele[ 0 ] ] = ele[ 2 ]
        return dict_company_number_to_price_info

    def process_output_file_path( self, str_output_path, list_file_exist, str_folder_name, str_file_name, n_year, n_season, b_overwrite, b_unit_test ):
        str_season = ''
        if n_season == 1 or n_season == 2 or n_season == 3 or n_season == 4:
            str_season = '_Q' + str( n_season )
        if str_output_path == None:
            if b_unit_test:
                str_output_path = os.path.join( g_data_dir, 'StockInventory', 'UnitTestData', str_file_name + str( n_year ) + str_season + '.txt' )
            else:
                str_output_path = os.path.join( g_data_dir, 'StockInventory', str_folder_name, str_file_name + str( n_year ) + str_season + '.txt' )
        # 確保目錄存在，若不存在則遞歸創建
        os.makedirs( os.path.dirname( str_output_path ), exist_ok = True )
        if b_overwrite or not os.path.exists( str_output_path ):
            list_file_exist[0] = False
        
        return str_output_path

    def download_general_company_all_yearly_dividend_data( self, n_dividend_data_start_year, str_date ):
        print( "\033[32m>>>>>>>>>>>>>>> Start to download all yearly dividend data.\033[0m" )
        current_date = datetime.datetime.today()
        n_current_year = current_date.year

        for n_year in range( n_dividend_data_start_year, n_current_year + 1 ):
            # 假如是西元，轉成民國
            if n_current_year > 1990:
                n_current_year -= 1911
            if n_year > 1990:
                n_year -= 1911

            b_overwrite = False
            file_exist = [ True ]
            str_output_path = self.process_output_file_path( None, file_exist, 'Dividend', 'GeneralCompanyDividend_', n_year, 0, b_overwrite, False )
            if not file_exist[0] or n_year == n_current_year:
                self.download_general_company_yearly_dividend_data( n_year, str_date, str_output_path, True )
                print(f"Finish {n_year} yearly dividend " )

        print( "\033[32m<<<<<<<<<<<<<<< Finish downloading all yearly dividend data.\033[0m" )

    def download_general_company_yearly_dividend_data( self, n_year, str_date, str_output_path, b_overwrite ):
        # 假如是西元，轉成民國
        if n_year > 1990:
            n_year -= 1911

        file_exist = [ True ]
        str_output_path = self.process_output_file_path( str_output_path, file_exist, 'Dividend', 'GeneralCompanyDividend_', n_year, 0, b_overwrite, False )
        if file_exist[0]:
            print("dividend file exists")
            return

        b_need_to_download = False
        if os.path.exists( str_output_path ):
            with open( str_output_path, 'r', encoding='utf-8' ) as f:
                date = f.readline().strip()
                if date != str_date:
                    if self.check_internet_via_http(): #日期不一樣，且又有網路時才重新下載，不然就用舊的
                        b_need_to_download = True
        else:
            b_need_to_download = True

        if b_need_to_download:
            # 請求的 URL
            url = 'https://mops.twse.com.tw/mops/web/ajax_t108sb27'

            # POST 請求的數據
            payload = {
                # 'TYPEK': 'sii' if e_company_type2 == CompanyType2.LISTED else 'otc',
                'encodeURIComponent': '1',
                'firstin': '1',
                'off': '1',
                'step': '1',
                'year': str(n_year)
            }

            all_company_dividend = []
            try:
                for n_type in range( 2 ):
                    if n_type == 0:
                        payload[ 'TYPEK' ] = 'sii'
                    else:
                        payload[ 'TYPEK' ] = 'otc'

                    res = self.send_post_request( url, payload )

                    soup = BeautifulSoup( res.text, "lxml" )
                    tr = soup.findAll( 'tr' )
                    for raw in tr:
                        if not raw.find( 'th' ):
                            data = []
                            td_elements = raw.findAll( "td" )
                            if n_year >= 99 and n_year < 105:
                                if len( td_elements ) == 23:
                                    # [0]公司代號
                                    # [1]公司名稱	
                                    # [2]股利所屬年度	
                                    # [3]權利分派基準日	
                                    # [4]股票股利_盈餘轉增資配股(元/股)
                                    # [5]股票股利_法定盈餘公積、資本公積轉增資配股(元/股)
                                    # [6]股票股利_除權交易日
                                    # X [7]配股總股數(股)
                                    # X [8]配股總金額(元)
                                    # X [9]配股總股數佔盈餘配股總股數之比例(%)
                                    # X [10]員工紅利配股率
                                    # [11]現金股利_盈餘分配之股東現金股利(元/股)
                                    # [12]現金股利_法定盈餘公積、資本公積發放之現金(元/股)
                                    # [13]現金股利_除息交易日
                                    # [14]現金股利_現金股利發放日
                                    # X [15]員工紅利總金額(元)
                                    # [16]現金增資總股數(股)	
                                    # [17]現金增資認股比率(%)	
                                    # [18]現金增資認購價(元/股)		
                                    # X [19]董監酬勞(元)
                                    # [20]公告日期
                                    # [21]公告時間
                                    # [22]普通股每股面額
                                    for index, td in enumerate( td_elements ):
                                        text = td.get_text().strip()
                                        if index == 4 or index == 5 or index == 11 or index == 12 or index == 17 or index == 18:
                                            if text == '\xa0' or text == ''  or text == '-' or text == '--':
                                                data.append( 0 )
                                            else:
                                                number = float( text.replace( ',', '' ) ) 
                                                data.append( number )
                                            if index == 12:
                                                data.append( 0 ) # 為了跟後續的格式有一致性，因為105年之後的表格有「現金股利_特別股配發現金股利(元/股)」，但因為現在沒有，所以直接補0
                                        elif index == 16:
                                            if text == '\xa0' or text == ''  or text == '-' or text == '--':
                                                data.append( 0 )
                                            else:
                                                number = int( text.replace( ',', '' ) ) 
                                                data.append( number )
                                        elif index == 7 or index == 8 or index == 9 or index == 10 or index == 15 or index == 19:
                                            continue
                                        else:
                                            if text == '\xa0' or text == ''  or text == '-' or text == '--':
                                                data.append( '--' )
                                            else:
                                                data.append( text )

                                    all_company_dividend.append( data )
                            elif n_year >= 105 and n_year < 108:
                                if len( td_elements ) == 18:
                                    # [0]公司代號,
                                    # [1]公司名稱,
                                    # [2]股利所屬期間,
                                    # [3]權利分派基準日,
                                    # [4]股票股利_盈餘轉增資配股(元/股),
                                    # [5]股票股利_法定盈餘公積、資本公積轉增資配股(元/股),                        
                                    # [6]股票股利_除權交易日,
                                    # [7]現金股利_盈餘分配之股東現金股利(元/股),
                                    # [8]現金股利_法定盈餘公積、資本公積發放之現金(元/股),
                                    # [9]現金股利_特別股配發現金股利(元/股),                        
                                    # [10]現金股利_除息交易日,
                                    # [11]現金股利_現金股利發放日,
                                    # [12]現金增資總股數(股),
                                    # [13]現金增資認股比率(%),
                                    # [14]現金增資認購價(元/股),                        
                                    # [15]公告日期,
                                    # [16]公告時間,
                                    # [17]普通股每股面額
                                    for index, td in enumerate( td_elements ):
                                        text = td.get_text().strip()
                                        if index == 4 or index == 5 or index == 7 or index == 8 or index == 9 or index == 13 or index == 14:
                                            if text == '\xa0' or text == ''  or text == '-' or text == '--':
                                                data.append( 0 )
                                            else:
                                                number = float( text.replace( ',', '' ) ) 
                                                data.append( number )
                                        elif index == 12:
                                            if text == '\xa0' or text == ''  or text == '-' or text == '--':
                                                data.append( 0 )
                                            else:
                                                number = int( text.replace( ',', '' ) ) 
                                                data.append( number )
                                        else:
                                            if text == '\xa0' or text == ''  or text == '-' or text == '--':
                                                data.append( '--' )
                                            else:
                                                data.append( text )

                                    all_company_dividend.append( data )
                            elif n_year >= 108:
                                if len( td_elements ) == 19:
                                    # [0]公司代號,
                                    # [1]公司名稱,
                                    # [2]股利所屬期間,
                                    # [3]權利分派基準日,
                                    # [4]股票股利_盈餘轉增資配股(元/股),
                                    # [5]股票股利_法定盈餘公積、資本公積轉增資配股(元/股),                        
                                    # [6]股票股利_除權交易日,
                                    # [7]現金股利_盈餘分配之股東現金股利(元/股),
                                    # [8]現金股利_法定盈餘公積、資本公積發放之現金(元/股),
                                    # [9]現金股利_特別股配發現金股利(元/股),                        
                                    # [10]現金股利_除息交易日,
                                    # [11]現金股利_現金股利發放日,
                                    # [12]現金增資總股數(股),
                                    # [13]現金增資認股比率(%),
                                    # [14]現金增資認購價(元/股),                        
                                    # [15]參加分派總股數,
                                    # [16]公告日期,
                                    # [17]公告時間,
                                    # [18]普通股每股面額
                                    for index, td in enumerate( td_elements ):
                                        text = td.get_text().strip()
                                        if index == 4 or index == 5 or index == 7 or index == 8 or index == 9 or index == 13 or index == 14:
                                            if text == '\xa0' or text == ''  or text == '-' or text == '--':
                                                data.append( 0 )
                                            else:
                                                number = float( text.replace( ',', '' ) ) 
                                                data.append( number )
                                        elif index == 12:
                                            if text == '\xa0' or text == ''  or text == '-' or text == '--':
                                                data.append( 0 )
                                            else:
                                                number = int( text.replace( ',', '' ) ) 
                                                data.append( number )
                                        elif index == 15:
                                            continue
                                        else:
                                            if text == '\xa0' or text == ''  or text == '-' or text == '--':
                                                data.append( '--' )
                                            else:
                                                data.append( text )

                                    all_company_dividend.append( data )
            except Exception as e:
                print(f"Final error: {e}")


            if len( all_company_dividend ) == 0:
                print( "no data" )
                return
            with open( str_output_path, 'w', encoding = 'utf-8' ) as f:
                f.write( str_date + '\n' )
                f.write( str( n_year ) + '\n' )
                f.write( '[0]公司代號,[1]公司名稱,[2]股利所屬期間,[3]權利分派基準日,[4]股票股利_盈餘轉增資配股(元/股),[5]股票股利_法定盈餘公積、資本公積轉增資配股(元/股),\
                          [6]股票股利_除權交易日,[7]現金股利_盈餘分配之股東現金股利(元/股),[8]現金股利_法定盈餘公積、資本公積發放之現金(元/股),[9]現金股利_特別股配發現金股利(元/股),\
                          [10]現金股利_除息交易日,[11]現金股利_現金股利發放日,[12]現金增資總股數(股),[13]現金增資認股比率(%),[14]現金增資認購價(元/股),\
                          [15]公告日期,[16]公告時間,[17]普通股每股面額\n' )
                for row in all_company_dividend:
                    b_first = True
                    for ele in row:
                        if b_first:
                            f.write( str( ele ) )
                            b_first = False
                        else:
                            f.write( ',' + str( ele ) )
                    f.write( '\n')  # 用逗號分隔每個元素，並換行

    def load_general_company_all_yearly_dividend_data( self, n_dividend_data_start_year, b_unit_test ):
        current_date = datetime.datetime.today()
        n_current_year = current_date.year
        dict_stock_yearly_dividned = {}
        for n_year in range( n_dividend_data_start_year, n_current_year + 1 ):
            # 假如是西元，轉成民國
            if n_current_year > 1990:
                n_current_year -= 1911
            if n_year > 1990:
                n_year -= 1911
            list_yearly_dividend = self.read_general_company_yearly_dividend_raw_data( n_year, b_unit_test )
            if list_yearly_dividend != None:
                for index, item in enumerate( list_yearly_dividend ):
                    f_stock_dividend_per_share = self.get_value_from_string( item[4] ) + self.get_value_from_string( item[5] )
                    str_stock_dividend_date = item[6]
                    f_cash_dividend_per_share = self.get_value_from_string( item[7] ) + self.get_value_from_string( item[8] )
                    str_cash_dividend_date = item[10]
                    str_cash_dividend_distribute_date = item[11]
                    str_year_month_date = ''

                    if ( f_stock_dividend_per_share != 0 and str_stock_dividend_date != '' ) and \
                       ( f_cash_dividend_per_share != 0 and str_cash_dividend_date != '' ):
                        #同時有現金股利和股票股利
                        
                        if str_stock_dividend_date == str_cash_dividend_date:
                            list_year_month_date = str_stock_dividend_date.split( '/' )
                            str_year = str( int( list_year_month_date[0] ) + 1911 )
                            str_year_month_date = str_year + '-' + list_year_month_date[1] + '-' + list_year_month_date[2]
                        else:
                            #股票股利和現金股利日期不同，理論上不應該出現
                            print("ERRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRR")
                            print("ERRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRR")
                            print("ERRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRR")
                            print("ERRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRR")
                            print("ERRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRR")
                            pass
                    elif ( f_stock_dividend_per_share != 0  and str_stock_dividend_date != '' ):
                        #只有股票股利
                        list_year_month_date = str_stock_dividend_date.split( '/' )
                        str_year = str( int( list_year_month_date[0] ) + 1911 )
                        str_year_month_date = str_year + '-' + list_year_month_date[1] + '-' + list_year_month_date[2]
                        pass
                    elif ( f_cash_dividend_per_share != 0 and str_cash_dividend_date != '' ):
                        #只有現金股利
                        list_year_month_date = str_cash_dividend_date.split( '/' )
                        str_year = str( int( list_year_month_date[0] ) + 1911 )
                        str_year_month_date = str_year + '-' + list_year_month_date[1] + '-' + list_year_month_date[2]
                        pass
                    
                    if str_year_month_date != '':
                        dict_dividend_data = Utility.generate_trading_data( str_year_month_date,        #交易日期
                                                                            TradingType.DIVIDEND,       #交易種類
                                                                            0,                          #交易價格                         
                                                                            0,                          #交易股數
                                                                            TradingFeeType.VARIABLE,    #手續費種類
                                                                            1,                          #手續費折扣
                                                                            0,                          #手續費最低金額
                                                                            0,                          #手續費固定金額                        
                                                                            f_stock_dividend_per_share, #每股股票股利
                                                                            f_cash_dividend_per_share,  #每股現金股利
                                                                            0,                          #每股減資金額
                                                                            False )                     #是否為當沖交易
                        dict_dividend_data[ TradingData.IS_AUTO_DIVIDEND_DATA_NON_SAVE ] = True

                        if item[0] in dict_stock_yearly_dividned:
                            dict_stock_yearly_dividned[ item[0] ].append( dict_dividend_data )
                        else:
                            dict_stock_yearly_dividned[ item[0] ] = [ dict_dividend_data ]

        return dict_stock_yearly_dividned

    def read_general_company_yearly_dividend_raw_data( self, n_year, b_unit_test ):
        if n_year > 1990:
            n_year -= 1911
        file_exist = [ True ]
        file_path = self.process_output_file_path( None, file_exist, 'Dividend', 'GeneralCompanyDividend_', n_year, 0, False, b_unit_test )
        if not os.path.exists( file_path ):
            return None
        
        list_all_company_dividend = []
        with open( file_path, 'r', encoding = 'utf-8' ) as f:
            data = f.readlines()
            for i, row in enumerate( data ):
                if i == 0 or i == 1 or i == 2:#i=0 檔案下載日期, i=1 資料年度, i=2 欄位名稱
                    continue
                else:
                    row = row.strip().split( ',' )
                    list_all_company_dividend.append( row )
            print( "Read " + 'StockDividend_Y' + str( n_year ) )

        return list_all_company_dividend

    def get_value_from_string( self, str_value ):
        if str_value == '' or str_value == '--':
            return 0
        return Decimal( str_value )

    def download_listed_etf_all_yearly_dividend_data( self, n_dividend_data_start_year, str_date ):
        print( "\033[32m>>>>>>>>>>>>>>> Start to download listed etf all yearly dividend data.\033[0m" )
        current_date = datetime.datetime.today()
        n_current_year = current_date.year

        for n_year in range( n_dividend_data_start_year, n_current_year + 1 ):
            # 假如是西元，轉成民國
            if n_current_year > 1990:
                n_current_year -= 1911
            if n_year > 1990:
                n_year -= 1911

            b_overwrite = False
            file_exist = [ True ]
            str_output_path = self.process_output_file_path( None, file_exist, 'Dividend', 'ListedETFDividend_', n_year, 0, b_overwrite, False )
            if not file_exist[0] or n_year == n_current_year:
                self.download_listed_etf_yearly_dividend_data( n_year, str_date, str_output_path, True )
                print(f"Finish Listed etf {n_year} yearly dividend " )

        print( "\033[32m<<<<<<<<<<<<<<< Finish downloading listed etf all yearly dividend data.\033[0m" )

    def download_listed_etf_yearly_dividend_data( self, n_year, str_date, str_output_path, b_overwrite ):
        # 假如是西元，轉成民國
        if n_year > 1990:
            n_year -= 1911

        file_exist = [ True ]
        str_output_path = self.process_output_file_path( str_output_path, file_exist, 'Dividend', 'ListedETFDividend_', n_year, 0, b_overwrite, False )
        if file_exist[0]:
            print("dividend file exists")
            return

        b_need_to_download = False
        if os.path.exists( str_output_path ):
            with open( str_output_path, 'r', encoding='utf-8' ) as f:
                date = f.readline().strip()
                if date != str_date:
                    if self.check_internet_via_http(): #日期不一樣，且又有網路時才重新下載，不然就用舊的
                        b_need_to_download = True
        else:
            b_need_to_download = True

        if n_year < 1990:
            n_year += 1911

        if b_need_to_download:
            url = "https://www.twse.com.tw/rwd/zh/ETF/etfDiv?stkNo=&startDate=" + str( n_year ) + "0101&endDate=" + str( n_year ) + "0101&response=json&_=1734754779791"
            try:
                res = self.send_get_request( url )
                json_value = json.loads( res.text )
                with open( str_output_path, 'w', encoding='utf-8' ) as f:
                    f.write( str_date + '\n' )
                    # 證券代號	
                    # 證券簡稱	
                    # 除息交易日	
                    # 收益分配基準日	
                    # 收益分配發放日	
                    # 收益分配金額 (每1受益權益單位)	
                    # 收益分配標準 (102年度起啟用)	
                    # 公告年度
                    json.dump( json_value[ 'data' ], f, ensure_ascii=False, indent=4 )
            except Exception as e:
                print(f"Final error: {e}")
            pass

    def load_listed_etf_all_yearly_dividend_data( self, n_dividend_data_start_year, b_unit_test ):
        current_date = datetime.datetime.today()
        n_current_year = current_date.year
        dict_stock_yearly_dividned = {}
        for n_year in range( n_dividend_data_start_year, n_current_year + 1 ):
            # 假如是西元，轉成民國
            if n_current_year > 1990:
                n_current_year -= 1911
            if n_year > 1990:
                n_year -= 1911
            list_yearly_dividend = self.read_listed_etf_yearly_dividend_raw_data( n_year, b_unit_test )
            if list_yearly_dividend != None:
                for index, item in enumerate( list_yearly_dividend ):
                    if item[5] != None:
                        try:
                            value = float( item[5].strip() )
                            if value > 0:
                                # 字串是數值且大於 0
                                str_stock_number = item[ 0 ]
                                taiwan_date_str  = item[ 2 ]
                                taiwan_year = int(taiwan_date_str .split('年')[0]) + 1911
                                str_year_month_date = f"{taiwan_year}-{taiwan_date_str.split('年')[1].replace('月', '-').replace('日', '')}"

                                f_cash_dividend_per_share = Decimal( item[ 5 ] )
                                dict_dividend_data = Utility.generate_trading_data( str_year_month_date,       #交易日期
                                                                                    TradingType.DIVIDEND,      #交易種類
                                                                                    0,                         #交易價格                         
                                                                                    0,                         #交易股數
                                                                                    TradingFeeType.VARIABLE,   #手續費種類
                                                                                    1,                         #手續費折扣
                                                                                    0,                         #手續費最低金額
                                                                                    0,                         #手續費固定金額                                   
                                                                                    0,                         #每股股票股利
                                                                                    f_cash_dividend_per_share, #每股現金股利
                                                                                    0,                         #每股減資金額
                                                                                    False )                    #是否為當沖交易
                                
                                dict_dividend_data[ TradingData.IS_AUTO_DIVIDEND_DATA_NON_SAVE ] = True
                                if str_stock_number in dict_stock_yearly_dividned:
                                    dict_stock_yearly_dividned[ str_stock_number ].append( dict_dividend_data )
                                else:
                                    dict_stock_yearly_dividned[ str_stock_number ] = [ dict_dividend_data ]
                                pass
                        except ValueError:
                            # 不是有效的數值
                            pass

        return dict_stock_yearly_dividned

    def read_listed_etf_yearly_dividend_raw_data( self, n_year, b_unit_test ):
        if n_year > 1990:
            n_year -= 1911
        file_exist = [ True ]
        file_path = self.process_output_file_path( None, file_exist, 'Dividend', 'ListedETFDividend_', n_year, 0, False, b_unit_test )
        if not os.path.exists( file_path ):
            return None
        
        with open( file_path, 'r', encoding = 'utf-8' ) as f:
            date = f.readline()
            list_data = json.load( f )

            return list_data
        return []

    def download_OTC_etf_all_yearly_dividend_data( self, n_dividend_data_start_year, str_date ):
        print( "\033[32m>>>>>>>>>>>>>>> Start to download OTC etf all yearly dividend data.\033[0m" )
        current_date = datetime.datetime.today()
        n_current_year = current_date.year

        for n_year in range( n_dividend_data_start_year, n_current_year + 1 ):
            # 假如是西元，轉成民國
            if n_current_year > 1990:
                n_current_year -= 1911
            if n_year > 1990:
                n_year -= 1911

            b_overwrite = False
            file_exist = [ True ]
            str_output_path = self.process_output_file_path( None, file_exist, 'Dividend', 'OTCETFDividend_', n_year, 0, b_overwrite, False )
            if not file_exist[0] or n_year == n_current_year:
                self.download_OTC_etf_yearly_dividend_data( n_year, str_date, str_output_path, True )
                print(f"Finish OTC etf {n_year} yearly dividend " )

        print( "\033[32m<<<<<<<<<<<<<<< Finish downloading OTC etf all yearly dividend data.\033[0m" )

    def download_OTC_etf_yearly_dividend_data( self, n_year, str_date, str_output_path, b_overwrite ):
        # 假如是西元，轉成民國
        if n_year > 1990:
            n_year -= 1911

        file_exist = [ True ]
        str_output_path = self.process_output_file_path( str_output_path, file_exist, 'Dividend', 'OTCETFDividend_', n_year, 0, b_overwrite, False )
        if file_exist[0]:
            print("dividend file exists")
            return

        b_need_to_download = False
        if os.path.exists( str_output_path ):
            with open( str_output_path, 'r', encoding='utf-8' ) as f:
                date = f.readline().strip()
                if date != str_date:
                    if self.check_internet_via_http(): #日期不一樣，且又有網路時才重新下載，不然就用舊的
                        b_need_to_download = True
        else:
            b_need_to_download = True

        if b_need_to_download:
            # 請求的 URL
            url = 'https://www.tpex.org.tw/www/zh-tw/bulletin/exDailyQ'

            if n_year < 1990:
                n_year += 1911
            # POST 請求的數據
            payload = {
                'startDate': str( n_year ) + '/01/01',
                'endDate': str( n_year ) + '/12/31',
                'response': 'json'
            }

            try:
                res = self.send_post_request( url, payload )
                json_value = json.loads( res.text )
                with open( str_output_path, 'w', encoding='utf-8' ) as f:
                    f.write( str_date + '\n' )
                    # 民國104年之前(包含104年)   
                    # 除權息日期                 
                    # 代號	                    
                    # 名稱	                    
                    # 除權息前收盤價             	
                    # 除權息參考價               	
                    # 權值                      	
                    # 息值                      	
                    # 權值＋息值	             
                    # 權/息                   
                    # 漲停價	
                    # 跌停價	
                    # 開始交易基準價	
                    # 減除股利參考價	
                    # 現金股利	
                    # 每仟股無償配股	
                    # 員工紅利轉增資	
                    # 現金增資股數	
                    # 現金增資認購價	
                    # 公開承銷股數	
                    # 員工認購股數	
                    # 原股東認購股數	
                    # 按持股比例仟股認購

                    # 民國105年之後
                    # 除權息日期                 
                    # 代號	                    
                    # 名稱	                    
                    # 除權息前收盤價             	
                    # 除權息參考價               	
                    # 權值                      	
                    # 息值                      	
                    # 權值＋息值	             
                    # 權/息                   
                    # 漲停價	
                    # 跌停價	
                    # 開始交易基準價	
                    # 減除股利參考價	
                    # 現金股利	
                    # 每仟股無償配股	
                    # 缺---------------員工紅利轉增資	
                    # 現金增資股數	
                    # 現金增資認購價	
                    # 公開承銷股數	
                    # 員工認購股數	
                    # 原股東認購股數	
                    # 按持股比例仟股認購
                    json.dump( json_value[ 'tables' ][0]['data'], f, ensure_ascii=False, indent=4 )

            except Exception as e:
                print(f"Final error: {e}")
        pass

    def load_OTC_etf_all_yearly_dividend_data( self, n_dividend_data_start_year, b_unit_test ):
        current_date = datetime.datetime.today()
        n_current_year = current_date.year
        dict_stock_yearly_dividned = {}
        for n_year in range( n_dividend_data_start_year, n_current_year + 1 ):
            # 假如是西元，轉成民國
            if n_current_year > 1990:
                n_current_year -= 1911
            if n_year > 1990:
                n_year -= 1911
            list_yearly_dividend = self.read_OTC_etf_yearly_dividend_raw_data( n_year, b_unit_test )
            if list_yearly_dividend != None:
                for index, item in enumerate( list_yearly_dividend ):
                    if item[13] != None:
                        try:
                            value = float( item[ 13 ].strip() )
                            if value > 0:
                                # 字串是數值且大於 0
                                str_stock_number = item[ 1 ]
                                taiwan_date_str  = item[ 0 ]
                                # taiwan_year = int(taiwan_date_str .split('年')[0]) + 1911
                                taiwan_year = int(taiwan_date_str .split('/')[0]) + 1911
                                str_year_month_date = f"{taiwan_year}-{taiwan_date_str.split('/')[1]}-{taiwan_date_str.split('/')[2]}"

                                str_cash_dividend = item[ 13 ].strip()
                                f_cash_dividend_per_share = Decimal( str_cash_dividend )
                                dict_dividend_data = Utility.generate_trading_data( str_year_month_date,       #交易日期
                                                                                    TradingType.DIVIDEND,      #交易種類
                                                                                    0,                         #交易價格                         
                                                                                    0,                         #交易股數
                                                                                    TradingFeeType.VARIABLE,   #手續費種類
                                                                                    1,                         #手續費折扣
                                                                                    0,                         #手續費最低金額
                                                                                    0,                         #手續費固定金額                                   
                                                                                    0,                         #每股股票股利
                                                                                    f_cash_dividend_per_share, #每股現金股利
                                                                                    0,                         #每股減資金額
                                                                                    False )                    #是否為當沖交易
                                
                                dict_dividend_data[ TradingData.IS_AUTO_DIVIDEND_DATA_NON_SAVE ] = True
                                if str_stock_number in dict_stock_yearly_dividned:
                                    dict_stock_yearly_dividned[ str_stock_number ].append( dict_dividend_data )
                                else:
                                    dict_stock_yearly_dividned[ str_stock_number ] = [ dict_dividend_data ]
                                pass
                        except ValueError:
                            # 不是有效的數值
                            pass

        return dict_stock_yearly_dividned
    
    def read_OTC_etf_yearly_dividend_raw_data( self, n_year, b_unit_test ):
        if n_year > 1990:
            n_year -= 1911
        file_exist = [ True ]
        file_path = self.process_output_file_path( None, file_exist, 'Dividend', 'OTCETFDividend_', n_year, 0, False, b_unit_test )
        if not os.path.exists( file_path ):
            return None
        
        with open( file_path, 'r', encoding = 'utf-8' ) as f:
            date = f.readline()
            list_data = json.load( f )

            return list_data
        return []
    #endregion

def run_app():
    app = QApplication( sys.argv )
    app.setStyle('Fusion')
    window = MainWindow( False, 
                        'TradingData.json',
                        'UISetting.config',
                        'StockNumber.txt',
                        'StockPrice.txt' )

    window.show()
    return app.exec()

if __name__ == "__main__":
    sys.exit( run_app() )

